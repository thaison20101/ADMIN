#!/usr/bin/env python3
"""Match PDF lab results to Medinet TTHC records and import into Khám cận lâm sàng.

Rules:
- Birth year <= 1967 -> M4 list; else M3
- M3: LoaiKham = 5152 (Khám Định Kỳ), fill non-DHDL fields (định kỳ section)
- Do NOT fill khám tuyển (5153 / DHDL_ section)
- Urine Âm tính -> Negative
- Skip / report missing info and already-updated CLS
"""

from __future__ import annotations

import argparse
import json
import os
import re
import time
import unicodedata
import uuid
import urllib.parse
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from medinet_client import API, MedinetClient

OUT = Path("/workspace/build for BIG DATA")

CLS_FIELDS = [
    "CongThucMau_SLHC",
    "XNM_HuyetSacTo",
    "XNM_Hematocrit",
    "XNM_MCV",
    "XNM_MCH",
    "XNM_MCHC",
    "XNM_RDW",
    "CongThucMau_SLBC",
    "SLBC_TrungTinh",
    "SLBC_lympho",
    "SLBC_DonNhan",
    "SLBC_AiToan",
    "SLBC_AiKiem",
    "CongThucMau_SLTC",
    "SinhHoaMau_DuongMau",
    "SinhHoaMau_Ure",
    "SinhHoaMau_Creatinin",
    "SinhHoaMau_ASAT_GOT",
    "SinhHoaMau_ALAT_GPT",
    "NuocTieu_TiTrong",
    "NuocTieu_pH",
    "NuocTieu_BC",
    "NuocTieu_HC",
    "NuocTieu_NiTrit",
    "NuocTieu_Protein",
    "NuocTieu_Duong",
    "NuocTieu_Cetonic",
    "NuocTieu_Bilirubin",
    "NuocTieu_Urobilinogen",
]

# Map định kỳ fields -> DHDL_ only if needed (not used by default)
DHDL_MAP = {f: f"DHDL_{f}" for f in CLS_FIELDS}


def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D")


def norm_name(s: str) -> str:
    return re.sub(r"\s+", " ", strip_accents(s).strip().lower())


def norm_phone(s: Any) -> str:
    return re.sub(r"\D", "", str(s or ""))


def birth_year(val: Any) -> Optional[int]:
    if val is None:
        return None
    if isinstance(val, int):
        return val
    s = str(val)
    m = re.search(r"(19|20)\d{2}", s)
    return int(m.group(0)) if m else None


def style_header(ws):
    fill = PatternFill("solid", fgColor="0F6A5A")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"


def read_pdf_excel(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("sid") or d.get("ho_ten"):
            rows.append(d)
    return rows


def load_web(path: Path) -> List[Dict[str, Any]]:
    return json.loads(path.read_text(encoding="utf-8"))


def index_web(rows: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    by_phone: Dict[str, List[Dict[str, Any]]] = {}
    by_name_year: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        phone = norm_phone(r.get("SDT"))
        name = norm_name(r.get("HoTen") or "")
        year = birth_year(r.get("NgaySinh"))
        if phone:
            by_phone.setdefault(phone, []).append(r)
        if name and year:
            by_name_year.setdefault(f"{name}|{year}", []).append(r)
        elif name:
            by_name_year.setdefault(f"{name}|", []).append(r)
    return {"by_phone": by_phone, "by_name_year": by_name_year}


def match_one(pdf: Dict[str, Any], idx: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], str]:
    phone = norm_phone(pdf.get("sdt"))
    name = norm_name(pdf.get("ho_ten") or pdf.get("name_from_file") or "")
    year = birth_year(pdf.get("nam_sinh"))

    cands: List[Dict[str, Any]] = []
    if phone and phone in idx["by_phone"]:
        cands = idx["by_phone"][phone]
        # refine by year/name if multiple
        if year:
            c2 = [c for c in cands if birth_year(c.get("NgaySinh")) == year]
            if c2:
                cands = c2
        if name:
            c2 = [c for c in cands if norm_name(c.get("HoTen") or "") == name]
            if c2:
                cands = c2
        if len(cands) == 1:
            return cands[0], "phone"
        if len(cands) > 1:
            return None, f"AMBIGUOUS_PHONE:{len(cands)}"

    key = f"{name}|{year or ''}"
    cands = idx["by_name_year"].get(key) or []
    if not cands and year:
        # try name only then filter year loosely
        cands = [c for c in idx["by_name_year"].get(f"{name}|", []) if birth_year(c.get("NgaySinh")) == year]
    if len(cands) == 1:
        return cands[0], "name_year"
    if len(cands) > 1:
        return None, f"AMBIGUOUS_NAME:{len(cands)}"
    return None, "NOT_FOUND"


def build_cls_payload(pdf: Dict[str, Any]) -> Dict[str, Any]:
    fd: Dict[str, Any] = {
        "LoaiKham": 5152,  # Khám Định Kỳ
        "__label_action_code": "CREATE_PHIEU,SAVE,BACK",
    }
    for f in CLS_FIELDS:
        val = pdf.get(f)
        if val is None or str(val).strip() == "":
            fd[f] = None
            continue
        s = str(val).strip()
        if s.lower() in ("negative", "neg"):
            fd[f] = "Negative"
            continue
        # keep qualitative +++, keep numbers as numbers when possible
        try:
            if re.fullmatch(r"-?\d+(?:\.\d+)?", s):
                num = float(s)
                fd[f] = int(num) if num.is_integer() else num
            else:
                fd[f] = s
        except Exception:
            fd[f] = s
    return fd


def get_cls_formdata(client: MedinetClient, tthc_id: int, maukham: str = "mauphieudk") -> Dict[str, Any]:
    url_page = urllib.parse.quote(
        f"/nav_group/kskdk_thongtinkham/app/main/dynamicform/viewer/KSKDK_Phieu_CanLamSang?TTHCId={tthc_id}&maukham={maukham}"
    )
    params = [
        {"Varible": "TTHCId", "Value": str(tthc_id)},
        {"Varible": "maukham", "Value": maukham},
    ]
    qs = (
        f"form_id=1000250&SessionSiteId={client.site_id}&record_id={tthc_id}"
        f"&UrlPage={url_page}&ispopup=false&istab=false"
    )
    res = client.call("POST", f"{API}/api/services/app/FormViewer/FormViewerDataByRecord?{qs}", params)
    data = (res.get("result") or {}).get("data") or {}
    fd = data.get("formData") or [{}]
    return fd[0] if isinstance(fd, list) and fd else {}


def cls_already_filled(fd: Dict[str, Any]) -> bool:
    # Count meaningful lab values in định kỳ section
    filled = 0
    for k in CLS_FIELDS:
        v = fd.get(k)
        if v not in (None, "", [], {}):
            filled += 1
    return filled >= 3


NITRIT_AM_TINH = 5120
NITRIT_DUONG_TINH = 5119

# NumberBox: send float/int with dot. TextBox urine: comma decimal or "Negative".
NUMBERBOX_FIELDS = {
    "CongThucMau_SLHC",
    "XNM_HuyetSacTo",
    "XNM_Hematocrit",
    "XNM_MCV",
    "XNM_MCH",
    "XNM_MCHC",
    "XNM_RDW",
    "CongThucMau_SLBC",
    "SLBC_TrungTinh",
    "SLBC_lympho",
    "SLBC_DonNhan",
    "SLBC_AiToan",
    "SLBC_AiKiem",
    "CongThucMau_SLTC",
    "SinhHoaMau_DuongMau",
    "SinhHoaMau_Ure",
    "SinhHoaMau_Creatinin",
    "SinhHoaMau_ASAT_GOT",
    "SinhHoaMau_ALAT_GPT",
}
TEXTBOX_FIELDS = {
    "NuocTieu_TiTrong",
    "NuocTieu_pH",
    "NuocTieu_BC",
    "NuocTieu_HC",
    "NuocTieu_Protein",
    "NuocTieu_Duong",
    "NuocTieu_Cetonic",
    "NuocTieu_Bilirubin",
    "NuocTieu_Urobilinogen",
}


def to_comma_decimal(val: Any) -> str:
    if isinstance(val, bool):
        raise ValueError("bool")
    if isinstance(val, int):
        return str(val)
    num = float(val) if not isinstance(val, float) else val
    if num.is_integer():
        return str(int(num))
    s = f"{num:.6f}".rstrip("0").rstrip(".")
    return s.replace(".", ",")


def to_numberbox(val: Any) -> Any:
    num = float(val) if not isinstance(val, (int, float)) else val
    if isinstance(num, float) and num.is_integer():
        return int(num)
    return num


def sanitize_cls_payload(form_data: Dict[str, Any]) -> Dict[str, Any]:
    """Build payload accepted by FormToDataBaseUpdate.

    - LoaiKham = 5152 (Khám Định Kỳ)
    - NumberBox: JSON numbers
    - TextBox urine: comma decimals or 'Negative'
    - Nitrit radio: 5120 (Âm tính) / 5119 (Dương tính)
    """
    out: Dict[str, Any] = {"LoaiKham": 5152}
    for k, v in form_data.items():
        if k in ("LoaiKham", "__label_action_code"):
            continue
        if v is None or str(v).strip() == "":
            continue
        if k == "NuocTieu_NiTrit":
            s = str(v).strip().lower()
            if s in ("negative", "âm tính", "am tinh", "5120"):
                out[k] = NITRIT_AM_TINH
            elif s in ("positive", "dương tính", "duong tinh", "5119"):
                out[k] = NITRIT_DUONG_TINH
            continue

        s = str(v).strip()
        if s.lower() == "negative":
            if k in TEXTBOX_FIELDS:
                out[k] = "Negative"
            continue

        num = None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            num = float(v)
        elif re.fullmatch(r"-?\d+(?:[.,]\d+)?", s):
            num = float(s.replace(",", "."))
        else:
            continue  # skip +++, ranges

        if k == "NuocTieu_pH" and (num < 3 or num > 9):
            continue
        if k == "NuocTieu_TiTrong" and (num < 1.0 or num > 1.05):
            continue

        if k in NUMBERBOX_FIELDS:
            out[k] = to_numberbox(num)
        elif k in TEXTBOX_FIELDS:
            out[k] = to_comma_decimal(num)
        else:
            out[k] = to_numberbox(num)
    return out


def submit_cls(client: MedinetClient, tthc_id: int, form_data: Dict[str, Any], maukham: str = "mauphieudk") -> Dict[str, Any]:
    url_page = urllib.parse.quote(
        f"/nav_group/kskdk_thongtinkham/app/main/dynamicform/viewer/KSKDK_Phieu_CanLamSang?TTHCId={tthc_id}&maukham={maukham}"
    )
    payload = sanitize_cls_payload(form_data)

    # 1) Open/create form context
    guid = str(uuid.uuid4())
    params = [
        {"Varible": "TTHCId", "Value": str(tthc_id)},
        {"Varible": "maukham", "Value": maukham},
        {"Varible": "KSKDK_Phieu_CanLamSang_guid", "Value": guid},
        {"Varible": "formdata_parent", "Value": "null"},
        {
            "Varible": "DataFlowData",
            "Value": json.dumps(
                {"LoaiKham": 5152, "__label_action_code": "CREATE_PHIEU,SAVE,BACK"},
                ensure_ascii=False,
            ),
        },
    ]
    qs = f"form_id=1000250&SessionSiteId={client.site_id}&UrlPage={url_page}&ispopup=false&istab=false"
    client.call("POST", f"{API}/api/services/app/FormViewer/FormViewerData?{qs}", params)

    # 2) Persist values
    url = (
        f"{API}/api/services/app/FormViewer/FormToDataBaseUpdate?"
        f"form_id=1000250&record_id={tthc_id}&SessionSiteId={client.site_id}"
        f"&UrlPage={url_page}&ispopup=false&istab=false"
    )
    res = client.call("POST", url, payload)
    saved = get_cls_formdata(client, tthc_id, maukham=maukham)
    ok = cls_already_filled(saved)
    # If bulk failed, retry field-by-field
    if not ok or not (res.get("result") or {}).get("isSucceeded"):
        ok_fields = 0
        last = res
        for k, v in payload.items():
            try:
                last = client.call("POST", url, {k: v})
                if (last.get("result") or {}).get("isSucceeded"):
                    ok_fields += 1
            except Exception:
                continue
        saved = get_cls_formdata(client, tthc_id, maukham=maukham)
        ok = cls_already_filled(saved)
        res = last
        res["_ok_fields"] = ok_fields
    res["_verified_filled"] = ok
    res["_saved_preview"] = {
        k: saved.get(k)
        for k in ("CongThucMau_SLBC", "SLBC_TrungTinh", "XNM_HuyetSacTo", "NuocTieu_Protein", "NuocTieu_NiTrit", "LoaiKham")
    }
    return res


def write_issues_excel(rows: List[Dict[str, Any]], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "CanKiemTra"
    cols = [
        "ly_do",
        "sid",
        "ho_ten",
        "nam_sinh",
        "sdt",
        "nhom",
        "match_status",
        "tthc_id",
        "web_hoten",
        "web_sdt",
        "web_ngaysinh",
        "ghi_chu",
        "file",
    ]
    ws.append(cols)
    style_header(ws)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    for i in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    wb.save(path)


def write_import_log(rows: List[Dict[str, Any]], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "ImportLog"
    cols = [
        "status",
        "sid",
        "ho_ten",
        "nam_sinh",
        "nhom",
        "tthc_id",
        "match_by",
        "message",
        "file",
    ]
    ws.append(cols)
    style_header(ws)
    for r in rows:
        ws.append([r.get(c) for c in cols])
    wb.save(path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--user", default="pkdkthuankieu")
    p.add_argument("--password", default=os.environ.get("MEDINET_PASSWORD",""))
    p.add_argument("--pdf-excel", default=str(OUT / "CLS_ket_qua_tu_PDF_de_kiem_tra.xlsx"))
    p.add_argument("--m3-json", default=str(OUT / "web_list_M3.json"))
    p.add_argument("--m4-json", default=str(OUT / "web_list_M4.json"))
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--skip-filled", action="store_true", default=True)
    p.add_argument("--only-nhom", default="", help="M3 or M4")
    p.add_argument("--delay", type=float, default=0.35)
    args = p.parse_args()

    pdf_rows = read_pdf_excel(Path(args.pdf_excel))
    if args.only_nhom:
        pdf_rows = [r for r in pdf_rows if r.get("nhom") == args.only_nhom]
    if args.limit:
        pdf_rows = pdf_rows[: args.limit]

    m3 = load_web(Path(args.m3_json)) if Path(args.m3_json).exists() else []
    m4 = load_web(Path(args.m4_json)) if Path(args.m4_json).exists() else []
    idx_m3 = index_web(m3)
    idx_m4 = index_web(m4)
    print(f"PDF={len(pdf_rows)} webM3={len(m3)} webM4={len(m4)}")

    client = MedinetClient(args.user, args.password)

    issues: List[Dict[str, Any]] = []
    logs: List[Dict[str, Any]] = []

    for i, pdf in enumerate(pdf_rows, 1):
        nhom = pdf.get("nhom") or ("M4" if (pdf.get("nam_sinh") or 9999) <= 1967 else "M3")
        idx = idx_m4 if nhom == "M4" else idx_m3
        web, how = match_one(pdf, idx)
        base = {
            "sid": pdf.get("sid"),
            "ho_ten": pdf.get("ho_ten"),
            "nam_sinh": pdf.get("nam_sinh"),
            "sdt": pdf.get("sdt"),
            "nhom": nhom,
            "file": pdf.get("file"),
        }
        if not web:
            issues.append({**base, "ly_do": "THIEU_THONG_TIN_HOAC_KHONG_TIM_THAY", "match_status": how, "ghi_chu": "Không khớp bản ghi trên web"})
            logs.append({**base, "status": "SKIP_NOT_FOUND", "match_by": how, "message": how, "tthc_id": None})
            continue

        tthc_id = web.get("Id") or web.get("phieukhamId")
        base.update(
            {
                "tthc_id": tthc_id,
                "web_hoten": web.get("HoTen"),
                "web_sdt": web.get("SDT"),
                "web_ngaysinh": web.get("NgaySinh"),
                "match_status": how,
            }
        )

        # Check existing CLS
        try:
            existing = get_cls_formdata(client, int(tthc_id), maukham="mauphieudk")
        except Exception as e:
            issues.append({**base, "ly_do": "LOI_DOC_CLS", "ghi_chu": str(e)[:200]})
            logs.append({**base, "status": "LOI_DOC", "match_by": how, "message": str(e)[:200]})
            continue

        if args.skip_filled and cls_already_filled(existing):
            issues.append({**base, "ly_do": "DA_CO_KET_QUA_CLS", "ghi_chu": "Đã có dữ liệu cận lâm sàng trên web"})
            logs.append({**base, "status": "SKIP_FILLED", "match_by": how, "message": "already filled", "tthc_id": tthc_id})
            continue

        if pdf.get("missing_core"):
            issues.append({**base, "ly_do": "THIEU_KET_QUA_PDF", "ghi_chu": pdf.get("missing_core")})
            logs.append({**base, "status": "SKIP_PDF_MISSING", "match_by": how, "message": pdf.get("missing_core"), "tthc_id": tthc_id})
            continue

        payload = build_cls_payload(pdf)
        if args.dry_run:
            logs.append({**base, "status": "DRY_RUN", "match_by": how, "message": f"keys={len(payload)}", "tthc_id": tthc_id})
            print(f"[{i}/{len(pdf_rows)}] DRY-RUN {pdf.get('ho_ten')} -> {tthc_id}")
            continue

        try:
            res = submit_cls(client, int(tthc_id), payload, maukham="mauphieudk")
            result = res.get("result") or {}
            ok = bool(res.get("_verified_filled"))
            msg = result.get("message") or ("OK" if ok else str(res)[:200])
            preview = res.get("_saved_preview") or {}
            status = "THANH_CONG" if ok else "LOI"
            if not ok:
                issues.append({**base, "ly_do": "LOI_IMPORT", "ghi_chu": f"{msg} | {preview}"})
            logs.append(
                {
                    **base,
                    "status": status,
                    "match_by": how,
                    "message": f"{msg} | {preview}",
                    "tthc_id": tthc_id,
                }
            )
            print(f"[{i}/{len(pdf_rows)}] {status} {pdf.get('ho_ten')} -> {tthc_id} {preview}")
        except Exception as e:
            issues.append({**base, "ly_do": "LOI_IMPORT", "ghi_chu": str(e)[:200]})
            logs.append({**base, "status": "LOI", "match_by": how, "message": str(e)[:200], "tthc_id": tthc_id})
            print(f"[{i}/{len(pdf_rows)}] LOI {pdf.get('ho_ten')}: {e}")

        if args.delay:
            time.sleep(args.delay)

        # periodic save
        if i % 25 == 0:
            write_issues_excel(issues, OUT / "CLS_can_kiem_tra_lai.xlsx")
            write_import_log(logs, OUT / "CLS_import_log.xlsx")
            (OUT / "CLS_import_log.jsonl").write_text(
                "\n".join(json.dumps(x, ensure_ascii=False) for x in logs), encoding="utf-8"
            )

    write_issues_excel(issues, OUT / "CLS_can_kiem_tra_lai.xlsx")
    write_import_log(logs, OUT / "CLS_import_log.xlsx")
    (OUT / "CLS_import_log.jsonl").write_text(
        "\n".join(json.dumps(x, ensure_ascii=False) for x in logs), encoding="utf-8"
    )
    summary = {
        "total": len(pdf_rows),
        "ok": sum(1 for x in logs if x["status"] == "THANH_CONG"),
        "dry": sum(1 for x in logs if x["status"] == "DRY_RUN"),
        "skip_filled": sum(1 for x in logs if x["status"] == "SKIP_FILLED"),
        "not_found": sum(1 for x in logs if x["status"] == "SKIP_NOT_FOUND"),
        "errors": sum(1 for x in logs if x["status"] in ("LOI", "LOI_DOC")),
        "issues": len(issues),
    }
    (OUT / "CLS_import_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print("SUMMARY", summary)


if __name__ == "__main__":
    main()
