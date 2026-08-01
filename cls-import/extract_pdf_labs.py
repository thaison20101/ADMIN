#!/usr/bin/env python3
"""Extract lab results from Thuận Kiều PDF reports into normalized Excel for web import."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

try:
    import fitz  # PyMuPDF
except ImportError:  # pragma: no cover
    fitz = None


OUT_DIR_DEFAULT = Path("/workspace/build for BIG DATA")


def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D")


def norm(s: str) -> str:
    return re.sub(r"\s+", " ", strip_accents(s or "").strip().lower())


def parse_filename(path: Path) -> Dict[str, Any]:
    # SID - NAME - YEAR - SEX.pdf
    stem = path.stem
    m = re.match(r"^(\d{6}-\d+)\s*-\s*(.+?)\s*-\s*(\d{4})\s*-\s*([MF])\s*$", stem)
    if not m:
        return {"sid_from_file": None, "name_from_file": stem, "year_from_file": None, "sex_from_file": None}
    return {
        "sid_from_file": m.group(1),
        "name_from_file": m.group(2).strip(),
        "year_from_file": int(m.group(3)),
        "sex_from_file": "Nam" if m.group(4) == "M" else "Nữ",
    }


def pdf_text(path: Path) -> str:
    if fitz is not None:
        doc = fitz.open(path)
        parts = []
        for page in doc:
            parts.append(page.get_text("text"))
        doc.close()
        return "\n".join(parts)
    import subprocess

    return subprocess.check_output(["pdftotext", "-layout", str(path), "-"], text=True, errors="ignore")


# PDF analyte aliases -> canonical key
ANALYTE_ALIASES = {
    "leukocytes(wbc)": "wbc",
    "leukocytes": "wbc",
    "wbc": "wbc",
    "neutrophils #": "neutrophils_abs",
    "neutrophils#": "neutrophils_abs",
    "neutrophil #": "neutrophils_abs",
    "neutrophils": "neutrophils_pct",
    "lymphocytes #": "lymphocytes_abs",
    "lymphocytes#": "lymphocytes_abs",
    "lymphocytes": "lymphocytes_pct",
    "monocytes #": "monocytes_abs",
    "monocytes#": "monocytes_abs",
    "monocytes": "monocytes_pct",
    "eosinophils #": "eosinophils_abs",
    "eosinophils#": "eosinophils_abs",
    "eosinophils": "eosinophils_pct",
    "basophils #": "basophils_abs",
    "basophils#": "basophils_abs",
    "basophils": "basophils_pct",
    "erythrocytes(rbc)": "rbc",
    "erythrocytes": "rbc",
    "rbc": "rbc",
    "hemoglobin (hgb)": "hgb",
    "hemoglobin (hb)": "hgb",
    "hemoglobin": "hgb",
    "hgb": "hgb",
    "hb": "hgb",
    "hematocrit (hct)": "hct",
    "hematocrit": "hct",
    "hct": "hct",
    "mcv": "mcv",
    "mch": "mch",
    "mchc": "mchc",
    "rdw": "rdw",
    "platelets (plt)": "plt",
    "platelets": "plt",
    "plt": "plt",
    "mpv": "mpv",
    "glucose": "glucose",
    "glucose ngau nhien": "glucose",
    "urea": "urea",
    "ure": "urea",
    "creatinine": "creatinine",
    "creatinin": "creatinine",
    "ast(sgot)": "ast",
    "ast": "ast",
    "sgot": "ast",
    "alt(sgpt)": "alt",
    "alt": "alt",
    "sgpt": "alt",
    "urobilinogen": "urobilinogen",
    "ketone": "ketone",
    "bilirubin": "bilirubin",
    "dam": "protein",
    "protein": "protein",
    "nitrite": "nitrite",
    "nitrit": "nitrite",
    "hong cau": "urine_rbc",
    "ti trong": "sg",
    "ty trong": "sg",
    "bach cau": "urine_wbc",
}

# Only matched inside urine section, with stricter line-start patterns
URINE_SPECIAL = [
    ("ph", re.compile(r"(?im)^\s*pH\s+(?P<value>\d+(?:[.,]\d+)?)")),
    ("urine_blood", re.compile(r"(?im)^\s*Máu\s+(?P<value>Âm tính|Am tinh|\+\+\+|\+\+|\+|\d+(?:[.,]\d+)?)")),
]


HEADER_RE = re.compile(
    r"Ho ten:\s*(?P<hoten>.+?)\s+Nam sinh\s+(?P<namsinh>\d{4})\s+Gioi tinh:\s*(?P<gioitinh>\S+)",
    re.I,
)
SID_RE = re.compile(r"SID:\s*([0-9]{6}-[0-9]+)", re.I)
PHONE_RE = re.compile(r"So DT:\s*([0-9\s]+)", re.I)
DATE_RE = re.compile(r"Ngay nhan mau:\s*([0-9/:\s]+)", re.I)
RESULT_DATE_RE = re.compile(r"Ngay co ket qua:\s*([0-9/:\s]+)", re.I)


def extract_header(text: str, meta: Dict[str, Any]) -> Dict[str, Any]:
    ntext = strip_accents(text)
    out = dict(meta)
    m = HEADER_RE.search(ntext)
    if m:
        out["ho_ten"] = re.sub(r"\s+", " ", m.group("hoten")).strip()
        out["nam_sinh"] = int(m.group("namsinh"))
        gt = m.group("gioitinh").strip().lower()
        out["gioi_tinh"] = "Nữ" if gt.startswith("nu") else "Nam"
    else:
        out["ho_ten"] = meta.get("name_from_file")
        out["nam_sinh"] = meta.get("year_from_file")
        out["gioi_tinh"] = meta.get("sex_from_file")
    m = SID_RE.search(ntext)
    out["sid"] = m.group(1) if m else meta.get("sid_from_file")
    m = PHONE_RE.search(ntext)
    if m:
        out["sdt"] = re.sub(r"\D", "", m.group(1))
    m = DATE_RE.search(ntext)
    if m:
        out["ngay_nhan_mau"] = m.group(1).strip()
    m = RESULT_DATE_RE.search(ntext)
    if m:
        out["ngay_co_ket_qua"] = m.group(1).strip()
    year = out.get("nam_sinh")
    if year:
        out["nhom"] = "M4" if int(year) <= 1967 else "M3"
        out["maukham"] = "MauPhieuKSKOT" if int(year) <= 1967 else "mauphieudk"
    return out


VALUE_RE = re.compile(
    r"(?P<name>[A-Za-zÀ-ỹ#\(\)/\.\s]+?)\s{2,}(?P<value>(?:Am tinh|\+\+\+|\+\+|\+|Âm tính|âm tính|\(?\d[\d\.,\s\-]*\)?|\d[\d\.,]*))\s*(?P<unit>[A-Za-zµμ/%\^\d\.\-]*)?",
    re.I,
)


# More reliable line-oriented extractor using known analyte names
KNOWN_PATTERNS: List[Tuple[str, re.Pattern]] = []
for alias, key in ANALYTE_ALIASES.items():
    # allow optional spaces around # and parentheses
    alias_re = re.escape(alias).replace(r"\#", r"\s*#\s*").replace(r"\ ", r"\s+")
    pat = re.compile(
        rf"(?im)^\s*.*?\b{alias_re}\b\s*(?P<value>Âm tính|Am tinh|\(\+\+\+\)|\(\+\+\)|\(\+\)|\+\+\+|\+\+|\+|\(?\d+(?:[.,]\d+)?(?:\s*-\s*\d+(?:[.,]\d+)?)?\)?|\d+(?:[.,]\d+)?)\s*(?P<unit>[A-Za-zµμ/%\^\d\.\-]*)?",
    )
    KNOWN_PATTERNS.append((key, pat))


def parse_number(val: str) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip().replace(",", ".")
    s = s.strip("()")
    if re.fullmatch(r"\d+(?:\.\d+)?\s*-\s*\d+(?:\.\d+)?", s):
        # range like 10 - 12 -> keep as text later
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_negative(val: str) -> bool:
    n = norm(val)
    return n in {"am tinh", "negative", "neg", "-"}


def convert_units(raw: Dict[str, Dict[str, Any]]) -> Tuple[Dict[str, Any], List[str]]:
    """Convert PDF units to web units. Returns (web_values, notes)."""
    notes: List[str] = []
    web: Dict[str, Any] = {}

    def get(key: str) -> Tuple[Optional[Any], Optional[str], Optional[str]]:
        item = raw.get(key) or {}
        return item.get("value"), item.get("unit"), item.get("raw")

    # RBC T/L or 10^12/L
    v, u, r = get("rbc")
    if v is not None:
        web["CongThucMau_SLHC"] = parse_number(v) if not isinstance(v, (int, float)) else v
        web["CongThucMau_SLHC_unit_src"] = u or ""

    # HGB -> g/L
    v, u, r = get("hgb")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        un = norm(u or "")
        if "g/dl" in un or un == "g/dl":
            web["XNM_HuyetSacTo"] = round(num * 10, 2)
            notes.append(f"HGB {num} g/dL -> {web['XNM_HuyetSacTo']} g/L")
        else:
            # already g/L (or blank but value looks like g/L if > 30)
            if num < 30 and "g/l" not in un:
                web["XNM_HuyetSacTo"] = round(num * 10, 2)
                notes.append(f"HGB {num} assumed g/dL -> {web['XNM_HuyetSacTo']} g/L")
            else:
                web["XNM_HuyetSacTo"] = num

    # HCT -> L/L
    v, u, r = get("hct")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        un = norm(u or "")
        if "%" in un or num > 1.5:
            web["XNM_Hematocrit"] = round(num / 100.0, 4)
            notes.append(f"HCT {num}% -> {web['XNM_Hematocrit']} L/L")
        else:
            web["XNM_Hematocrit"] = num

    for src, dst in [
        ("mcv", "XNM_MCV"),
        ("mch", "XNM_MCH"),
        ("rdw", "XNM_RDW"),
    ]:
        v, u, _ = get(src)
        num = parse_number(str(v)) if v is not None else None
        if num is not None:
            web[dst] = num

    # MCHC -> g/L
    v, u, _ = get("mchc")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        un = norm(u or "")
        if "g/dl" in un or (num < 50 and "g/l" not in un):
            web["XNM_MCHC"] = round(num * 10, 2)
            notes.append(f"MCHC {num} -> {web['XNM_MCHC']} g/L")
        else:
            web["XNM_MCHC"] = num

    # WBC and absolute diffs: G/L == 10^9/L
    for src, dst in [
        ("wbc", "CongThucMau_SLBC"),
        ("neutrophils_abs", "SLBC_TrungTinh"),
        ("lymphocytes_abs", "SLBC_lympho"),
        ("monocytes_abs", "SLBC_DonNhan"),
        ("eosinophils_abs", "SLBC_AiToan"),
        ("basophils_abs", "SLBC_AiKiem"),
        ("plt", "CongThucMau_SLTC"),
    ]:
        v, u, _ = get(src)
        num = parse_number(str(v)) if v is not None else None
        if num is not None:
            web[dst] = num
            # warn if only percent was present for diffs
    for pct_key, abs_dst in [
        ("neutrophils_pct", "SLBC_TrungTinh"),
        ("lymphocytes_pct", "SLBC_lympho"),
        ("monocytes_pct", "SLBC_DonNhan"),
        ("eosinophils_pct", "SLBC_AiToan"),
        ("basophils_pct", "SLBC_AiKiem"),
    ]:
        if abs_dst not in web:
            v, u, _ = get(pct_key)
            if v is not None:
                notes.append(f"WARNING: only % for {pct_key}={v}{u or ''}; need # absolute for {abs_dst}")

    # Glucose blood -> mmol/L
    v, u, _ = get("glucose")
    num = parse_number(str(v)) if v is not None and not is_negative(str(v)) else None
    if num is not None:
        un = norm(u or "")
        if "mg" in un or (num > 25 and "mmol" not in un):
            web["SinhHoaMau_DuongMau"] = round(num / 18.0, 2)
            notes.append(f"Glucose {num} {u or 'mg/dL?'} -> {web['SinhHoaMau_DuongMau']} mmol/L")
        else:
            web["SinhHoaMau_DuongMau"] = num

    # Urea -> mmol/L
    v, u, _ = get("urea")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        un = norm(u or "")
        if "mg" in un or (num > 15 and "mmol" not in un):
            web["SinhHoaMau_Ure"] = round(num / 6.0, 2)  # BUN/urea mg/dL -> mmol/L approx /6
            notes.append(f"Urea {num} {u or 'mg%?'} -> {web['SinhHoaMau_Ure']} mmol/L (÷6)")
        else:
            web["SinhHoaMau_Ure"] = num

    # Creatinine -> µmol/L
    v, u, _ = get("creatinine")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        un = norm(u or "")
        if "mg" in un:
            web["SinhHoaMau_Creatinin"] = round(num * 88.4, 2)
            notes.append(f"Creatinine {num} mg/dL -> {web['SinhHoaMau_Creatinin']} µmol/L")
        elif "mcmol" in un or "umol" in un or "µmol" in (u or "") or "μmol" in (u or ""):
            web["SinhHoaMau_Creatinin"] = num
        else:
            # heuristic: < 20 likely mg/dL
            if num < 20:
                web["SinhHoaMau_Creatinin"] = round(num * 88.4, 2)
                notes.append(f"Creatinine {num} assumed mg% -> {web['SinhHoaMau_Creatinin']} µmol/L")
            else:
                web["SinhHoaMau_Creatinin"] = num

    for src, dst in [("ast", "SinhHoaMau_ASAT_GOT"), ("alt", "SinhHoaMau_ALAT_GPT")]:
        v, u, _ = get(src)
        num = parse_number(str(v)) if v is not None else None
        if num is not None:
            web[dst] = num

    # Urine fields
    v, u, r = get("sg")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        web["NuocTieu_TiTrong"] = num
    v, u, r = get("ph")
    num = parse_number(str(v)) if v is not None else None
    if num is not None:
        web["NuocTieu_pH"] = num

    def urine_text(key: str, dst: str, negative_as="Negative"):
        v, u, r = get(key)
        if v is None:
            return
        if is_negative(str(v)):
            web[dst] = negative_as
            notes.append(f"{key}: Âm tính -> {negative_as}")
        else:
            # keep qualitative or numeric as-is (string)
            web[dst] = str(v).strip()
            if u:
                web[dst + "_unit_src"] = u

    # Nitrit neg/pos
    v, u, r = get("nitrite")
    if v is not None:
        if is_negative(str(v)):
            web["NuocTieu_NiTrit"] = "Negative"
        else:
            web["NuocTieu_NiTrit"] = str(v).strip()

    # Prefer dipstick Máu/Bạch cầu Âm tính over sediment counts
    # Urobilinogen -> µmol/L (only convert when unit is mg/dL or clearly mg-scale)
    v, u, _ = get("urobilinogen")
    if v is not None and not is_negative(str(v)):
        num = parse_number(str(v))
        un = norm(u or "")
        if num is not None:
            if "mg" in un or (not un and num <= 1.5):
                web["NuocTieu_Urobilinogen"] = round(num * 16.9, 2)
                notes.append(f"Urobilinogen {num} {u or 'mg/dL?'} -> {web['NuocTieu_Urobilinogen']} µmol/L")
            else:
                web["NuocTieu_Urobilinogen"] = num

    for key, dst in [
        ("protein", "NuocTieu_Protein"),
        ("ketone", "NuocTieu_Cetonic"),
        ("bilirubin", "NuocTieu_Bilirubin"),
        ("urine_blood", "NuocTieu_HC"),
        ("urine_rbc", "NuocTieu_HC"),
        ("urine_wbc", "NuocTieu_BC"),
    ]:
        v, u, r = get(key)
        if v is None:
            continue
        if dst in web and web[dst] == "Negative":
            continue
        if is_negative(str(v)):
            web[dst] = "Negative"
            notes.append(f"{key}: Âm tính -> Negative")
        else:
            if dst in web:
                continue
            web[dst] = str(v).strip()

    # Urine glucose specifically (may collide with blood glucose key if poorly parsed)
    # Prefer dedicated urine glucose if raw has urine section marker later.

    return web, notes


UNIT_TOKEN_RE = re.compile(
    r"(10\^9/L|10\^12/L|g/dL|g/L|mg/dL|mg/%|mg%|mmol/L|mcmol/L|µmol/L|μmol/L|U/L|fL|pg|G/L|T/L|L/L|%|Leu/µL|Leu/uL|Ery/µL|Ery/uL)\s*$",
    re.I,
)
VALUE_TOKEN_RE = re.compile(
    r"(âm tính|am tinh|\(\+\+\+\)|\(\+\+\)|\(\+\)|\+\+\+|\+\+|\+|\(\s*\d+(?:[.,]\d+)?\s*-\s*\d+(?:[.,]\d+)?\s*\)|\d+(?:[.,]\d+)?)",
    re.I,
)


def parse_line_value_unit(line: str, alias: str) -> Optional[Tuple[str, str]]:
    nline = norm(line)
    needle = alias
    idx = nline.find(needle)
    if idx < 0:
        return None
    after = nline[idx + len(needle) :]
    # unit is usually the last token on the layout line
    unit = ""
    um = UNIT_TOKEN_RE.search(nline)
    if um:
        unit = um.group(1)
    vm = VALUE_TOKEN_RE.search(after)
    if not vm:
        return None
    value = vm.group(1)
    # Ignore values that are clearly the lower bound of a ref range when another value exists?
    return value.replace(",", "."), unit


def extract_analytes(text: str) -> Dict[str, Dict[str, Any]]:
    """Extract analytes with a more structured approach on layout lines."""
    raw: Dict[str, Dict[str, Any]] = {}
    lines = text.splitlines()
    # Determine section for each line
    section = "other"
    section_of_line: List[str] = []
    for line in lines:
        nline = norm(line)
        if "sinh hoa" in nline:
            section = "sinhhoa"
        elif "nuoc tieu" in nline or "phan tich nuoc tieu" in nline:
            section = "urine"
        elif "huyet hoc" in nline or "cong thuc mau" in nline:
            section = "huyethoc"
        section_of_line.append(section)

    # Sort aliases longest-first to match neutrophils # before neutrophils
    aliases_sorted = sorted(ANALYTE_ALIASES.items(), key=lambda kv: -len(kv[0]))

    for i, line in enumerate(lines):
        nline = norm(line)
        sec = section_of_line[i]
        for alias, key in aliases_sorted:
            if alias not in nline:
                continue
            if key.endswith("_abs") and "#" not in nline:
                continue
            if key.endswith("_pct") and "#" in nline:
                continue
            # glucose collision: blood vs urine
            use_key = key
            if key == "glucose":
                if sec == "urine":
                    use_key = "urine_glucose"
                elif sec != "sinhhoa" and "glucose" in raw:
                    continue
            # urine-only analytes
            if key in {"protein", "ketone", "bilirubin", "nitrite", "urobilinogen", "sg", "urine_wbc", "urine_rbc"}:
                if sec != "urine":
                    continue
            parsed = parse_line_value_unit(line, alias)
            if not parsed:
                continue
            value, unit = parsed
            if use_key in raw:
                continue
            raw[use_key] = {"value": value, "unit": unit, "raw_line": line.strip(), "section": sec}

    # Special urine patterns (pH / Máu)
    for i, line in enumerate(lines):
        if section_of_line[i] != "urine":
            continue
        for key, cre in URINE_SPECIAL:
            if key in raw:
                continue
            m = cre.search(line)
            if m:
                raw[key] = {
                    "value": m.group("value").replace(",", "."),
                    "unit": "",
                    "raw_line": line.strip(),
                    "section": "urine",
                }

    return raw


WEB_FIELDS_DINH_KY = [
    "CongThucMau_SLHC",
    "XNM_HuyetSacTo",
    "XNM_Hematocrit",
    "XNM_MCV",
    "XNM_MCH",
    "XNM_MCHC",
    "XNM_RDW",
    "CongThucMau_SLBC",
    "SLBC_TrungTinh",
    "SLBC_lympho",
    "SLBC_DonNhan",
    "SLBC_AiToan",
    "SLBC_AiKiem",
    "CongThucMau_SLTC",
    "SinhHoaMau_DuongMau",
    "SinhHoaMau_Ure",
    "SinhHoaMau_Creatinin",
    "SinhHoaMau_ASAT_GOT",
    "SinhHoaMau_ALAT_GPT",
    "NuocTieu_TiTrong",
    "NuocTieu_pH",
    "NuocTieu_BC",
    "NuocTieu_HC",
    "NuocTieu_NiTrit",
    "NuocTieu_Protein",
    "NuocTieu_Duong",
    "NuocTieu_Cetonic",
    "NuocTieu_Bilirubin",
    "NuocTieu_Urobilinogen",
]


def process_pdf(path: Path) -> Dict[str, Any]:
    meta = parse_filename(path)
    text = pdf_text(path)
    # Prefer layout text for parsing
    try:
        import subprocess

        layout = subprocess.check_output(["pdftotext", "-layout", str(path), "-"], text=True, errors="ignore")
    except Exception:
        layout = text
    header = extract_header(layout, meta)
    raw = extract_analytes(layout)
    web, notes = convert_units(raw)
    # urine glucose
    ug = raw.get("urine_glucose")
    if ug:
        if is_negative(str(ug["value"])):
            web["NuocTieu_Duong"] = "Negative"
            notes.append("urine glucose: Âm tính -> Negative")
        else:
            web["NuocTieu_Duong"] = str(ug["value"])
    row = {
        "file": str(path),
        "batch": path.parent.name,
        **header,
        "raw_json": json.dumps(raw, ensure_ascii=False),
        "notes": " | ".join(notes),
        "missing_core": "",
    }
    for f in WEB_FIELDS_DINH_KY:
        row[f] = web.get(f)
    # core required for useful import
    core = ["CongThucMau_SLBC", "SLBC_TrungTinh", "XNM_HuyetSacTo", "CongThucMau_SLHC"]
    missing = [c for c in core if row.get(c) in (None, "")]
    row["missing_core"] = ",".join(missing)
    return row


def style_header(ws):
    fill = PatternFill("solid", fgColor="0F6A5A")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def write_excel(rows: List[Dict[str, Any]], out_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "KetQuaCanNhap"
    cols = [
        "batch",
        "file",
        "sid",
        "ho_ten",
        "nam_sinh",
        "gioi_tinh",
        "sdt",
        "nhom",
        "maukham",
        "LoaiKham",
        "ngay_nhan_mau",
        "ngay_co_ket_qua",
        *WEB_FIELDS_DINH_KY,
        "missing_core",
        "notes",
        "raw_json",
    ]
    ws.append(cols)
    style_header(ws)
    for r in rows:
        r = dict(r)
        r["LoaiKham"] = 5152  # Khám Định Kỳ
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    # Review sheet: unit conversions
    ws2 = wb.create_sheet("DoiDonVi")
    ws2.append(["sid", "ho_ten", "nhom", "notes", "missing_core"])
    style_header(ws2)
    for r in rows:
        if r.get("notes") or r.get("missing_core"):
            ws2.append([r.get("sid"), r.get("ho_ten"), r.get("nhom"), r.get("notes"), r.get("missing_core")])

    wb.save(out_path)


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--root", default="/workspace", help="Root containing Batch_* folders")
    p.add_argument("--out-dir", default=str(OUT_DIR_DEFAULT))
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--batches", default="", help="Comma batches e.g. Batch_001,Batch_002")
    args = p.parse_args()

    root = Path(args.root)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    batches = [b.strip() for b in args.batches.split(",") if b.strip()]
    pdfs: List[Path] = []
    if batches:
        for b in batches:
            pdfs.extend(sorted((root / b).glob("*.pdf")))
    else:
        pdfs = sorted(root.glob("Batch_*/*.pdf"))
    if args.limit:
        pdfs = pdfs[: args.limit]

    rows = []
    errors = []
    for i, pdf in enumerate(pdfs, 1):
        try:
            rows.append(process_pdf(pdf))
        except Exception as e:
            errors.append({"file": str(pdf), "error": str(e)})
        if i % 50 == 0:
            print(f"processed {i}/{len(pdfs)}")

    out_xlsx = out_dir / "CLS_ket_qua_tu_PDF_de_kiem_tra.xlsx"
    write_excel(rows, out_xlsx)
    with open(out_dir / "CLS_extract_errors.json", "w", encoding="utf-8") as f:
        json.dump(errors, f, ensure_ascii=False, indent=2)
    with open(out_dir / "CLS_extract_summary.json", "w", encoding="utf-8") as f:
        json.dump(
            {
                "total_pdfs": len(pdfs),
                "ok": len(rows),
                "errors": len(errors),
                "m3": sum(1 for r in rows if r.get("nhom") == "M3"),
                "m4": sum(1 for r in rows if r.get("nhom") == "M4"),
                "missing_core": sum(1 for r in rows if r.get("missing_core")),
            },
            f,
            ensure_ascii=False,
            indent=2,
        )
    print(f"Wrote {out_xlsx} rows={len(rows)} errors={len(errors)}")


if __name__ == "__main__":
    main()
