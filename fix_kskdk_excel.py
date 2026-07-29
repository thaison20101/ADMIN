#!/usr/bin/env python3
"""Fix KSKDK Excel: standard data validation + searchable aliases for dropdowns."""

import re
import unicodedata
from copy import copy

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import column_index_from_string

SRC = "/home/ubuntu/.cursor/projects/workspace/uploads/KSKDK_TTHC_mau_nhap_839d.xlsx"
OUT = "/workspace/KSKDK_TTHC_mau_nhap.xlsx"

# NhapLieu column -> DM sheet (fields with dropdown catalogs)
FIELD_MAP = {
    "B": "DM_DoiTuong",
    "C": "DM_DiaDiemKham",
    "D": "DM_HinhThucChiTra",
    "E": "DM_HinhThucKham",
    "J": "DM_GioiTinh",
    "K": "DM_DanToc",
    "L": "DM_NhomMau",
    "M": "DM_YeuToNhomMau",
    "Q": "DM_TinhThanh",
    "R": "DM_XaPhuong",
    "U": "DM_NoiCongTac",
}

ADMIN_PREFIXES = [
    re.compile(r"^Phường\s+", re.I),
    re.compile(r"^Xã\s+", re.I),
    re.compile(r"^Thị trấn\s+", re.I),
    re.compile(r"^Thành phố\s+", re.I),
    re.compile(r"^Thành Phố\s+", re.I),
    re.compile(r"^Tỉnh\s+", re.I),
]

ORG_PREFIXES = [
    re.compile(r"^Phòng khám đa khoa thuộc\s+", re.I),
    re.compile(r"^Chi nhánh\s+", re.I),
]

DATA_START_ROW = 3
DATA_END_ROW = 1000


def norm_key(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFC", str(s).strip().casefold())
    return re.sub(r"\s+", " ", s)


def strip_prefixes(name: str, patterns) -> str | None:
    if not name:
        return None
    result = str(name).strip()
    for pat in patterns:
        result = pat.sub("", result).strip()
    if norm_key(result) != norm_key(name):
        return result
    return None


def make_alias(name: str, sheet_name: str) -> str | None:
    if not name:
        return None
    name = str(name).strip()

    if sheet_name in ("DM_XaPhuong", "DM_TinhThanh"):
        return strip_prefixes(name, ADMIN_PREFIXES)

    if sheet_name == "DM_NoiCongTac":
        alias = strip_prefixes(name, ORG_PREFIXES)
        if alias:
            return alias
        # Short label: text inside parentheses at end, e.g. "(Phòng khám số 1)"
        m = re.search(r"\(([^)]+)\)\s*$", name)
        if m and len(m.group(1)) >= 3:
            return m.group(1).strip()
        return None

    if sheet_name == "DM_DoiTuong":
        if "(" in name:
            before = name.split("(", 1)[0].strip()
            if before and norm_key(before) != norm_key(name):
                return before
        return None

    if sheet_name in ("DM_DiaDiemKham", "DM_HinhThucChiTra", "DM_HinhThucKham"):
        # Drop leading generic words for easier search
        alias = re.sub(
            r"^(Cơ Sở|Khám|Ngân sách|Người|Nguồn|Tự)\s+",
            "",
            name,
            flags=re.I,
        ).strip()
        if alias and norm_key(alias) != norm_key(name):
            return alias
        return None

    return None


def build_search_lists(ws, sheet_name: str) -> int:
    """Add TimKiem (C) and GoiY (D); return last row of GoiY list."""
    ws.cell(1, 3, "TimKiem")
    ws.cell(1, 4, "GoiY")

    last_data = ws.max_row
    seen = set()
    goi_y_rows = []

    for r in range(2, last_data + 1):
        official = ws.cell(r, 2).value
        if official is None or str(official).strip() == "":
            continue
        official = str(official).strip()
        alias = make_alias(official, sheet_name)
        ws.cell(r, 3, alias if alias else "")

        key = norm_key(official)
        if key not in seen:
            seen.add(key)
            goi_y_rows.append(official)

        if alias:
            akey = norm_key(alias)
            if akey not in seen:
                seen.add(akey)
                goi_y_rows.append(alias)

    for i, val in enumerate(goi_y_rows, start=2):
        ws.cell(i, 4, val)

    return max(2, 1 + len(goi_y_rows))


def add_lookup_sheet(wb):
    if "_Lookup" in wb.sheetnames:
        del wb["_Lookup"]
    ws = wb.create_sheet("_Lookup", len(wb.sheetnames))
    ws.sheet_state = "hidden"
    headers = ["Sheet", "Official", "TimKiem"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)

    row = 2
    for dm in sorted(s for s in wb.sheetnames if s.startswith("DM_")):
        dm_ws = wb[dm]
        for r in range(2, dm_ws.max_row + 1):
            official = dm_ws.cell(r, 2).value
            if not official:
                continue
            alias = dm_ws.cell(r, 3).value
            ws.cell(row, 1, dm)
            ws.cell(row, 2, official)
            ws.cell(row, 3, alias or "")
            row += 1


def add_chuan_hoa_columns(wb):
    """Hidden helper columns: map alias -> official name for import scripts."""
    ws = wb["NhapLieu"]
    specs = [
        ("X", "B", "DM_DoiTuong"),
        ("Y", "C", "DM_DiaDiemKham"),
        ("Z", "D", "DM_HinhThucChiTra"),
        ("AA", "E", "DM_HinhThucKham"),
        ("AB", "J", "DM_GioiTinh"),
        ("AC", "K", "DM_DanToc"),
        ("AD", "L", "DM_NhomMau"),
        ("AE", "M", "DM_YeuToNhomMau"),
        ("AF", "Q", "DM_TinhThanh"),
        ("AG", "R", "DM_XaPhuong"),
        ("AH", "U", "DM_NoiCongTac"),
    ]
    for col_chuan, col_src, dm in specs:
        col_idx = column_index_from_string(col_chuan)
        ws.cell(1, col_idx, f"_Chuan_{col_src}")
        ws.cell(2, col_idx, f"Tên chuẩn (ẩn) — tra cứu từ {dm}")
        for r in range(DATA_START_ROW, DATA_END_ROW + 1):
            src = f"{col_src}{r}"
            formula = (
                f'=IF({src}="","",IF(COUNTIF({dm}!$B:$B,{src}),{src},'
                f"IFERROR(INDEX({dm}!$B:$B,MATCH({src},{dm}!$C:$C,0)),{src})))"
            )
            ws.cell(r, col_idx, formula)
        ws.column_dimensions[col_chuan].hidden = True


def update_huong_dan(wb):
    ws = wb["HuongDan"]
    ws.cell(4, 2, "2) Cột danh mục: gõ tên viết tắt hoặc chọn từ gợi ý (vd: Phú Định → Phường Phú Định)")
    ws.cell(8, 1, "Gợi ý nhập")
    ws.cell(8, 2, "Gõ vài ký tự đầu (vd: Phú Định, Hồ Chí Minh) — Excel lọc danh sách gợi ý")
    ws.cell(9, 1, "Tên chuẩn")
    ws.cell(9, 2, "Cột _Chuan_* (ẩn) tự khớp tên đầy đủ khi nhập tên viết tắt")


def apply_data_validations(wb, goi_y_end_rows: dict[str, int]):
    ws = wb["NhapLieu"]
    # Remove any existing validations openpyxl loaded (x14 already stripped on load)
    ws.data_validations.dataValidation.clear()

    for col, dm in FIELD_MAP.items():
        end = goi_y_end_rows[dm]
        formula = f"{dm}!$D$2:$D${end}"
        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=True,
            showErrorMessage=True,
            showInputMessage=True,
        )
        dv.errorTitle = "Sai danh mục"
        dv.error = "Chọn giá trị trong danh sách gợi ý hoặc nhập đúng tên viết tắt"
        dv.promptTitle = "Gợi ý danh mục"
        dv.prompt = "Gõ tên viết tắt (vd: Phú Định) hoặc chọn từ danh sách"
        ws.add_data_validation(dv)
        dv.add(f"{col}{DATA_START_ROW}:{col}{DATA_END_ROW}")


def main():
    wb = load_workbook(SRC)

    goi_y_end_rows = {}
    for dm in set(FIELD_MAP.values()):
        end = build_search_lists(wb[dm], dm)
        goi_y_end_rows[dm] = end

    apply_data_validations(wb, goi_y_end_rows)
    add_lookup_sheet(wb)
    add_chuan_hoa_columns(wb)
    update_huong_dan(wb)

    wb.save(OUT)
    print(f"Saved: {OUT}")

    # Verify
    import zipfile

    with zipfile.ZipFile(OUT) as z:
        xml = z.read("xl/worksheets/sheet2.xml").decode()
        std = xml.count("<dataValidation")
        x14 = xml.count("x14:dataValidation")
        print(f"NhapLieu validations: standard={std}, x14={x14}")

    wb2 = load_workbook(OUT, data_only=True)
    d = wb2["DM_XaPhuong"]
    aliases = [d.cell(r, 3).value for r in range(2, d.max_row + 1) if d.cell(r, 3).value]
    assert "Phú Định" in aliases, "Phú Định alias missing"
    goi_y = [d.cell(r, 4).value for r in range(2, 200) if d.cell(r, 4).value]
    assert "Phú Định" in goi_y and "Phường Phú Định" in goi_y
    print("Verified: Phú Định alias and GoiY list OK")


if __name__ == "__main__":
    main()
