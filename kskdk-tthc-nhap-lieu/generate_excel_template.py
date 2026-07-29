#!/usr/bin/env python3
"""Tạo file Excel mẫu nhập liệu KSKDK_TTHC từ schema/lookup thật trên hệ thống."""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

API = "https://be-qlskcd.medinet.org.vn"
FORM_CODE = "KSKDK_TTHC"
FORM_ID = 1000092
DEFAULT_SITE = 130

# Cột nhập trên sheet NhapLieu (tiếng Việt) → field API
COLUMNS: List[Tuple[str, str, str]] = [
    # header, api_field, hint
    ("NgayKham", "NgayKham", "dd/MM/yyyy — Ngày khám"),
    ("DoiTuong", "DoiTuong_M13", "Chọn từ sheet DM_DoiTuong (điền Name hoặc Id)"),
    ("DiaDiemKham", "DoiTuongKham", "Chọn từ DM_DiaDiemKham"),
    ("HinhThucChiTra", "HinhThucChiTraKhamSK", "Chọn từ DM_HinhThucChiTra"),
    ("HinhThucKham", "HinhThucChiTraKhamSK_ChiTiet", "Chọn từ DM_HinhThucKham"),
    ("NguonKhac_GhiRo", "NguonKhac_GhiRo", "Nếu hình thức chi trả = nguồn khác"),
    ("CCCD", "DinhDanhCaNhan", "Số CCCD / định danh / hộ chiếu"),
    ("HoTen", "HoTen", "Họ tên IN HOA"),
    ("NgaySinh", "NgaySinh", "dd/MM/yyyy"),
    ("GioiTinh", "GioiTinh", "Nam hoặc Nữ"),
    ("DanToc", "DanTocId", "Chọn từ DM_DanToc (mặc định: Kinh)"),
    ("NhomMau", "NhomMauId", "A/B/O/AB — để trống nếu không có"),
    ("YeuToNhomMau", "YeuToNhomMauId", "Rh+ / Rh-"),
    ("BHYT", "BHYT", "Số thẻ BHYT"),
    ("SDT", "SDT", "Điện thoại"),
    ("NoiOHienTai", "DiaChiHienTai", "Số nhà, đường..."),
    ("TinhThanh", "DiaChiHienTai_Tinh", "Gõ một phần (vd. HO CHI MINH) — xem GoiY_TinhThanh"),
    ("XaPhuong", "DiaChiHienTai_XaPhuong", "Gõ một phần (vd. MINH PHUNG) — xem GoiY_XaPhuong"),
    ("NgheNghiepId", "NgheNghiepId", "Id nghề nghiệp nếu biết"),
    ("NgheNghiep", "NgheNghiep", "Tên nghề nghiệp (text)"),
    ("NoiCongTac", "NoiCongTac", "Id hoặc Name từ DM_NoiCongTac"),
    ("XaPhuongCongTac", "NoiCongTac_XaPhuong", "Xã/phường nơi công tác"),
    ("LyDoKham", "LyDoKham", "Lý do khám sức khỏe"),
    ("MaBanGhi", "_output", "Script ghi — mã bản ghi trên web"),
    ("TrangThai", "_output", "THANH_CONG / TRUNG / LOI"),
    ("GhiChu", "_output", "Thông báo từ hệ thống"),
]


def http_json(method: str, url: str, token: str, site_id: int, body: Any = None) -> Dict[str, Any]:
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={
            "Authorization": f"Bearer {token}",
            "SessionSiteId": str(site_id),
            "displaymode": "0",
            "Accept": "application/json",
            "Content-Type": "application/json",
        },
        method=method,
    )
    with urllib.request.urlopen(req, timeout=90) as resp:
        return json.loads(resp.read().decode("utf-8"))


def login(username: str, password: str) -> str:
    url = f"{API}/api/TokenAuth/Authenticate"
    payload = {"userNameOrEmailAddress": username, "password": password}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        res = json.loads(resp.read().decode("utf-8"))
    if not res.get("success"):
        raise RuntimeError(res.get("error") or res)
    return res["result"]["accessToken"]


def resolve_site(token: str) -> int:
    url = f"{API}/api/services/app/User/GetSessionSiteByViewCode?viewType=form&viewCode={FORM_CODE}"
    # Một số SessionSiteId đặc biệt (-100) gây 401 — thử 0 rồi fallback DEFAULT_SITE
    for site in (0, DEFAULT_SITE):
        try:
            res = http_json("GET", url, token, site)
            data = (res.get("result") or {}).get("data")
            if data not in (None, "", 0, "0"):
                return int(data)
            if data in (0, "0"):
                # API đôi khi trả 0 khi cache miss nhưng thực tế site form là DEFAULT_SITE
                continue
        except Exception:
            continue
    return DEFAULT_SITE


def hf(token: str, site_id: int, service_id: int, params: Optional[list] = None) -> list:
    qs = urllib.parse.urlencode({"serviceId": service_id, "SessionSiteId": site_id})
    url = f"{API}/api/services/app/DRReportService/HF_ExecuteServiceWithParam?{qs}"
    res = http_json("POST", url, token, site_id, params or [])
    data = (res.get("result") or {}).get("data")
    return data if isinstance(data, list) else []


def fetch_lookups(token: str, site_id: int) -> Dict[str, List[Dict[str, Any]]]:
    mapping = {
        "DoiTuong": 1000195,
        "DiaDiemKham": 1000198,
        "HinhThucChiTra": 1000190,
        "HinhThucKham": 1000265,
        "GioiTinh": 1000056,
        "DanToc": 1000266,
        "NhomMau": 1000260,
        "YeuToNhomMau": 1000261,
        "TinhThanh": 1001337,
        "NgheNghiep": 1000294,
        "NoiCongTac": 1000292,
    }
    out: Dict[str, List[Dict[str, Any]]] = {}
    for name, sid in mapping.items():
        out[name] = hf(token, site_id, sid, [])
    # Xã/phường theo HCM (Id=50) — phổ biến; import sẽ resolve theo tỉnh dòng dữ liệu
    hcm = next((t for t in out["TinhThanh"] if "Hồ Chí Minh" in t.get("Name", "")), None)
    tinh_id = hcm["Id"] if hcm else 50
    out["XaPhuong"] = hf(token, site_id, 1000058, [{"Varible": "Id", "Value": tinh_id}])
    return out


def style_header(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="0F6A5A")
    font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin
    ws.row_dimensions[row].height = 36


def write_dm_sheet(wb: Workbook, title: str, items: List[Dict[str, Any]]):
    from excel_autocomplete import normalize_search

    ws = wb.create_sheet(title[:31])
    ws.append(["Id", "Name", "TimKiem"])
    style_header(ws)
    for it in items:
        name = it.get("Name")
        ws.append([it.get("Id"), name, normalize_search(name)])
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 28
    return ws


def build_workbook(lookups: Dict[str, List[Dict[str, Any]]], out_path: Path):
    wb = Workbook()

    # Hướng dẫn
    hd = wb.active
    hd.title = "HuongDan"
    lines = [
        ["FORM", "KSKDK_TTHC — Thông tin hành chính"],
        ["FormId", FORM_ID],
        ["Cách dùng", "1) Điền dữ liệu vào sheet NhapLieu (mỗi dòng = 1 người)"],
        ["", "2) Cột danh mục: gõ đúng Name hoặc Id như sheet DM_*"],
        ["", "3) Ngày tháng: dd/MM/yyyy (ví dụ 29/07/2026)"],
        ["", "4) Giới tính: Nam hoặc Nữ"],
        ["", "5) Chạy import: python3 import_excel.py --excel KSKDK_TTHC_mau_nhap.xlsx --user ... --password ..."],
        ["", "6) Tìm một phần: nhập CCCD/họ tên/SĐT ở sheet TimKiem → python3 search_fill.py --excel ..."],
        ["", "7) Danh mục (xã/phường...): gõ một phần → cột GoiY_* gợi ý tên đầy đủ"],
        ["", "8) Nếu người đã khám cùng ngày: TrangThai=TRUNG (không tạo bản ghi mới)"],
        ["Lưu ý", "Web không có nút Import — script này gọi API FormToDatabaseInsert thay thế"],
        ["", "Không commit mật khẩu vào git"],
    ]
    for row in lines:
        hd.append(row)
    hd.column_dimensions["A"].width = 16
    hd.column_dimensions["B"].width = 100

    # Mapping ẩn/tham chiếu
    mp = wb.create_sheet("Mapping")
    mp.append(["CotExcel", "FieldAPI", "GoiY"])
    style_header(mp)
    for h, api, hint in COLUMNS:
        mp.append([h, api, hint])
    mp.column_dimensions["A"].width = 22
    mp.column_dimensions["B"].width = 28
    mp.column_dimensions["C"].width = 55

    # DM sheets
    dm_sheets = {}
    for key, sheet_name in [
        ("DoiTuong", "DM_DoiTuong"),
        ("DiaDiemKham", "DM_DiaDiemKham"),
        ("HinhThucChiTra", "DM_HinhThucChiTra"),
        ("HinhThucKham", "DM_HinhThucKham"),
        ("GioiTinh", "DM_GioiTinh"),
        ("DanToc", "DM_DanToc"),
        ("NhomMau", "DM_NhomMau"),
        ("YeuToNhomMau", "DM_YeuToNhomMau"),
        ("TinhThanh", "DM_TinhThanh"),
        ("XaPhuong", "DM_XaPhuong"),
        ("NgheNghiep", "DM_NgheNghiep"),
        ("NoiCongTac", "DM_NoiCongTac"),
    ]:
        dm_sheets[key] = write_dm_sheet(wb, sheet_name, lookups.get(key) or [])

    # Nhập liệu
    ws = wb.create_sheet("NhapLieu", 1)
    headers = [c[0] for c in COLUMNS]
    hints = [c[2] for c in COLUMNS]
    ws.append(headers)
    ws.append(hints)
    style_header(ws, 1)
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
        cell.font = Font(italic=True, color="475569", size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 48

    # Dòng 1: TRẦN NGỌC DUNG (đã có trên hệ thống — test trùng)
    sample1 = {
        "NgayKham": "29/07/2026",
        "DoiTuong": "Người lao động (theo pháp luật về an toàn, vệ sinh lao động)",
        "DiaDiemKham": "Khám Lưu Động",
        "HinhThucChiTra": "Ngân sách thành phố hỗ trợ ",
        "HinhThucKham": "Khám Theo Hợp Đồng",
        "NguonKhac_GhiRo": "",
        "CCCD": "087187000236",
        "HoTen": "TRẦN NGỌC DUNG",
        "NgaySinh": "27/06/1987",
        "GioiTinh": "Nữ",
        "DanToc": "Kinh",
        "SDT": "0989681925",
        "NoiOHienTai": "262/20 LẠC LONG QUÂN",
        "TinhThanh": next(
            (x["Name"] for x in (lookups.get("TinhThanh") or []) if "Hồ Chí Minh" in x.get("Name", "")),
            "Thành Phố Hồ Chí Minh",
        ),
        "XaPhuong": "Phường Bình Thới",
        "NgheNghiep": "GIÁO VIÊN MẦM NON",
        "LyDoKham": "Khám sức khỏe định kỳ",
        "MaBanGhi": "",
        "TrangThai": "",
        "GhiChu": "",
    }
    ws.append([sample1.get(h, "") for h in headers])

    # Dòng 2: người mới
    sample2 = dict(sample1)
    sample2.update({
        "CCCD": "079195031510",
        "HoTen": "LE THI MAI",
        "NgaySinh": "15/03/1995",
        "SDT": "0989681920",
        "NoiOHienTai": "45 Le Loi",
        "XaPhuong": "Phường Bình Thới",
        "MaBanGhi": "",
        "TrangThai": "",
        "GhiChu": "",
    })
    ws.append([sample2.get(h, "") for h in headers])

    for col, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A3"

    # Data validations for a few columns
    def add_list_validation(col_idx: int, sheet_name: str, max_row: int = 200):
        col = get_column_letter(col_idx)
        # Name column B on DM sheet
        dv = DataValidation(
            type="list",
            formula1=f"'{sheet_name}'!$B$2:$B${max(2, max_row)}",
            allow_blank=True,
        )
        dv.error = "Chọn giá trị trong danh mục"
        dv.errorTitle = "Sai danh mục"
        ws.add_data_validation(dv)
        dv.add(f"{col}3:{col}500")

    header_index = {h: i + 1 for i, h in enumerate(headers)}
    for excel_col, dm_key, dm_sheet in [
        ("DoiTuong", "DoiTuong", "DM_DoiTuong"),
        ("DiaDiemKham", "DiaDiemKham", "DM_DiaDiemKham"),
        ("HinhThucChiTra", "HinhThucChiTra", "DM_HinhThucChiTra"),
        ("HinhThucKham", "HinhThucKham", "DM_HinhThucKham"),
        ("GioiTinh", "GioiTinh", "DM_GioiTinh"),
        ("DanToc", "DanToc", "DM_DanToc"),
        ("NhomMau", "NhomMau", "DM_NhomMau"),
        ("YeuToNhomMau", "YeuToNhomMau", "DM_YeuToNhomMau"),
        ("TinhThanh", "TinhThanh", "DM_TinhThanh"),
        ("XaPhuong", "XaPhuong", "DM_XaPhuong"),
        ("NoiCongTac", "NoiCongTac", "DM_NoiCongTac"),
    ]:
        n = len(lookups.get(dm_key) or []) + 1
        add_list_validation(header_index[excel_col], dm_sheet, n)

    # Sheet TimKiem — tìm theo một phần
    tk = wb.create_sheet("TimKiem")
    tk["A1"] = "TÌM KIẾM (nhập một phần cũng được)"
    tk["A3"], tk["B3"] = "CCCD (đủ hoặc vài số cuối)", ""
    tk["A4"], tk["B4"] = "HoTen (một phần)", ""
    tk["A5"], tk["B5"] = "SDT (một phần)", ""
    tk["A6"], tk["B6"] = "NgaySinh", ""
    tk["A8"], tk["B8"] = "Số kết quả", ""
    tk["A9"], tk["B9"] = "HoTen (kết quả)", ""
    tk["A10"], tk["B10"] = "CCCD", ""
    tk["A11"], tk["B11"] = "SDT", ""
    tk["A12"], tk["B12"] = "NgaySinh", ""
    tk["A13"], tk["B13"] = "MaBanGhi", ""
    tk["A14"], tk["B14"] = "TrangThai", ""
    tk["A15"], tk["B15"] = "GhiChu", ""
    tk["A17"] = "Chạy: python3 search_fill.py --excel KSKDK_TTHC_mau_nhap.xlsx"
    tk.column_dimensions["A"].width = 28
    tk.column_dimensions["B"].width = 40

    # Sheet DaImport — lịch sử import thành công
    di = wb.create_sheet("DaImport")
    di.append(headers)
    style_header(di)
    # Ghi sẵn 1 bản ghi đã import (Traan Ngoc Dung)
    di.append([
        "29/07/2026",
        "Người lao động (theo pháp luật về an toàn, vệ sinh lao động)",
        "Khám Lưu Động",
        "Ngân sách thành phố hỗ trợ ",
        "Khám Theo Hợp Đồng",
        "",
        "087187000236",
        "TRẦN NGỌC DUNG",
        "27/06/1987",
        "Nữ",
        "",
        "",
        "",
        "",
        "0989681925",
        "262/20 LẠC LONG QUÂN",
        "Thành Phố Hồ Chí Minh",
        "Phường Bình Thới",
        "",
        "GIÁO VIÊN MẦM NON",
        "",
        "",
        "Khám sức khỏe định kỳ",
        904629,
        "THANH_CONG",
        "Đã import trước đó",
    ])

    wb.save(out_path)
    # Thêm cột GoiY_* + sheet TraCuuDM
    from excel_autocomplete import patch_workbook

    patch_workbook(str(out_path))
    print(f"Đã tạo: {out_path}")


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Tạo Excel mẫu KSKDK_TTHC")
    p.add_argument("--user", required=True)
    p.add_argument("--password", required=True)
    p.add_argument("--out", default=str(Path(__file__).with_name("KSKDK_TTHC_mau_nhap.xlsx")))
    p.add_argument("--site-id", type=int, default=None)
    p.add_argument("--lookups-json", default="", help="Dùng file lookup có sẵn (bỏ qua API)")
    args = p.parse_args(argv)

    if args.lookups_json:
        lookups_raw = json.loads(Path(args.lookups_json).read_text(encoding="utf-8"))
        # normalize to flat lists
        lookups = {}
        rename = {
            "DoiTuong_M13": "DoiTuong",
            "DoiTuongKham": "DiaDiemKham",
            "HinhThucChiTraKhamSK": "HinhThucChiTra",
            "HinhThucChiTraKhamSK_ChiTiet": "HinhThucKham",
            "DiaChiHienTai_Tinh": "TinhThanh",
            "DiaChiHienTai_XaPhuong": "XaPhuong",
            "NgheNghiepId": "NgheNghiep",
            "DanTocId": "DanToc",
            "NhomMauId": "NhomMau",
            "YeuToNhomMauId": "YeuToNhomMau",
            "GioiTinh": "GioiTinh",
            "NoiCongTac": "NoiCongTac",
        }
        for k, v in lookups_raw.items():
            items = v.get("items") if isinstance(v, dict) else v
            key = rename.get(k, k)
            lookups[key] = items or []
    else:
        token = login(args.user, args.password)
        site_id = args.site_id or resolve_site(token)
        print(f"Đăng nhập OK, SessionSiteId={site_id}")
        lookups = fetch_lookups(token, site_id)

    build_workbook(lookups, Path(args.out))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
