#!/usr/bin/env python3
"""Gợi ý danh mục trong Excel khi nhập một phần (vd. MINH PHUNG → Phường Minh Phụng)."""

from __future__ import annotations

import re
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# (cột nhập, sheet danh mục)
SUGGEST_FIELDS: List[Tuple[str, str]] = [
    ("DoiTuong", "DM_DoiTuong"),
    ("DiaDiemKham", "DM_DiaDiemKham"),
    ("HinhThucChiTra", "DM_HinhThucChiTra"),
    ("HinhThucKham", "DM_HinhThucKham"),
    ("DanToc", "DM_DanToc"),
    ("TinhThanh", "DM_TinhThanh"),
    ("XaPhuong", "DM_XaPhuong"),
    ("NgheNghiep", "DM_NgheNghiep"),
    ("NoiCongTac", "DM_NoiCongTac"),
    ("XaPhuongCongTac", "DM_XaPhuong"),
]

GOIY_PREFIX = "GoiY_"
DATA_START_ROW = 3
DATA_END_ROW = 500


def normalize_search(text: Any) -> str:
    if text is None:
        return ""
    s = unicodedata.normalize("NFD", str(text))
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    return re.sub(r"\s+", "", s.upper())


def first_match_formula(input_cell: str, dm_sheet: str, last_row: int) -> str:
    """Công thức Excel: gợi ý kết quả đầu tiên (tương thích Excel 2016+)."""
    b_rng = f"'{dm_sheet}'!$B$2:$B${last_row}"
    c_rng = f"'{dm_sheet}'!$C$2:$C${last_row}"
    row_base = f"ROW('{dm_sheet}'!$B$2)"
    q = f'SUBSTITUTE(UPPER({input_cell})," ","")'
    return (
        f'=IF({input_cell}="","",IFERROR('
        f"INDEX({b_rng},AGGREGATE(15,6,ROW({b_rng})-{row_base}+1"
        f"/ISNUMBER(SEARCH({q},{c_rng})),1)),"
        f'"Không tìm thấy"))'
    )


def nth_match_formula(input_cell: str, dm_sheet: str, last_row: int, n: int) -> str:
    b_rng = f"'{dm_sheet}'!$B$2:$B${last_row}"
    c_rng = f"'{dm_sheet}'!$C$2:$C${last_row}"
    row_base = f"ROW('{dm_sheet}'!$B$2)"
    q = f'SUBSTITUTE(UPPER({input_cell})," ","")'
    return (
        f'=IF({input_cell}="","",IFERROR('
        f"INDEX({b_rng},AGGREGATE(15,6,ROW({b_rng})-{row_base}+1"
        f"/ISNUMBER(SEARCH({q},{c_rng})),{n})),"
        f'""))'
    )


def ensure_dm_search_column(ws: Worksheet) -> int:
    """Thêm cột C (TimKiem) nếu chưa có. Trả về số dòng cuối có dữ liệu."""
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    if "TimKiem" not in headers:
        ws.cell(1, 3, "TimKiem")
        ws.cell(1, 3).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, 3).fill = PatternFill("solid", fgColor="0F6A5A")
    last = 1
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if name is None and ws.cell(r, 1).value is None:
            continue
        if name is not None:
            ws.cell(r, 3, normalize_search(name))
            last = r
    ws.column_dimensions["C"].width = 28
    return max(last, 2)


def get_headers(ws: Worksheet) -> List[str]:
    return [str(c.value).strip() if c.value is not None else "" for c in ws[1]]


def style_goiy_header(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="FEF9C3")
    cell.font = Font(bold=True, color="854D0E", size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_goiy_cell(cell) -> None:
    cell.fill = PatternFill("solid", fgColor="FFFBEB")
    cell.font = Font(color="713F12", italic=True, size=10)


def add_goiy_columns_to_nhaplieu(ws: Worksheet, dm_last_rows: Dict[str, int]) -> None:
    headers = get_headers(ws)
    # Chèn từ phải sang trái để không lệch chỉ số
    inserts: List[Tuple[int, str, str]] = []
    for field, dm_sheet in SUGGEST_FIELDS:
        if field not in headers:
            continue
        col_idx = headers.index(field) + 1
        goiy_name = f"{GOIY_PREFIX}{field}"
        if goiy_name in headers:
            continue
        inserts.append((col_idx, goiy_name, dm_sheet))
    inserts.sort(key=lambda x: x[0], reverse=True)

    for col_idx, goiy_name, dm_sheet in inserts:
        ws.insert_cols(col_idx + 1)
        ws.cell(1, col_idx + 1, goiy_name)
        style_goiy_header(ws.cell(1, col_idx + 1))
        ws.cell(2, col_idx + 1, "Gợi ý tự động — gõ một phần bên trái (có thể không dấu)")
        ws.cell(2, col_idx + 1).font = Font(italic=True, color="92400E", size=9)
        input_col = get_column_letter(col_idx)
        last_row = dm_last_rows.get(dm_sheet, 200)
        formula = first_match_formula(f"${input_col}{{row}}", dm_sheet, last_row)
        for row in range(DATA_START_ROW, DATA_END_ROW + 1):
            cell = ws.cell(row, col_idx + 1)
            cell.value = formula.format(row=row)
            style_goiy_cell(cell)
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = 28
        headers = get_headers(ws)


def build_tracuu_sheet(wb, dm_last_rows: Dict[str, int]) -> None:
    title = "TraCuuDM"
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    ws["A1"] = "TRA CỨU DANH MỤC — nhập một phần cũng được (vd. MINH PHUNG → Phường Minh Phụng)"
    ws["A1"].font = Font(bold=True, size=12, color="0F6A5A")
    ws.merge_cells("A1:D1")

    ws["A3"], ws["B3"] = "Loại danh mục", "XaPhuong"
    ws["A4"], ws["B4"] = "Từ khóa tìm", "MINH PHUNG"
    ws["A6"] = "Gợi ý (tối đa 8 kết quả)"
    ws["A6"].font = Font(bold=True)

    # Mapping loại → sheet (đặt ẩn ở cột F)
    ws["F1"] = "LoaiDM"
    ws["G1"] = "SheetDM"
    mapping = [(f, dm) for f, dm in SUGGEST_FIELDS]
    # bỏ trùng sheet
    seen = set()
    unique = []
    for f, dm in mapping:
        if dm not in seen:
            unique.append((f, dm))
            seen.add(dm)
    for i, (f, dm) in enumerate(unique, 2):
        ws.cell(i, 6, f)
        ws.cell(i, 7, dm)

    # Dropdown loại danh mục
    types = ",".join(f for f, _ in unique)
    from openpyxl.worksheet.datavalidation import DataValidation

    dv = DataValidation(type="list", formula1=f'"{types}"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add("B3")

    # Công thức gợi ý — tham chiếu B4 và sheet theo VLOOKUP B3
    for i in range(8):
        r = 7 + i
        ws.cell(r, 1, i + 1)
        # Dùng INDIRECT để chọn sheet theo B3
        # Row trong mapping: MATCH(B3,F:F,0)+1 → sheet name in G
        b_rng = 'INDIRECT("\'"&VLOOKUP(B3,$F$2:$G$20,2,0)&"\'!B2:B"&COUNTA(INDIRECT("\'"&VLOOKUP(B3,$F$2:$G$20,2,0)&"\'!B:B")))'
        c_rng = 'INDIRECT("\'"&VLOOKUP(B3,$F$2:$G$20,2,0)&"\'!C2:C"&COUNTA(INDIRECT("\'"&VLOOKUP(B3,$F$2:$G$20,2,0)&"\'!C:C")))'
        q = 'SUBSTITUTE(UPPER($B$4)," ","")'
        n = i + 1
        formula = (
            f'=IF($B$4="","",IFERROR('
            f'INDEX({b_rng},AGGREGATE(15,6,ROW({b_rng})-1+1/ISNUMBER(SEARCH({q},{c_rng})),{n})),""))'
        )
        ws.cell(r, 2, formula)
        style_goiy_cell(ws.cell(r, 2))

    ws["A16"] = "Cách dùng"
    ws["B16"] = "1) Chọn loại (XaPhuong, TinhThanh...)  2) Gõ một phần tên  3) Copy gợi ý sang cột nhập"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 18


def remove_blocking_validations(ws: Worksheet) -> None:
    """Bỏ dropdown cứng ở cột cho phép gõ một phần."""
    free_type = {f for f, _ in SUGGEST_FIELDS}
    headers = get_headers(ws)
    cols = {headers.index(f) + 1 for f in free_type if f in headers}
    if not cols or not ws.data_validations:
        return
    keep = []
    for dv in ws.data_validations.dataValidation:
        # dv.sqref có thể là MultiCellRange
        try:
            ref = str(dv.sqref)
        except Exception:
            keep.append(dv)
            continue
        skip = False
        for c in cols:
            letter = get_column_letter(c)
            if ref.startswith(letter) or f"{letter}3:" in ref or f":{letter}" in ref:
                skip = True
                break
        if not skip:
            keep.append(dv)
    ws.data_validations.dataValidation = keep


def patch_workbook(path: str, nhap_sheet: str = "NhapLieu") -> Dict[str, int]:
    wb = load_workbook(path)
    dm_last_rows: Dict[str, int] = {}
    for _, dm_sheet in SUGGEST_FIELDS:
        if dm_sheet in wb.sheetnames:
            dm_last_rows[dm_sheet] = ensure_dm_search_column(wb[dm_sheet])

    if nhap_sheet in wb.sheetnames:
        ws = wb[nhap_sheet]
        remove_blocking_validations(ws)
        add_goiy_columns_to_nhaplieu(ws, dm_last_rows)

    build_tracuu_sheet(wb, dm_last_rows)
    wb.save(path)
    return dm_last_rows


def main() -> int:
    import argparse

    p = argparse.ArgumentParser(description="Thêm cột GoiY_* và sheet TraCuuDM vào Excel KSKDK_TTHC")
    p.add_argument("--excel", required=True)
    args = p.parse_args()
    rows = patch_workbook(args.excel)
    print(f"Đã cập nhật: {args.excel}")
    for k, v in rows.items():
        print(f"  {k}: {v} dòng")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
