#!/usr/bin/env python3
"""Tạo file Excel tổng hợp KSKDK M13 — tất cả phần nhập liệu trong 1 file.

Sheet:
  HuongDan       — hướng dẫn
  NhapLieu       — Thông tin hành chính (TTHC)
  NhapLieu_TSBT  — Tiền sử bản thân
  NhapLieu_KLS   — Khám lâm sàng
  NhapLieu_CLS   — Khám cận lâm sàng
  DM_*           — danh mục dropdown (từ file TTHC)
"""

from __future__ import annotations

import argparse
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from generate_excel_phieukham import TSBT_COLS, KLS_COLS, CLS_COLS, style_header

OUT_DEFAULT = "KSKDK_M13_mau_nhap.xlsx"
TTHC_SOURCE = "KSKDK_TTHC_mau_nhap.xlsx"

# Cột liên kết giữa các sheet (cùng 1 người)
LINK_COLS = [
    ("STT", "Số thứ tự (1, 2, 3...) — cùng STT trên mọi sheet"),
    ("CCCD", "Số CCCD — khớp với sheet NhapLieu"),
    ("HoTen", "Họ tên IN HOA — khớp với sheet NhapLieu"),
    ("MaBanGhi", "Mã bản ghi TTHC (script ghi sau import; dùng cho import phiếu khám)"),
]


def copy_sheet(source_wb, target_wb, sheet_name):
    if sheet_name not in source_wb.sheetnames:
        return
    src = source_wb[sheet_name]
    if sheet_name in target_wb.sheetnames:
        del target_wb[sheet_name]
    tgt = target_wb.create_sheet(sheet_name)
    for row in src.iter_rows():
        for cell in row:
            new = tgt.cell(cell.row, cell.column, value=cell.value)
            if cell.has_style:
                new.font = copy(cell.font)
                new.fill = copy(cell.fill)
                new.border = copy(cell.border)
                new.alignment = copy(cell.alignment)
                new.number_format = cell.number_format
    for col, dim in src.column_dimensions.items():
        tgt.column_dimensions[col].width = dim.width
    for row, dim in src.row_dimensions.items():
        tgt.row_dimensions[row].height = dim.height
    tgt.freeze_panes = src.freeze_panes


def make_linked_sheet(wb, title: str, data_cols, sample_row: dict | None = None):
    link_data = [(h, g) for h, g in LINK_COLS] + list(data_cols)
    if title in wb.sheetnames:
        del wb[title]
    headers = [c[0] for c in link_data] + ["TrangThai", "GhiChu"]
    hints = [c[1] for c in link_data] + ["Script ghi", "Ghi chú"]
    ws = wb.create_sheet(title)
    ws.append(headers)
    ws.append(hints)
    style_header(ws, 1)
    hint_fill = PatternFill("solid", fgColor="E2E8F0")
    hint_font = Font(italic=True, color="475569", size=9)
    for cell in ws[2]:
        cell.fill = hint_fill
        cell.font = hint_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 36
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20 if col <= 4 else 18
    ws.freeze_panes = "A3"
    if sample_row:
        ws.append([sample_row.get(h, "") for h in headers])
    return ws


def build_full(out_path: Path, tthc_path: Path):
    # Bắt đầu từ file TTHC (có DM sheets)
    if tthc_path.exists():
        shutil.copy(tthc_path, out_path)
        wb = load_workbook(out_path)
    else:
        raise FileNotFoundError(
            f"Thiếu {tthc_path}. Chạy: python3 generate_excel_template.py --user ... --password ..."
        )

    # Cập nhật HuongDan
    if "HuongDan" in wb.sheetnames:
        del wb["HuongDan"]
    hd = wb.create_sheet("HuongDan", 0)
    lines = [
        ["FILE", "KSKDK M13 — Nhập liệu đầy đủ (1 file)"],
        ["Mẫu", "KSKDK_DanhSach_KSK_M13"],
        ["", ""],
        ["SHEET", "NỘI DUNG"],
        ["NhapLieu", "1) Thông tin hành chính — import trước"],
        ["NhapLieu_TSBT", "2) Tiền sử bản thân + thể lực"],
        ["NhapLieu_KLS", "3) Khám lâm sàng (15 chuyên khoa)"],
        ["NhapLieu_CLS", "4) Khám cận lâm sàng (XN máu, nước tiểu, XQuang)"],
        ["DM_*", "Danh mục dropdown — tham khảo khi điền NhapLieu"],
        ["", ""],
        ["CÁCH DÙNG", ""],
        ["Bước 1", "Điền sheet NhapLieu (mỗi dòng = 1 người). Cột danh mục: chọn từ DM_*"],
        ["Bước 2", "python3 import_excel.py --excel KSKDK_M13_mau_nhap.xlsx --user USER --password PASS"],
        ["Bước 3", "Ghi MaBanGhi từ cột kết quả → điền cùng STT/CCCD vào TSBT/KLS/CLS"],
        ["Bước 4", "Điền NhapLieu_TSBT, NhapLieu_KLS, NhapLieu_CLS (cùng STT/CCCD)"],
        ["Bước 5", "python3 import_phieukham.py --excel KSKDK_M13_mau_nhap.xlsx --tthc-id <MaBanGhi> ..."],
        ["", ""],
        ["GỢI Ý NHANH", ""],
        ["TSBT", "TheLuc_ChieuCao, CanNang, NhipTho, Mach, HuyetAp bắt buộc"],
        ["KLS", "ChuaPhatHienBatThuong = 1 nếu bình thường (không phát hiện bất thường)"],
        ["CLS", "Điền cột DHDL_* (định kỳ) HOẶC cột không prefix (lao động)"],
    ]
    for row in lines:
        hd.append(row)
    hd.column_dimensions["A"].width = 16
    hd.column_dimensions["B"].width = 80

    # Thêm sheet TSBT/KLS/CLS với cột liên kết + dòng mẫu
    sample_tsbt = {
        "STT": 1, "CCCD": "", "HoTen": "", "MaBanGhi": "",
        "TheLuc_ChieuCao": 160, "TheLuc_CanNang": 55,
        "TheLuc_NhipTho": 18, "TheLuc_Mach": 72,
        "TheLuc_HuyetAp_TT": 120, "TheLuc_HuyetAp_TTr": 80,
    }
    sample_kls = {
        "STT": 1, "CCCD": "", "HoTen": "", "MaBanGhi": "",
        "NoiKhoa_ChuaPhatHienBatThuong": 1,
    }
    for prefix, _ in [
        ("NoiKhoa_TuanHoan", ""), ("NoiKhoa_HoHap", ""), ("NoiKhoa_TieuHoa", ""),
        ("NoiKhoa_ThanTietNieu", ""), ("NoiKhoa_NoiTiet", ""), ("NoiKhoa_CoXuongKhop", ""),
        ("NoiKhoa_ThanKinh", ""), ("NoiKhoa_TamThan", ""), ("NgoaiKhoa", ""),
        ("DaLieu", ""), ("SanKhoa", ""), ("PhuKhoa", ""), ("Mat", ""), ("TMH", ""), ("RHM", ""),
    ]:
        sample_kls[f"{prefix}_ChuaPhatHienBatThuong"] = 1

    sample_cls = {"STT": 1, "CCCD": "", "HoTen": "", "MaBanGhi": "", "LoaiKham": 5152}

    for title, cols, sample in [
        ("NhapLieu_TSBT", TSBT_COLS, sample_tsbt),
        ("NhapLieu_KLS", KLS_COLS, sample_kls),
        ("NhapLieu_CLS", CLS_COLS, sample_cls),
    ]:
        if title in wb.sheetnames:
            del wb[title]
        make_linked_sheet(wb, title, cols, sample)

    # Thêm cột STT vào NhapLieu nếu chưa có
    if "NhapLieu" in wb.sheetnames:
        ws = wb["NhapLieu"]
        headers = [c.value for c in ws[1]]
        if headers and headers[0] != "STT":
            ws.insert_cols(1)
            ws.cell(1, 1, "STT")
            ws.cell(2, 1, "Số thứ tự (1, 2, 3...)")
            ws.cell(3, 1, 1)

    wb.save(out_path)
    print(f"Đã tạo: {out_path}")
    print("Sheets:", wb.sheetnames)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Tạo Excel M13 tổng hợp")
    p.add_argument("--out", default=OUT_DEFAULT)
    p.add_argument("--tthc", default=TTHC_SOURCE, help="File TTHC nguồn (có DM sheets)")
    args = p.parse_args()
    build_full(Path(args.out), Path(args.tthc))
