#!/usr/bin/env python3
"""Import file Excel KSKDK_TTHC lên hệ thống (thay cho chức năng Import thiếu trên web).

Đăng nhập bằng tài khoản của bạn → map Name/Id danh mục → gọi FormToDatabaseInsert.
KHÔNG lưu mật khẩu trong source code / git.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import time
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

API = "https://be-qlskcd.medinet.org.vn"
FORM_CODE = "KSKDK_TTHC"
FORM_ID = 1000092
URL_PAGE = "/nav_group/kskdk_thongtinkham/app/main/dynamicform/viewer/KSKDK_TTHC"
DEFAULT_SITE = 130

# Excel header → API field
COL_MAP = {
    "NgayKham": "NgayKham",
    "DoiTuong": "DoiTuong_M13",
    "DiaDiemKham": "DoiTuongKham",
    "HinhThucChiTra": "HinhThucChiTraKhamSK",
    "HinhThucKham": "HinhThucChiTraKhamSK_ChiTiet",
    "NguonKhac_GhiRo": "NguonKhac_GhiRo",
    "CCCD": "DinhDanhCaNhan",
    "HoTen": "HoTen",
    "NgaySinh": "NgaySinh",
    "GioiTinh": "GioiTinh",
    "DanToc": "DanTocId",
    "NhomMau": "NhomMauId",
    "YeuToNhomMau": "YeuToNhomMauId",
    "BHYT": "BHYT",
    "SDT": "SDT",
    "NoiOHienTai": "DiaChiHienTai",
    "TinhThanh": "DiaChiHienTai_Tinh",
    "XaPhuong": "DiaChiHienTai_XaPhuong",
    "NgheNghiepId": "NgheNghiepId",
    "NgheNghiep": "NgheNghiep",
    "NoiCongTac": "NoiCongTac",
    "XaPhuongCongTac": "NoiCongTac_XaPhuong",
    "LyDoKham": "LyDoKham",
}

# Cột kết quả (script ghi lại, không gửi API)
OUTPUT_COLS = ["MaBanGhi", "TrangThai", "GhiChu"]

def lookup_candidates(row: Dict[str, Any], excel_key: str, primary: Any) -> List[Any]:
    out: List[Any] = []
    if primary is not None and str(primary).strip() != "":
        out.append(primary)
    goiy = row.get(f"GoiY_{excel_key}")
    if goiy is not None:
        gs = str(goiy).strip()
        if gs and "không tìm thấy" not in gs.lower() and gs not in out:
            out.append(goiy)
    return out


def resolve_lookup(idx: LookupIndex, row: Dict[str, Any], excel_key: str, primary: Any) -> Any:
    last_err: Optional[Exception] = None
    for val in lookup_candidates(row, excel_key, primary):
        try:
            return idx.resolve(val)
        except KeyError as e:
            last_err = e
    raise KeyError(last_err or f"Không map được '{primary}'")


DUPLICATE_PATTERNS = (
    "đã khám",
    "da kham",
    "đã được khám",
    "da duoc kham",
)

# Fields that should be resolved via lookup Name→Id
LOOKUP_FIELDS = {
    "DoiTuong_M13": 1000195,
    "DoiTuongKham": 1000198,
    "HinhThucChiTraKhamSK": 1000190,
    "HinhThucChiTraKhamSK_ChiTiet": 1000265,
    "GioiTinh": 1000056,
    "DanTocId": 1000266,
    "NhomMauId": 1000260,
    "YeuToNhomMauId": 1000261,
    "DiaChiHienTai_Tinh": 1001337,
    "NoiCongTac": 1000292,
}


def load_indexes_from_excel(path: Path) -> Dict[str, LookupIndex]:
    """Đọc danh mục từ các sheet DM_* trong file Excel (ổn định hơn gọi lại API lookup)."""
    wb = load_workbook(path, data_only=True)
    sheet_map = {
        "DoiTuong_M13": "DM_DoiTuong",
        "DoiTuongKham": "DM_DiaDiemKham",
        "HinhThucChiTraKhamSK": "DM_HinhThucChiTra",
        "HinhThucChiTraKhamSK_ChiTiet": "DM_HinhThucKham",
        "GioiTinh": "DM_GioiTinh",
        "DanTocId": "DM_DanToc",
        "NhomMauId": "DM_NhomMau",
        "YeuToNhomMauId": "DM_YeuToNhomMau",
        "DiaChiHienTai_Tinh": "DM_TinhThanh",
        "NoiCongTac": "DM_NoiCongTac",
    }
    indexes: Dict[str, LookupIndex] = {}
    for api_key, sheet_name in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            indexes[api_key] = LookupIndex([])
            continue
        ws = wb[sheet_name]
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None and row[1] is None:
                continue
            items.append({"Id": row[0], "Name": row[1]})
        indexes[api_key] = LookupIndex(items)
    # Xã phường mặc định từ sheet (theo HCM); nếu đổi tỉnh sẽ refetch
    if "DM_XaPhuong" in wb.sheetnames:
        items = []
        for row in wb["DM_XaPhuong"].iter_rows(min_row=2, values_only=True):
            if row[0] is None and row[1] is None:
                continue
            items.append({"Id": row[0], "Name": row[1]})
        indexes["_XaPhuongDefault"] = LookupIndex(items)
    else:
        indexes["_XaPhuongDefault"] = LookupIndex([])
    return indexes


def http_json(
    method: str,
    url: str,
    token: str,
    site_id: int,
    body: Any = None,
    timeout: int = 90,
    retries: int = 2,
) -> Dict[str, Any]:
    last_err: Optional[Exception] = None
    for attempt in range(retries + 1):
        data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            url,
            data=data,
            headers={
                "Authorization": f"Bearer {token}",
                "SessionSiteId": str(site_id),
                "displaymode": "0",
                "Accept": "application/json",
                "Content-Type": "application/json",
            },
            method=method,
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode("utf-8", "ignore")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as e:
            raw = e.read().decode("utf-8", "ignore")
            try:
                payload = json.loads(raw)
            except Exception:
                payload = {"raw": raw[:1000]}
            last_err = RuntimeError(f"HTTP {e.code} {url}: {payload}")
            if e.code in (401, 403, 429, 502, 503) and attempt < retries:
                time.sleep(0.8 * (attempt + 1))
                continue
            raise last_err from e
    raise last_err or RuntimeError("http_json failed")


def login(username: str, password: str) -> str:
    url = f"{API}/api/TokenAuth/Authenticate"
    payload = {"userNameOrEmailAddress": username, "password": password}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=data,
        headers={"Content-Type": "application/json", "Accept": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as resp:
        res = json.loads(resp.read().decode("utf-8"))
    if not res.get("success"):
        raise RuntimeError(f"Đăng nhập thất bại: {res.get('error') or res}")
    return res["result"]["accessToken"]


def resolve_site(token: str) -> int:
    url = f"{API}/api/services/app/User/GetSessionSiteByViewCode?viewType=form&viewCode={FORM_CODE}"
    for site in (0, DEFAULT_SITE):
        try:
            res = http_json("GET", url, token, site)
            data = (res.get("result") or {}).get("data")
            if data not in (None, "", 0, "0"):
                return int(data)
        except Exception:
            continue
    return DEFAULT_SITE


def hf(token: str, site_id: int, service_id: int, params: Optional[list] = None) -> list:
    qs = urllib.parse.urlencode({"serviceId": service_id, "SessionSiteId": site_id})
    url = f"{API}/api/services/app/DRReportService/HF_ExecuteServiceWithParam?{qs}"
    res = http_json("POST", url, token, site_id, params or [])
    data = (res.get("result") or {}).get("data")
    return data if isinstance(data, list) else []


def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    return s


def normalize_name(s: str) -> str:
    return re.sub(r"\s+", " ", strip_accents(str(s)).strip().lower())


class LookupIndex:
    def __init__(self, items: List[Dict[str, Any]]):
        self.by_id: Dict[str, Any] = {}
        self.by_name: Dict[str, Any] = {}
        for it in items:
            iid = it.get("Id")
            name = it.get("Name")
            if iid is not None:
                self.by_id[str(iid)] = iid
            if name is not None:
                self.by_name[normalize_name(name)] = iid
                self.by_name[normalize_name(str(name).strip())] = iid

    def resolve(self, value: Any) -> Any:
        if value is None or str(value).strip() == "":
            return None
        s = str(value).strip()
        if s in self.by_id:
            return self.by_id[s]
        if s.isdigit() and s in self.by_id:
            return self.by_id[s]
        key = normalize_name(s)
        if key in self.by_name:
            return self.by_name[key]
        # partial contains
        for n, iid in self.by_name.items():
            if key == n or key in n or n in key:
                return iid
        raise KeyError(f"Không map được giá trị '{value}'")


def parse_date(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime("%Y-%m-%dT00:00:00")
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%S")
        except ValueError:
            continue
    # excel serial?
    try:
        n = float(s)
        # openpyxl usually gives datetime already; skip serial edge cases
    except Exception:
        pass
    raise ValueError(f"Ngày không hợp lệ: {value}")


def validate_phone(sdt: str) -> Optional[str]:
    """Kiểm tra sơ bộ SĐT trước khi gọi API (hệ thống còn kiểm tra đầu số thật)."""
    d = re.sub(r"\D", "", str(sdt or ""))
    if len(d) != 10:
        return f"SĐT phải đủ 10 chữ số (đang có {len(d)}): {sdt}"
    if not d.startswith("0"):
        return f"SĐT phải bắt đầu bằng 0: {sdt}"
    return None


def is_duplicate_message(msg: str) -> bool:
    m = normalize_name(msg or "")
    return any(p in m for p in DUPLICATE_PATTERNS)


def extract_record_id(res: Dict[str, Any]) -> Optional[Any]:
    result = res.get("result") or {}
    data = result.get("data")
    if isinstance(data, list) and data:
        row = data[0]
        if isinstance(row, dict):
            return row.get("id") or row.get("Id")
    return None


def append_daimport(wb_path: Path, row: Dict[str, Any], record_id: Any, status: str, note: str) -> None:
    wb = load_workbook(wb_path)
    if "DaImport" not in wb.sheetnames:
        return
    ws = wb["DaImport"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    values = []
    out_map = dict(row)
    out_map["MaBanGhi"] = record_id
    out_map["TrangThai"] = status
    out_map["GhiChu"] = note
    for h in headers:
        values.append(out_map.get(h, ""))
    ws.append(values)
    wb.save(wb_path)


def write_row_status(wb_path: Path, excel_row: int, record_id: Any, status: str, note: str) -> None:
    wb = load_workbook(wb_path)
    if "NhapLieu" not in wb.sheetnames:
        return
    ws = wb["NhapLieu"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    for name, val in (("MaBanGhi", record_id or ""), ("TrangThai", status), ("GhiChu", note)):
        if name not in col:
            col[name] = len(headers) + 1
            ws.cell(1, col[name], name)
            headers.append(name)
        ws.cell(excel_row, col[name], val)
    wb.save(wb_path)


def ensure_output_columns(wb_path: Path) -> None:
    wb = load_workbook(wb_path)
    if "NhapLieu" not in wb.sheetnames:
        wb.save(wb_path)
        return
    ws = wb["NhapLieu"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    for col_name in OUTPUT_COLS:
        if col_name not in headers:
            ws.cell(1, len(headers) + 1, col_name)
            headers.append(col_name)
    wb.save(wb_path)


def write_review_excel(
    source_path: Path,
    review_path: Path,
    review_rows: List[Dict[str, Any]],
) -> None:
    """Ghi file Excel riêng các dòng TRUNG / LOI để kiểm tra."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    wb_src = load_workbook(source_path, data_only=True)
    src_headers = []
    if "NhapLieu" in wb_src.sheetnames:
        src_headers = [str(c.value).strip() if c.value else "" for c in wb_src["NhapLieu"][1]]

    wb = Workbook()
    hd = wb.active
    hd.title = "TongHop"
    hd.append(["TrangThai", "SoDong", "MoTa"])
    trung = sum(1 for r in review_rows if r.get("TrangThai") == "TRUNG")
    loi = sum(1 for r in review_rows if r.get("TrangThai") == "LOI")
    hd.append(["TRUNG", trung, "Đã khám / trùng trên hệ thống — không tạo bản ghi mới"])
    hd.append(["LOI", loi, "Lỗi khi import — cần sửa và import lại"])
    hd.append(["TONG", len(review_rows), "Tổng dòng cần kiểm tra"])
    hd.column_dimensions["A"].width = 14
    hd.column_dimensions["B"].width = 10
    hd.column_dimensions["C"].width = 60

    fill_hdr = PatternFill("solid", fgColor="B45309")
    font_hdr = Font(color="FFFFFF", bold=True)

    def cell_value(item: Dict[str, Any], row_data: Dict[str, Any], header: str) -> Any:
        if header in OUTPUT_COLS:
            return item.get(header, row_data.get(header, ""))
        val = row_data.get(header, item.get(header, ""))
        return "" if val is None else val

    def make_data_sheet(title: str, status_filter: str, tab_color: str):
        ws = wb.create_sheet(title)
        ws.sheet_properties.tabColor = tab_color
        headers = src_headers if src_headers else list(review_rows[0].keys()) if review_rows else []
        if "TrangThai" not in headers:
            headers = headers + ["MaBanGhi", "TrangThai", "GhiChu"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = fill_hdr
            cell.font = font_hdr
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        filtered = [r for r in review_rows if r.get("TrangThai") == status_filter]
        for item in filtered:
            row_data = item.get("_row_data") or item
            ws.append([cell_value(item, row_data, h) for h in headers])
        ws.freeze_panes = "A2"
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 18
        return len(filtered)

    all_ws = wb.create_sheet("CanKiemTra")
    all_ws.sheet_properties.tabColor = "DC2626"
    headers = src_headers if src_headers else []
    extra = ["MaBanGhi", "TrangThai", "GhiChu", "_excel_row"]
    for h in extra:
        if h not in headers:
            headers.append(h)
    all_ws.append(headers)
    for cell in all_ws[1]:
        cell.fill = fill_hdr
        cell.font = font_hdr
    for item in review_rows:
        row_data = item.get("_row_data") or item
        all_ws.append([cell_value(item, row_data, h) for h in headers])
    all_ws.freeze_panes = "A2"

    make_data_sheet("Trung", "TRUNG", "F59E0B")
    make_data_sheet("Loi", "LOI", "EF4444")

    wb.save(review_path)
    print(f"File kiểm tra: {review_path} (TRUNG={trung}, LOI={loi})")


def append_review_row(
    review_rows: List[Dict[str, Any]],
    row: Dict[str, Any],
    status: str,
    note: str,
) -> None:
    review_rows.append({**row, "TrangThai": status, "GhiChu": note, "_row_data": dict(row)})


def load_review_rows_from_jsonl(log_path: Path) -> List[Dict[str, Any]]:
    """Đọc log import (.jsonl) và trích các dòng TRUNG / LOI."""
    review_rows: List[Dict[str, Any]] = []
    with open(log_path, encoding="utf-8") as f:
        for line in f:
            entry = json.loads(line)
            if entry.get("status") == "TRUNG":
                append_review_row(review_rows, entry["row"], "TRUNG", entry.get("note", ""))
            elif "error" in entry:
                append_review_row(review_rows, entry["row"], "LOI", entry["error"])
    return review_rows


def read_excel_rows(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if "NhapLieu" not in wb.sheetnames:
        raise RuntimeError("File Excel thiếu sheet NhapLieu")
    ws = wb["NhapLieu"]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = list(row)
        if i == 2 and values and isinstance(values[0], str) and "dd/MM" in values[0]:
            continue
        item: Dict[str, Any] = {"_excel_row": i}
        empty = True
        for h, v in zip(headers, values):
            if not h:
                continue
            if v is not None and str(v).strip() != "":
                empty = False
            item[h] = v
        if empty:
            continue
        rows.append(item)
    return rows


def build_payload(row: Dict[str, Any], indexes: Dict[str, LookupIndex], token: str, site_id: int) -> Dict[str, Any]:
    raw: Dict[str, Any] = {}
    for excel_key, api_key in COL_MAP.items():
        if excel_key not in row:
            continue
        val = row[excel_key]
        if val is None or str(val).strip() == "":
            continue
        raw[api_key] = val

    payload: Dict[str, Any] = {}

    for dk in ("NgayKham", "NgaySinh"):
        if dk in raw:
            payload[dk] = parse_date(raw[dk])

    for tk in ("NguonKhac_GhiRo", "DinhDanhCaNhan", "HoTen", "BHYT", "SDT", "DiaChiHienTai", "NgheNghiep", "LyDoKham"):
        if tk in raw:
            payload[tk] = str(raw[tk]).strip()
            if tk == "HoTen":
                payload[tk] = payload[tk].upper()

    if "HoTen" in payload:
        payload["HoTenKhongDau"] = strip_accents(payload["HoTen"]).upper()

    if "SDT" in payload:
        phone_err = validate_phone(payload["SDT"])
        if phone_err:
            raise ValueError(phone_err)

    if "NgheNghiepId" in raw and str(raw["NgheNghiepId"]).strip() != "":
        payload["NgheNghiepId"] = int(raw["NgheNghiepId"]) if str(raw["NgheNghiepId"]).isdigit() else raw["NgheNghiepId"]
    elif "NgheNghiep" in payload and payload["NgheNghiep"]:
        nghe_candidates = lookup_candidates(row, "NgheNghiep", payload["NgheNghiep"])
        resolved = False
        last_err: Optional[Exception] = None
        for nghe_val in nghe_candidates:
            try:
                items = hf(
                    token,
                    site_id,
                    1000294,
                    [{"Varible": "SearchValue", "Value": str(nghe_val).strip()}],
                )
                idx = LookupIndex(items)
                payload["NgheNghiepId"] = idx.resolve(nghe_val)
                for it in items:
                    if it.get("Id") == payload["NgheNghiepId"]:
                        payload["NgheNghiep"] = it.get("Name") or str(nghe_val).strip()
                        break
                resolved = True
                break
            except Exception as e:
                last_err = e
        if not resolved:
            raise ValueError(
                f"NgheNghiep: không tìm thấy trong danh mục. "
                f"Hãy gõ một phần và xem cột GoiY_NgheNghiep. ({last_err})"
            ) from last_err

    excel_by_api = {v: k for k, v in COL_MAP.items()}
    for api_key in LOOKUP_FIELDS:
        if api_key not in raw:
            continue
        idx = indexes.get(api_key) or LookupIndex([])
        excel_key = excel_by_api.get(api_key, api_key)
        try:
            payload[api_key] = resolve_lookup(idx, row, excel_key, raw[api_key])
        except KeyError as e:
            raise ValueError(f"{api_key}: {e}") from e

    def resolve_xa(field_name: str, excel_key: str) -> None:
        if field_name not in raw:
            return
        default_idx = indexes.get("_XaPhuongDefault") or LookupIndex([])
        for val in lookup_candidates(row, excel_key, raw[field_name]):
            try:
                payload[field_name] = default_idx.resolve(val)
                return
            except KeyError:
                pass
        tinh_id = payload.get("DiaChiHienTai_Tinh") or 50
        try:
            xa_items = hf(token, site_id, 1000058, [{"Varible": "Id", "Value": tinh_id}])
            xa_idx = LookupIndex(xa_items)
            payload[field_name] = resolve_lookup(xa_idx, row, excel_key, raw[field_name])
        except Exception as e:
            raise ValueError(
                f"{field_name}: không map được '{raw[field_name]}' — thử xem GoiY_{excel_key}. ({e})"
            ) from e

    resolve_xa("DiaChiHienTai_XaPhuong", "XaPhuong")
    resolve_xa("NoiCongTac_XaPhuong", "XaPhuongCongTac")

    if payload.get("DinhDanhCaNhan"):
        payload.setdefault("CoCCCD", 264)

    return payload


def insert_one(token: str, site_id: int, form_data: Dict[str, Any]) -> Dict[str, Any]:
    qs = urllib.parse.urlencode(
        {
            "form_id": FORM_ID,
            "UrlPage": URL_PAGE,
            "ispopup": "false",
            "istab": "false",
        }
    )
    url = f"{API}/api/services/app/FormViewer/FormToDatabaseInsert?{qs}"
    return http_json("POST", url, token, site_id, form_data)


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Import Excel lên form KSKDK_TTHC")
    p.add_argument("--excel", required=True, help="Đường dẫn file Excel NhapLieu")
    p.add_argument("--user", default="")
    p.add_argument("--password", default="")
    p.add_argument("--site-id", type=int, default=None)
    p.add_argument("--delay", type=float, default=0.4)
    p.add_argument("--limit", type=int, default=0)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--skip-sample", action="store_true", help="Bỏ dòng mẫu NGUYEN VAN A")
    p.add_argument("--on-duplicate", choices=["skip", "fail"], default="skip",
                   help="skip=đánh dấu TRUNG và bỏ qua; fail=dừng import")
    p.add_argument("--write-excel", action="store_true", default=True,
                   help="Ghi MaBanGhi/TrangThai/GhiChu vào Excel và sheet DaImport")
    p.add_argument("--no-write-excel", action="store_false", dest="write_excel")
    p.add_argument("--out", default="import_result.jsonl")
    p.add_argument("--review-excel", default="", metavar="PATH",
                   help="File Excel riêng ghi các dòng TRUNG/LOI (mặc định: <ten_file>_kiem_tra.xlsx)")
    p.add_argument("--review-from-log", default="", metavar="PATH",
                   help="Tạo file kiểm tra từ log .jsonl (không import lại)")
    args = p.parse_args(argv)

    excel_path = Path(args.excel)
    review_path = Path(args.review_excel) if args.review_excel else excel_path.with_name(
        excel_path.stem + "_kiem_tra.xlsx"
    )
    if args.review_from_log:
        review_rows = load_review_rows_from_jsonl(Path(args.review_from_log))
        if not review_rows:
            print("Không có dòng TRUNG/LOI trong log.", file=sys.stderr)
            return 2
        write_review_excel(excel_path, review_path, review_rows)
        return 0

    if not args.user or not args.password:
        print("Cần --user và --password để import.", file=sys.stderr)
        return 2

    token = login(args.user, args.password)
    site_id = args.site_id or resolve_site(token)
    print(f"Đăng nhập OK — user={args.user}, SessionSiteId={site_id}")

    if args.write_excel:
        ensure_output_columns(excel_path)

    review_rows: List[Dict[str, Any]] = []

    indexes = load_indexes_from_excel(Path(args.excel))
    for k, idx in indexes.items():
        print(f"  danh mục {k}: {len(idx.by_id)} items")

    rows = read_excel_rows(Path(args.excel))
    if args.skip_sample:
        rows = [r for r in rows if str(r.get("HoTen") or "").strip().upper() != "NGUYEN VAN A"]
    if args.limit and args.limit > 0:
        rows = rows[: args.limit]
    if not rows:
        print("Không có dòng dữ liệu để import.", file=sys.stderr)
        return 2

    ok = fail = skip = 0
    with open(args.out, "w", encoding="utf-8") as log:
        for i, row in enumerate(rows, start=1):
            excel_row = row.get("_excel_row", i + 2)
            try:
                payload = build_payload(row, indexes, token, site_id)
                if args.dry_run:
                    res = {"dry_run": True, "payload": payload}
                    print(f"[{i}/{len(rows)}] DRY-RUN {payload.get('HoTen')}")
                    status, note, rec_id = "DRY-RUN", "Chưa gửi lên web", None
                else:
                    try:
                        res = insert_one(token, site_id, payload)
                    except RuntimeError as e:
                        if "HTTP 401" in str(e):
                            token = login(args.user, args.password)
                            res = insert_one(token, site_id, payload)
                        else:
                            raise
                    result = res.get("result") if isinstance(res.get("result"), dict) else {}
                    succeeded = bool(result.get("isSucceeded", res.get("success")))
                    msg = str(result.get("message") or result.get("errorMessage") or "")
                    if succeeded:
                        rec_id = extract_record_id(res)
                        status, note = "THANH_CONG", msg or "Lưu thành công"
                        print(f"[{i}/{len(rows)}] OK {payload.get('HoTen')} id={rec_id}")
                    elif is_duplicate_message(msg):
                        rec_id = None
                        status, note = "TRUNG", msg
                        if args.on_duplicate == "fail":
                            raise RuntimeError(msg)
                        skip += 1
                        print(f"[{i}/{len(rows)}] TRUNG {payload.get('HoTen')}: {msg}")
                    else:
                        raise RuntimeError(msg or res)
                ok += 1 if status != "TRUNG" else 0
                log.write(json.dumps({"index": i, "ok": status != "TRUNG", "skipped": status == "TRUNG",
                                      "row": row, "payload": payload, "response": res,
                                      "status": status, "note": note}, ensure_ascii=False, default=str) + "\n")
                if args.write_excel and not args.dry_run:
                    prev_id = row.get("MaBanGhi")
                    write_row_status(
                        excel_path,
                        excel_row,
                        rec_id if status == "THANH_CONG" else prev_id,
                        status,
                        note,
                    )
                    if status == "THANH_CONG":
                        append_daimport(excel_path, row, rec_id, status, note)
                    elif status == "TRUNG":
                        out_row = dict(row)
                        out_row["TrangThai"] = status
                        out_row["GhiChu"] = note
                        append_daimport(excel_path, out_row, prev_id or "", status, note)
                if status == "TRUNG" and not args.dry_run:
                    append_review_row(review_rows, row, status, note)
            except Exception as e:
                fail += 1
                print(f"[{i}/{len(rows)}] FAIL {row.get('HoTen')}: {e}", file=sys.stderr)
                log.write(json.dumps({"index": i, "ok": False, "row": row, "error": str(e)}, ensure_ascii=False, default=str) + "\n")
                if args.write_excel and not args.dry_run:
                    write_row_status(excel_path, excel_row, None, "LOI", str(e))
                if not args.dry_run:
                    append_review_row(review_rows, row, "LOI", str(e))
            if args.delay > 0 and not args.dry_run:
                time.sleep(args.delay)

    print(f"Xong: ok={ok}, trung={skip}, fail={fail}, log={args.out}")
    if review_rows and not args.dry_run:
        write_review_excel(excel_path, review_path, review_rows)
    elif not review_rows and not args.dry_run:
        print("Không có dòng TRUNG/LOI — không tạo file kiểm tra.")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
