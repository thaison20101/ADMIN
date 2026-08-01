#!/usr/bin/env python3
"""Tạo file Excel mẫu nhập liệu phiếu khám KSKDK M13.

Tạo 3 sheet:
  - NhapLieu_TSBT  : Tiền sử bản thân
  - NhapLieu_KLS   : Khám lâm sàng
  - NhapLieu_CLS   : Khám cận lâm sàng
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

OUT_DEFAULT = str(Path(__file__).with_name("KSKDK_PhieuKham_mau_nhap.xlsx"))

OUTPUT_COLS = ["TrangThai", "GhiChu"]

# =============================================================
# TSBT columns: (header, goi_y)
# =============================================================
TSBT_COLS = [
    # --- Tiền sử gia đình ---
    ("TS_GiaDinh_MacBenh",
     "Có bệnh gia đình? true/false"),
    ("TS_GiaDinh_MacBenh_DanhSachBenh",
     "Danh sách mã ICD bệnh gia đình (phân cách phẩy)"),
    ("TS_GiaDinh_MacBenh_DanhSachBenh_ICD",
     "Mã ICD bổ sung"),
    ("TS_GiaDinh_MacBenh_TenBenh",
     "Tên bệnh gia đình (nếu có)"),
    # --- Tiền sử bản thân - sinh đẻ ---
    ("TS_BanThan_ConThuMay",      "Con thứ mấy (số)"),
    ("TS_BanThan_TongSoCon",      "Tổng số con"),
    ("TS_BanThan_TuoiThai",       "Tuổi thai (tuần)"),
    ("TS_BanThan_CanNangLucSinh", "Cân nặng lúc sinh (kg)"),
    ("TS_BanThan_CachSinh",       "Cách sinh (tự nhiên/mổ)"),
    ("TS_BanThan_DiTatBamSinh",   "Dị tật bẩm sinh: true/false"),
    # --- Tiền sử bản thân - bệnh tật ---
    ("TS_BanThan_MacBenh",            "Từng mắc bệnh? true/false"),
    ("TS_BanThan_MacBenh_DanhSachBenh","Danh sách bệnh ICD"),
    ("TS_BanThan_MacBenh_Khac_ICD",   "ICD khác"),
    ("TS_BanThan_TrieuChung",         "Triệu chứng"),
    ("TS_BanThan_MacBenh_NgheNghiep", "Bệnh nghề nghiệp: true/false"),
    ("TS_BanThan_DangDieuTriBenh",    "Đang điều trị bệnh: true/false"),
    ("TS_HoiBenh",      "Hỏi bệnh (văn bản)"),
    ("TS_HoiBenh_ICD",  "ICD hỏi bệnh"),
    # --- Tiền sử thai sản (phụ nữ) ---
    ("TSSan_BatDauCoKinh",         "Bắt đầu có kinh (tuổi)"),
    ("TSSan_TinhChatKinhNguyet",   "Tính chất kinh nguyệt"),
    ("TSSan_ChuKyKinh",            "Chu kỳ kinh (ngày)"),
    ("TSSan_LuongKinh",            "Lượng kinh"),
    ("TSSan_DauBungKinh",          "Đau bụng kinh: true/false"),
    ("TSSan_DaLapGiaDinh",         "Đã lập gia đình: true/false"),
    ("TSSan_Para",                  "Para (số con)"),
    ("TSSan_MoSanPhuKhoa",         "Mổ sản phụ khoa: true/false"),
    ("TSSan_SoLanMo",              "Số lần mổ"),
    ("TSSan_DangApDungBPTT",       "Đang dùng biện pháp tránh thai: true/false"),
    ("TSSan_DangApDungBPTT_GhiRo", "BPTT ghi rõ"),
    # --- Thể lực (bắt buộc) ---
    ("TheLuc_ChieuCao",   "Chiều cao (cm) *"),
    ("TheLuc_CanNang",    "Cân nặng (kg) *"),
    ("TheLuc_NhipTho",    "Nhịp thở (lần/phút) *"),
    ("TheLuc_Mach",       "Mạch (lần/phút) *"),
    ("TheLuc_HuyetAp_TT", "Huyết áp tâm thu (mmHg) *"),
    ("TheLuc_HuyetAp_TTr","Huyết áp tâm trương (mmHg) *"),
    ("TheLuc_BMI",        "BMI (tự tính hoặc để trống)"),
    ("TheLuc_PhanLoai",   "Phân loại thể lực (I-V)"),
]

# =============================================================
# KLS columns
# =============================================================
CHUYEN_KHOA = [
    ("NoiKhoa_TuanHoan", "Tuần hoàn"),
    ("NoiKhoa_HoHap", "Hô hấp"),
    ("NoiKhoa_TieuHoa", "Tiêu hóa"),
    ("NoiKhoa_ThanTietNieu", "Thận - Tiết niệu"),
    ("NoiKhoa_NoiTiet", "Nội tiết"),
    ("NoiKhoa_CoXuongKhop", "Cơ - Xương - Khớp"),
    ("NoiKhoa_ThanKinh", "Thần kinh"),
    ("NoiKhoa_TamThan", "Tâm thần"),
    ("NgoaiKhoa", "Ngoại khoa"),
    ("DaLieu", "Da liễu"),
    ("SanKhoa", "Sản khoa"),
    ("PhuKhoa", "Phụ khoa"),
    ("Mat", "Mắt"),
    ("TMH", "Tai mũi họng"),
    ("RHM", "Răng hàm mặt"),
]

def kls_col(prefix: str, name: str):
    return [
        (f"{prefix}_ChuaPhatHienBatThuong", f"{name} - Chưa phát hiện bất thường (0/1)"),
        (f"{prefix}_ChanDoanSoBo",          f"{name} - Chẩn đoán sơ bộ"),
        (f"{prefix}_ChanDoanXacDinh",       f"{name} - Chẩn đoán xác định"),
        (f"{prefix}_PhanLoai",              f"{name} - Phân loại (I-V)"),
    ]

KLS_COLS = [("NoiKhoa_ChuaPhatHienBatThuong", "Nội khoa - Chưa phát hiện bất thường (0/1)")]
for prefix, name in CHUYEN_KHOA:
    KLS_COLS.extend(kls_col(prefix, name))
KLS_COLS.append(("AnPhuKhoa", "Ẩn phụ khoa (0/1)"))

# =============================================================
# CLS columns (2 loại mẫu)
# =============================================================
CLS_COLS = [
    ("LoaiKham", "5152=Định kỳ; để trống nếu lao động"),
    # Định kỳ (DHDL_)
    ("DHDL_CongThucMau_SLHC",    "HC (T/L) - Định kỳ"),
    ("DHDL_XNM_HuyetSacTo",      "Huyết sắc tố (g/L) - Định kỳ"),
    ("DHDL_XNM_Hematocrit",      "Hematocrit (L/L) - Định kỳ"),
    ("DHDL_XNM_MCV",             "MCV (fL) - Định kỳ"),
    ("DHDL_XNM_MCH",             "MCH (pg) - Định kỳ"),
    ("DHDL_XNM_MCHC",            "MCHC (g/L) - Định kỳ"),
    ("DHDL_XNM_RDW",             "RDW (%) - Định kỳ"),
    ("DHDL_CongThucMau_SLBC",    "BC (G/L) - Định kỳ"),
    ("DHDL_SLBC_TrungTinh",      "BC trung tính (G/L) - Định kỳ"),
    ("DHDL_SLBC_lympho",         "Lympho (G/L) - Định kỳ"),
    ("DHDL_SLBC_DonNhan",        "Đơn nhân (G/L) - Định kỳ"),
    ("DHDL_SLBC_AiToan",         "Ái toan (G/L) - Định kỳ"),
    ("DHDL_SLBC_AiKiem",         "Ái kiềm (G/L) - Định kỳ"),
    ("DHDL_CongThucMau_SLTC",    "Tiểu cầu (G/L) - Định kỳ"),
    ("DHDL_SinhHoaMau_DuongMau", "Đường máu (mmol/L) - Định kỳ"),
    ("DHDL_SinhHoaMau_Ure",      "Urê (mmol/L) - Định kỳ"),
    ("DHDL_SinhHoaMau_Creatinin","Creatinin (µmol/L) - Định kỳ"),
    ("DHDL_SinhHoaMau_ASAT_GOT", "ASAT/GOT (U/L) - Định kỳ"),
    ("DHDL_SinhHoaMau_ALAT_GPT", "ALAT/GPT (U/L) - Định kỳ"),
    ("DHDL_NuocTieu_TiTrong",    "NT Tỉ trọng - Định kỳ"),
    ("DHDL_NuocTieu_pH",         "NT pH - Định kỳ"),
    ("DHDL_NuocTieu_BC",         "NT BC (Leu/µL) - Định kỳ"),
    ("DHDL_NuocTieu_HC",         "NT HC (Ery/uL) - Định kỳ"),
    ("DHDL_NuocTieu_NiTrit",     "NT Nitrit (neg/pos) - Định kỳ"),
    ("DHDL_NuocTieu_Protein",    "NT Protein (g/L) - Định kỳ"),
    ("DHDL_NuocTieu_Duong",      "NT Glucose (mmol/L) - Định kỳ"),
    ("DHDL_NuocTieu_Cetonic",    "NT Cetonic (mmol/L) - Định kỳ"),
    ("DHDL_NuocTieu_Bilirubin",  "NT Bilirubin (µmol/L) - Định kỳ"),
    ("DHDL_NuocTieu_Urobilinogen","NT Urobilinogen (µmol/L) - Định kỳ"),
    ("DHDL_NuocTieu_Khac",       "NT Khác - Định kỳ"),
    ("DHDL_CDHA_XQuangTimPhoiThang","XQuang tim phổi - Định kỳ"),
    ("DHDL_CLS_Khac",            "CLS khác - Định kỳ"),
    ("DHDL_CLS_Khac_ChiTiet",    "CLS khác chi tiết - Định kỳ"),
    # Lao động (không có prefix)
    ("CongThucMau_SLHC",    "HC (T/L) - Lao động"),
    ("XNM_HuyetSacTo",      "Huyết sắc tố (g/L) - Lao động"),
    ("XNM_Hematocrit",      "Hematocrit (L/L) - Lao động"),
    ("XNM_MCV",             "MCV (fL) - Lao động"),
    ("XNM_MCH",             "MCH (pg) - Lao động"),
    ("XNM_MCHC",            "MCHC (g/L) - Lao động"),
    ("XNM_RDW",             "RDW (%) - Lao động"),
    ("CongThucMau_SLBC",    "BC (G/L) - Lao động"),
    ("SLBC_TrungTinh",      "BC trung tính (G/L) - Lao động"),
    ("SLBC_lympho",         "Lympho (G/L) - Lao động"),
    ("SLBC_DonNhan",        "Đơn nhân (G/L) - Lao động"),
    ("SLBC_AiToan",         "Ái toan (G/L) - Lao động"),
    ("SLBC_AiKiem",         "Ái kiềm (G/L) - Lao động"),
    ("CongThucMau_SLTC",    "Tiểu cầu (G/L) - Lao động"),
    ("SinhHoaMau_DuongMau", "Đường máu (mmol/L) - Lao động"),
    ("SinhHoaMau_Ure",      "Urê (mmol/L) - Lao động"),
    ("SinhHoaMau_Creatinin","Creatinin (µmol/L) - Lao động"),
    ("SinhHoaMau_ASAT_GOT", "ASAT/GOT (U/L) - Lao động"),
    ("SinhHoaMau_ALAT_GPT", "ALAT/GPT (U/L) - Lao động"),
    ("NuocTieu_TiTrong",    "NT Tỉ trọng - Lao động"),
    ("NuocTieu_pH",         "NT pH - Lao động"),
    ("NuocTieu_BC",         "NT BC (Leu/uL) - Lao động"),
    ("NuocTieu_HC",         "NT HC (Ery/uL) - Lao động"),
    ("NuocTieu_NiTrit",     "NT Nitrit (neg/pos) - Lao động"),
    ("NuocTieu_Protein",    "NT Protein (g/L) - Lao động"),
    ("NuocTieu_Duong",      "NT Glucose (mmol/L) - Lao động"),
    ("NuocTieu_Cetonic",    "NT Cetonic (mmol/L) - Lao động"),
    ("NuocTieu_Bilirubin",  "NT Bilirubin (µmol/L) - Lao động"),
    ("NuocTieu_Urobilinogen","NT Urobilinogen (µmol/L) - Lao động"),
    ("NuocTieu_Khac",       "NT Khác - Lao động"),
    ("CDHA_XQuangTimPhoiThang","XQuang tim phổi - Lao động"),
    ("XN_CTC",              "Xét nghiệm CTC (nữ)"),
    ("XN_HPV",              "Xét nghiệm HPV (nữ)"),
    ("CDHA_XQuangNhu",      "X-Quang nhũ (nữ)"),
    ("CDHA_SieuAm02TuyenVu","Siêu âm 2 tuyến vú (nữ)"),
]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="0F6A5A")
    font = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="CBD5E1"), right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"), bottom=Side(style="thin", color="CBD5E1"),
    )
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = thin
    ws.row_dimensions[row].height = 36


def make_sheet(wb: Workbook, title: str, cols):
    ws = wb.create_sheet(title)
    headers = [c[0] for c in cols] + OUTPUT_COLS
    hints = [c[1] for c in cols] + ["Script ghi", "Ghi chú lỗi"]
    ws.append(headers)
    ws.append(hints)
    style_header(ws, 1)
    hint_fill = PatternFill("solid", fgColor="E2E8F0")
    hint_font = Font(italic=True, color="475569", size=9)
    for cell in ws[2]:
        cell.fill = hint_fill
        cell.font = hint_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 36
    for col, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws.freeze_panes = "A3"
    return ws


def build_workbook(out_path: Path):
    wb = Workbook()

    # Hướng dẫn
    hd = wb.active
    hd.title = "HuongDan"
    lines = [
        ["FORM", "Phiếu khám KSKDK M13 — 3 form con"],
        ["Form 1", "KSKDK_TTHC_TienSu (Tiền sử bản thân) — DRViewer/PostData?id=1002342"],
        ["Form 2", "KSKDK_ThongTinKham (Khám lâm sàng) — DRViewer/PostData?id=1002124"],
        ["Form 3", "KSKDK_Phieu_CanLamSang (Cận lâm sàng) — FormViewerData?form_id=1000250"],
        ["Cách dùng", "1) Điền dữ liệu vào sheet NhapLieu_TSBT, NhapLieu_KLS, NhapLieu_CLS"],
        ["", "2) Chạy: python3 import_phieukham.py --excel KSKDK_PhieuKham_mau_nhap.xlsx --user ... --password ... --tthc-id <ID>"],
        ["", "3) --tthc-id = cột MaBanGhi của bản ghi TTHC (từ import_excel.py)"],
        ["", "4) --forms tsbt,kls,cls   (chọn 1 hoặc nhiều form)"],
        ["Lưu ý", "Phải import KSKDK_TTHC trước (dùng import_excel.py), rồi mới import phiếu khám"],
        ["", "Không commit mật khẩu vào git"],
    ]
    for row in lines:
        hd.append(row)
    hd.column_dimensions["A"].width = 14
    hd.column_dimensions["B"].width = 90

    make_sheet(wb, "NhapLieu_TSBT", TSBT_COLS)
    make_sheet(wb, "NhapLieu_KLS", KLS_COLS)
    make_sheet(wb, "NhapLieu_CLS", CLS_COLS)

    wb.save(out_path)
    print(f"Đã tạo: {out_path}")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--out", default=OUT_DEFAULT)
    args = p.parse_args()
    build_workbook(Path(args.out))
