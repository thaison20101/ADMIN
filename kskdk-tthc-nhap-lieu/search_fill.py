#!/usr/bin/env python3
"""Tìm dữ liệu theo CCCD / Họ tên / SĐT (kể cả nhập một phần) và điền vào sheet TimKiem."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

INPUT_FIELDS = ["CCCD", "HoTen", "SDT", "NgaySinh"]
OUTPUT_FIELDS = [
    "NgayKham", "DoiTuong", "DiaDiemKham", "HinhThucChiTra", "HinhThucKham",
    "CCCD", "HoTen", "NgaySinh", "GioiTinh", "SDT", "NoiOHienTai", "TinhThanh",
    "XaPhuong", "NgheNghiep", "LyDoKham", "MaBanGhi", "TrangThai", "GhiChu",
]


def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return s.replace("đ", "d").replace("Đ", "D")


def norm(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", strip_accents(str(s)).strip().lower())


def digits_only(s: Any) -> str:
    return re.sub(r"\D", "", str(s or ""))


def row_to_dict(headers: List[str], values: tuple) -> Dict[str, Any]:
    return {h: v for h, v in zip(headers, values) if h}


def load_records_from_sheet(ws, skip_hint: bool = True) -> List[Dict[str, Any]]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    out: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if skip_hint and i == 2 and row and isinstance(row[0], str) and "dd/MM" in row[0]:
            continue
        item = row_to_dict(headers, row)
        if any(v not in (None, "") for v in item.values()):
            item["_row"] = i
            item["_sheet"] = ws.title
            out.append(item)
    return out


def score_match(query: Dict[str, str], cand: Dict[str, Any]) -> int:
    score = 0
    q_cccd = digits_only(query.get("CCCD", ""))
    q_hoten = query.get("HoTen", "")
    q_sdt = digits_only(query.get("SDT", ""))
    q_ns = query.get("NgaySinh", "")

    c_cccd = digits_only(cand.get("CCCD") or cand.get("DinhDanhCaNhan"))
    c_hoten = norm(cand.get("HoTen"))
    c_sdt = digits_only(cand.get("SDT"))
    c_ns = norm(cand.get("NgaySinh"))

    if q_cccd and c_cccd:
        if q_cccd == c_cccd:
            score += 100
        elif c_cccd.endswith(q_cccd):
            score += 85
        elif q_cccd in c_cccd:
            score += 70
    if q_hoten and c_hoten:
        if q_hoten == c_hoten:
            score += 50
        elif q_hoten in c_hoten or c_hoten in q_hoten:
            score += 35
    if q_sdt and c_sdt:
        if q_sdt == c_sdt:
            score += 40
        elif c_sdt.endswith(q_sdt):
            score += 30
        elif q_sdt in c_sdt:
            score += 25
    if q_ns and c_ns and (q_ns in c_ns or c_ns in q_ns):
        score += 15
    return score


def search_all(wb, query: Dict[str, str], min_score: int = 25) -> List[Dict[str, Any]]:
    q = {k: norm(v) for k, v in query.items() if v}
    if not q:
        return []
    pools: List[Dict[str, Any]] = []
    for name in ("DaImport", "NhapLieu"):
        if name in wb.sheetnames:
            pools.extend(load_records_from_sheet(wb[name]))
    scored = []
    for cand in pools:
        s = score_match(q, cand)
        if s >= min_score:
            scored.append((s, cand))
    scored.sort(key=lambda x: (-x[0], x[1].get("_sheet", ""), x[1].get("_row", 0)))
    return [c for _, c in scored]


def main(argv: Optional[List[str]] = None) -> int:
    p = argparse.ArgumentParser(description="Tìm kiếm một phần trong Excel KSKDK_TTHC")
    p.add_argument("--excel", required=True)
    p.add_argument("--cccd", default="")
    p.add_argument("--hoten", default="")
    p.add_argument("--sdt", default="")
    p.add_argument("--ngaysinh", default="")
    p.add_argument("--fill-nhaplieu", action="store_true", help="Chép kết quả tốt nhất sang dòng trống NhapLieu")
    p.add_argument("--min-score", type=int, default=25)
    args = p.parse_args(argv)

    path = Path(args.excel)
    wb = load_workbook(path)
    if "TimKiem" not in wb.sheetnames:
        raise SystemExit("File thiếu sheet TimKiem")

    ws = wb["TimKiem"]
    query = {
        "CCCD": args.cccd or (ws["B3"].value or ""),
        "HoTen": args.hoten or (ws["B4"].value or ""),
        "SDT": args.sdt or (ws["B5"].value or ""),
        "NgaySinh": args.ngaysinh or (ws["B6"].value or ""),
    }
    matches = search_all(wb, query, min_score=args.min_score)

    # Ghi kết quả vào TimKiem (cột B)
    ws["B8"] = len(matches)
    if matches:
        best = matches[0]
        ws["B9"] = best.get("HoTen")
        ws["B10"] = best.get("CCCD")
        ws["B11"] = best.get("SDT")
        ws["B12"] = best.get("NgaySinh")
        ws["B13"] = best.get("MaBanGhi")
        ws["B14"] = best.get("TrangThai")
        ws["B15"] = best.get("GhiChu") or f"Tìm thấy tại {best.get('_sheet')} dòng {best.get('_row')}"
        # Chi tiết nhiều kết quả từ dòng 18
        start = 18
        ws.cell(start, 1, "STT")
        ws.cell(start, 2, "HoTen")
        ws.cell(start, 3, "CCCD")
        ws.cell(start, 4, "SDT")
        ws.cell(start, 5, "MaBanGhi")
        ws.cell(start, 6, "TrangThai")
        ws.cell(start, 7, "Nguon")
        for i, m in enumerate(matches[:20], 1):
            r = start + i
            ws.cell(r, 1, i)
            ws.cell(r, 2, m.get("HoTen"))
            ws.cell(r, 3, m.get("CCCD"))
            ws.cell(r, 4, m.get("SDT"))
            ws.cell(r, 5, m.get("MaBanGhi"))
            ws.cell(r, 6, m.get("TrangThai"))
            ws.cell(r, 7, f"{m.get('_sheet')}:{m.get('_row')}")
    else:
        ws["B9"] = "(Không tìm thấy)"
        ws["B15"] = "Thử nhập thêm CCCD hoặc họ tên. Sau import, dữ liệu sẽ có trong sheet DaImport."

    if args.fill_nhaplieu and matches and "NhapLieu" in wb.sheetnames:
        nl = wb["NhapLieu"]
        headers = [str(c.value).strip() if c.value is not None else "" for c in nl[1]]
        target_row = nl.max_row + 1
        best = matches[0]
        for col, h in enumerate(headers, 1):
            if h in best:
                nl.cell(target_row, col, best.get(h))
        ws["B16"] = f"Đã chép sang NhapLieu dòng {target_row}"

    wb.save(path)
    print(json.dumps({"query": query, "matches": len(matches), "best": matches[0] if matches else None}, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
