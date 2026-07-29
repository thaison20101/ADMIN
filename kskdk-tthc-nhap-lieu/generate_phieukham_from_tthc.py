#!/usr/bin/env python3
"""Tạo file Excel nhập TSBT/KLS/CLS, điền sẵn STT/CCCD/HoTen/MaBanGhi từ file TTHC đã import."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

from generate_excel_m13_full import LINK_COLS, build_full

SHEETS = ("NhapLieu_TSBT", "NhapLieu_KLS", "NhapLieu_CLS")


def read_success_rows(tthc_path: Path) -> list[dict]:
    wb = load_workbook(tthc_path, data_only=True)
    if "NhapLieu" not in wb.sheetnames:
        raise RuntimeError("File TTHC thiếu sheet NhapLieu")
    ws = wb["NhapLieu"]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, row))
        if str(item.get("TrangThai") or "").strip() != "THANH_CONG":
            continue
        ma = item.get("MaBanGhi")
        if ma is None or str(ma).strip() == "":
            continue
        rows.append(item)
    return rows


def prefill_phieukham_sheets(out_path: Path, people: list[dict]) -> None:
    wb = load_workbook(out_path)
    link_headers = [h for h, _ in LINK_COLS]
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        # Xóa dòng mẫu (từ dòng 3 trở đi)
        if ws.max_row >= 3:
            ws.delete_rows(3, ws.max_row - 2)
        for i, person in enumerate(people, start=1):
            row_vals = []
            for h in headers:
                if h == "STT":
                    row_vals.append(i)
                elif h == "CCCD":
                    row_vals.append(str(person.get("CCCD") or "").strip())
                elif h == "HoTen":
                    row_vals.append(str(person.get("HoTen") or "").strip().upper())
                elif h == "MaBanGhi":
                    row_vals.append(person.get("MaBanGhi"))
                elif h in ("TrangThai", "GhiChu"):
                    row_vals.append(None)
                else:
                    row_vals.append(None)
            ws.append(row_vals)
    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser(description="Tạo Excel nhập TSBT/KLS/CLS từ file TTHC đã import")
    p.add_argument("--tthc", required=True, help="File TTHC đã có MaBanGhi (THANH_CONG)")
    p.add_argument("--out", default="KSKDK_PhieuKham_nhap.xlsx")
    args = p.parse_args()

    tthc_path = Path(args.tthc)
    out_path = Path(args.out)
    people = read_success_rows(tthc_path)
    if not people:
        raise SystemExit("Không có dòng THANH_CONG có MaBanGhi trong file TTHC.")

    build_full(out_path, tthc_path)
    prefill_phieukham_sheets(out_path, people)
    print(f"Đã tạo: {out_path} — {len(people)} người (TSBT/KLS/CLS)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
