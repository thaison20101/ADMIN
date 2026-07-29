#!/usr/bin/env python3
"""Import phiếu khám KSKDK M13 lên hệ thống.

Workflow: Sau khi đã có bản ghi KSKDK_TTHC (thông tin hành chính), script này
gọi lần lượt 3 form:
  1. KSKDK_TTHC_TienSu  (Tiền sử bản thân)  — DRViewer/PostData?id=1002342
  2. KSKDK_ThongTinKham (Khám lâm sàng)     — DRViewer/PostData?id=1002124
  3. KSKDK_Phieu_CanLamSang (Cận lâm sàng)  — FormViewer/FormViewerData?form_id=1000250

Đọc dữ liệu từ file Excel (sheet NhapLieu_TSBT, NhapLieu_KLS, NhapLieu_CLS).
"""

from __future__ import annotations

import argparse
import json
import sys
import time
import uuid
import urllib.parse
import urllib.request
import urllib.error
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

API = "https://be-qlskcd.medinet.org.vn"
DEFAULT_SITE = 130
MAU_KHAM = "mauphieudk"

# Form IDs và Report IDs đã xác định
FORMS = {
    "KSKDK_TTHC_TienSu": {
        "form_id": 1000252,
        "report_id": 1002342,
        "endpoint": "drviewer",  # DRViewer/PostData
        "guid_key": "KSKDK_TTHC_TienSu_guid",
        "sheet": "NhapLieu_TSBT",
        "name": "Tiền sử bản thân",
    },
    "KSKDK_ThongTinKham": {
        "form_id": 1000249,
        "report_id": 1002124,
        "endpoint": "drviewer",
        "guid_key": "KSKDK_ThongTinKham_guid",
        "sheet": "NhapLieu_KLS",
        "name": "Khám lâm sàng",
    },
    "KSKDK_Phieu_CanLamSang": {
        "form_id": 1000250,
        "report_id": None,
        "endpoint": "formviewer",  # FormViewer/FormViewerData
        "guid_key": "KSKDK_Phieu_CanLamSang_guid",
        "sheet": "NhapLieu_CLS",
        "name": "Khám cận lâm sàng",
    },
}

# ==================== TSBT fields (Tiền sử bản thân) ====================
TSBT_FIELDS = [
    "TS_GiaDinh_MacBenh", "TS_GiaDinh_MacBenh_DanhSachBenh",
    "TS_GiaDinh_MacBenh_DanhSachBenh_ICD", "TS_GiaDinh_MacBenh_TenBenh",
    "TS_BanThan_ConThuMay", "TS_BanThan_TongSoCon", "TS_BanThan_TuoiThai",
    "TS_BanThan_CanNangLucSinh", "TS_BanThan_CachSinh", "TS_BanThan_DiTatBamSinh",
    "TS_BanThan_MacBenh", "TS_BanThan_MacBenh_DanhSachBenh",
    "TS_BanThan_MacBenh_Khac_ICD", "TS_BanThan_TrieuChung",
    "TS_BanThan_MacBenh_NgheNghiep", "TS_BanThan_DangDieuTriBenh",
    "TS_HoiBenh", "TS_HoiBenh_ICD",
    "TSSan_BatDauCoKinh", "TSSan_TinhChatKinhNguyet", "TSSan_ChuKyKinh",
    "TSSan_LuongKinh", "TSSan_DauBungKinh", "TSSan_DaLapGiaDinh",
    "TSSan_Para", "TSSan_MoSanPhuKhoa", "TSSan_SoLanMo",
    "TSSan_DangApDungBPTT", "TSSan_DangApDungBPTT_GhiRo",
    "TheLuc_ChieuCao", "TheLuc_CanNang", "TheLuc_NhipTho",
    "TheLuc_Mach", "TheLuc_HuyetAp_TT", "TheLuc_HuyetAp_TTr",
    "TheLuc_BMI", "TheLuc_PhanLoai",
]

# ==================== KLS fields (Khám lâm sàng) ====================
KLS_FIELDS = [
    "NoiKhoa_ChuaPhatHienBatThuong",
    "NoiKhoa_TuanHoan_ChuaPhatHienBatThuong", "NoiKhoa_TuanHoan_ChanDoanSoBo",
    "NoiKhoa_TuanHoan_ChanDoanXacDinh", "NoiKhoa_TuanHoan_PhanLoai",
    "NoiKhoa_HoHap_ChuaPhatHienBatThuong", "NoiKhoa_HoHap_ChanDoanSoBo",
    "NoiKhoa_HoHap_ChanDoanXacDinh", "NoiKhoa_HoHap_PhanLoai",
    "NoiKhoa_TieuHoa_ChuaPhatHienBatThuong", "NoiKhoa_TieuHoa_ChanDoanSoBo",
    "NoiKhoa_TieuHoa_ChanDoanXacDinh", "NoiKhoa_TieuHoa_PhanLoai",
    "NoiKhoa_ThanTietNieu_ChuaPhatHienBatThuong", "NoiKhoa_ThanTietNieu_ChanDoanSoBo",
    "NoiKhoa_ThanTietNieu_ChanDoanXacDinh", "NoiKhoa_ThanTietNieu_PhanLoai",
    "NoiKhoa_NoiTiet_ChuaPhatHienBatThuong", "NoiKhoa_NoiTiet_ChanDoanSoBo",
    "NoiKhoa_NoiTiet_ChanDoanXacDinh", "NoiKhoa_NoiTiet_PhanLoai",
    "NoiKhoa_CoXuongKhop_ChuaPhatHienBatThuong", "NoiKhoa_CoXuongKhop_ChanDoanSoBo",
    "NoiKhoa_CoXuongKhop_ChanDoanXacDinh", "NoiKhoa_CoXuongKhop_PhanLoai",
    "NoiKhoa_ThanKinh_ChuaPhatHienBatThuong", "NoiKhoa_ThanKinh_ChanDoanSoBo",
    "NoiKhoa_ThanKinh_ChanDoanXacDinh", "NoiKhoa_ThanKinh_PhanLoai",
    "NoiKhoa_TamThan_ChuaPhatHienBatThuong", "NoiKhoa_TamThan_ChanDoanSoBo",
    "NoiKhoa_TamThan_ChanDoanXacDinh", "NoiKhoa_TamThan_PhanLoai",
    "NgoaiKhoa_ChuaPhatHienBatThuong", "NgoaiKhoa_ChanDoanSoBo",
    "NgoaiKhoa_ChanDoanXacDinh", "NgoaiKhoa_PhanLoai",
    "DaLieu_ChuaPhatHienBatThuong", "DaLieu_ChanDoanSoBo",
    "DaLieu_ChanDoanXacDinh", "DaLieu_PhanLoai",
    "SanKhoa_ChuaPhatHienBatThuong", "SanKhoa_ChanDoanSoBo",
    "SanKhoa_ChanDoanXacDinh", "SanKhoa_PhanLoai",
    "PhuKhoa_ChuaPhatHienBatThuong", "PhuKhoa_ChanDoanSoBo",
    "PhuKhoa_ChanDoanXacDinh", "PhuKhoa_PhanLoai",
    "Mat_ChuaPhatHienBatThuong", "Mat_ChanDoanSoBo",
    "Mat_ChanDoanXacDinh", "Mat_PhanLoai",
    "TMH_ChuaPhatHienBatThuong", "TMH_ChanDoanSoBo",
    "TMH_ChanDoanXacDinh", "TMH_PhanLoai",
    "RHM_ChuaPhatHienBatThuong", "RHM_ChanDoanSoBo",
    "RHM_ChanDoanXacDinh", "RHM_PhanLoai",
    "AnPhuKhoa",
]

# ==================== CLS fields (Cận lâm sàng) ====================
CLS_FIELDS_DINH_KI = [  # Khám định kỳ (LoaiKham=5152)
    "DHDL_CongThucMau_SLHC", "DHDL_XNM_HuyetSacTo", "DHDL_XNM_Hematocrit",
    "DHDL_XNM_MCV", "DHDL_XNM_MCH", "DHDL_XNM_MCHC", "DHDL_XNM_RDW",
    "DHDL_CongThucMau_SLBC", "DHDL_SLBC_TrungTinh", "DHDL_SLBC_lympho",
    "DHDL_SLBC_DonNhan", "DHDL_SLBC_AiToan", "DHDL_SLBC_AiKiem",
    "DHDL_CongThucMau_SLTC",
    "DHDL_SinhHoaMau_DuongMau", "DHDL_SinhHoaMau_Ure", "DHDL_SinhHoaMau_Creatinin",
    "DHDL_SinhHoaMau_ASAT_GOT", "DHDL_SinhHoaMau_ALAT_GPT",
    "DHDL_NuocTieu_TiTrong", "DHDL_NuocTieu_pH", "DHDL_NuocTieu_BC",
    "DHDL_NuocTieu_HC", "DHDL_NuocTieu_NiTrit", "DHDL_NuocTieu_Protein",
    "DHDL_NuocTieu_Duong", "DHDL_NuocTieu_Cetonic", "DHDL_NuocTieu_Bilirubin",
    "DHDL_NuocTieu_Urobilinogen", "DHDL_NuocTieu_Khac",
    "DHDL_CDHA_XQuangTimPhoiThang", "DHDL_CLS_Khac", "DHDL_CLS_Khac_ChiTiet",
]
CLS_FIELDS_LAO_DONG = [  # Khám lao động (LoaiKham khác)
    "CongThucMau_SLHC", "XNM_HuyetSacTo", "XNM_Hematocrit",
    "XNM_MCV", "XNM_MCH", "XNM_MCHC", "XNM_RDW",
    "CongThucMau_SLBC", "SLBC_TrungTinh", "SLBC_lympho",
    "SLBC_DonNhan", "SLBC_AiToan", "SLBC_AiKiem", "CongThucMau_SLTC",
    "SinhHoaMau_DuongMau", "SinhHoaMau_Ure", "SinhHoaMau_Creatinin",
    "SinhHoaMau_ASAT_GOT", "SinhHoaMau_ALAT_GPT",
    "NuocTieu_TiTrong", "NuocTieu_pH", "NuocTieu_BC", "NuocTieu_HC",
    "NuocTieu_NiTrit", "NuocTieu_Protein", "NuocTieu_Duong",
    "NuocTieu_Cetonic", "NuocTieu_Bilirubin", "NuocTieu_Urobilinogen", "NuocTieu_Khac",
    "CDHA_XQuangTimPhoiThang",
    "XN_CTC", "XN_HPV", "CDHA_XQuangNhu", "CDHA_SieuAm02TuyenVu",
]


def http_json(method, url, token, site_id, body=None, timeout=90, retries=2):
    last_err = None
    for attempt in range(retries + 1):
        data = None if body is None else json.dumps(body, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(
            url, data=data,
            headers={
                "Authorization": f"Bearer {token}",
                "SessionSiteId": str(site_id),
                "displaymode": "0",
                "Accept": "application/json",
                "Content-Type": "application/json",
            },
            method=method,
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                raw = resp.read().decode("utf-8", "ignore")
                return json.loads(raw) if raw else {}
        except urllib.error.HTTPError as e:
            raw = e.read().decode("utf-8", "ignore")
            try:
                payload = json.loads(raw)
            except Exception:
                payload = {"raw": raw[:500]}
            last_err = RuntimeError(f"HTTP {e.code}: {payload}")
            if e.code in (401, 403, 429, 502, 503) and attempt < retries:
                time.sleep(0.8 * (attempt + 1))
                continue
            raise last_err from e
    raise last_err or RuntimeError("http_json failed")


def login(username, password):
    url = f"{API}/api/TokenAuth/Authenticate"
    data = json.dumps({"userNameOrEmailAddress": username, "password": password}).encode()
    req = urllib.request.Request(url, data=data,
        headers={"Content-Type": "application/json", "Accept": "application/json"}, method="POST")
    with urllib.request.urlopen(req, timeout=60) as resp:
        res = json.loads(resp.read().decode())
    if not res.get("success"):
        raise RuntimeError(f"Đăng nhập thất bại: {res.get('error') or res}")
    return res["result"]["accessToken"]


def parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).strftime("%Y-%m-%dT00:00:00")
    s = str(value).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%S")
        except ValueError:
            continue
    return s


def coerce_value(val):
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip()
    if s.lower() in ("true", "1", "có", "co", "x"):
        return True
    if s.lower() in ("false", "0", "không", "khong"):
        return False
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s


def read_excel_rows(path: Path, sheet_name: str):
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {"_excel_row": i}
        empty = True
        for h, v in zip(headers, row):
            if not h:
                continue
            if v is not None and str(v).strip():
                empty = False
            item[h] = v
        if not empty:
            rows.append(item)
    return rows


def build_form_data(row: Dict[str, Any], field_list: List[str]) -> Dict[str, Any]:
    fd: Dict[str, Any] = {}
    date_fields = {"NgayKham", "NgayBatDauLam"}
    for field in field_list:
        val = row.get(field)
        if val is None or str(val).strip() == "":
            fd[field] = None
            continue
        if field in date_fields:
            fd[field] = parse_date(val)
        else:
            fd[field] = coerce_value(val)
    return fd


def submit_tsbt(token: str, site: int, tthc_id: int, form_data: Dict[str, Any]) -> Dict[str, Any]:
    form_data["__label_action_code"] = "CREATE_PHIEU,SAVE,BACK"
    guid = str(uuid.uuid4())
    url_page = urllib.parse.quote(
        f"/nav_group/kskdk_thongtinkham/app/main/dynamicform/viewer/KSKDK_TTHC_TienSu?TTHCId={tthc_id}&maukham={MAU_KHAM}"
    )
    qs = f"id=1002342&SessionSiteId={site}&UrlPage={url_page}&ispopup=false&istab=true"
    params = [
        {"Varible": "ReportId", "Value": 1002342},
        {"Varible": "KSKDK_TTHC_TienSu_guid", "Value": guid},
        {"Varible": "maukham", "Value": MAU_KHAM},
        {"Varible": "TTHCId", "Value": str(tthc_id)},
    ]
    return http_json("POST", f"{API}/api/services/app/DRViewer/PostData?{qs}", token, site, params)


def submit_kls(token: str, site: int, tthc_id: int, form_data: Dict[str, Any]) -> Dict[str, Any]:
    form_data["__label_action_code"] = "CREATE_PHIEU,SAVE,BACK"
    guid = str(uuid.uuid4())
    url_page = urllib.parse.quote(
        f"/nav_group/kskdk_thongtinkham/app/main/dynamicform/viewer/KSKDK_ThongTinKham?TTHCId={tthc_id}&maukham={MAU_KHAM}"
    )
    qs = f"id=1002124&SessionSiteId={site}&UrlPage={url_page}&ispopup=false&istab=true"
    params = [
        {"Varible": "ReportId", "Value": 1002124},
        {"Varible": "KSKDK_ThongTinKham_guid", "Value": guid},
        {"Varible": "maukham", "Value": MAU_KHAM},
        {"Varible": "TTHCId", "Value": str(tthc_id)},
    ]
    return http_json("POST", f"{API}/api/services/app/DRViewer/PostData?{qs}", token, site, params)


def submit_cls(token: str, site: int, tthc_id: int, form_data: Dict[str, Any], is_dinh_ki: bool = False) -> Dict[str, Any]:
    form_data["__label_action_code"] = "CREATE_PHIEU,SAVE,BACK"
    form_data["LoaiKham"] = 5152 if is_dinh_ki else form_data.get("LoaiKham", 5152)
    guid = str(uuid.uuid4())
    url_page = urllib.parse.quote(
        f"/nav_group/kskdk_thongtinkham/app/main/dynamicform/viewer/KSKDK_Phieu_CanLamSang?TTHCId={tthc_id}&maukham={MAU_KHAM}"
    )
    qs = f"form_id=1000250&SessionSiteId={site}&UrlPage={url_page}&ispopup=false&istab=false"
    params = [
        {"Varible": "TTHCId", "Value": str(tthc_id)},
        {"Varible": "maukham", "Value": MAU_KHAM},
        {"Varible": "KSKDK_Phieu_CanLamSang_guid", "Value": guid},
        {"Varible": "formdata_parent", "Value": "null"},
        {"Varible": "DataFlowData", "Value": json.dumps(form_data, ensure_ascii=False)},
    ]
    return http_json("POST", f"{API}/api/services/app/FormViewer/FormViewerData?{qs}", token, site, params)


def write_status(path: Path, sheet_name: str, excel_row: int, status: str, note: str):
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        return
    ws = wb[sheet_name]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}
    if "TrangThai" in col:
        ws.cell(excel_row, col["TrangThai"], status)
    if "GhiChu" in col:
        ws.cell(excel_row, col["GhiChu"], note)
    wb.save(path)


def main(argv=None):
    p = argparse.ArgumentParser(description="Import phiếu khám KSKDK_M13 (TSBT, KLS, CLS)")
    p.add_argument("--excel", required=True)
    p.add_argument("--user", required=True)
    p.add_argument("--password", required=True)
    p.add_argument("--site-id", type=int, default=DEFAULT_SITE)
    p.add_argument("--tthc-id", type=int, required=True,
                   help="ID bản ghi KSKDK_TTHC (lấy từ cột MaBanGhi sau khi import TTHC)")
    p.add_argument("--forms", default="tsbt,kls,cls",
                   help="Các form cần import, cách nhau bởi dấu phẩy: tsbt,kls,cls")
    p.add_argument("--delay", type=float, default=0.4)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--out", default="import_phieukham_result.jsonl")
    args = p.parse_args(argv)

    token = login(args.user, args.password)
    site = args.site_id
    tthc_id = args.tthc_id
    print(f"Đăng nhập OK — user={args.user}, site={site}, tthcId={tthc_id}")

    forms_to_run = [f.strip().lower() for f in args.forms.split(",")]
    excel_path = Path(args.excel)

    results = []

    def run_form(form_key: str, sheet_name: str, field_list: List[str], submit_fn):
        if form_key not in forms_to_run:
            return
        rows = read_excel_rows(excel_path, sheet_name)
        if not rows:
            print(f"[{form_key}] Không có dữ liệu trong sheet {sheet_name}")
            return
        print(f"\n[{form_key}] Xử lý {len(rows)} dòng từ sheet {sheet_name}...")
        for i, row in enumerate(rows, 1):
            excel_row = row.get("_excel_row", i + 1)
            form_data = build_form_data(row, field_list)
            try:
                if args.dry_run:
                    print(f"  [{i}] DRY-RUN form_data keys: {list(form_data.keys())}")
                    write_status(excel_path, sheet_name, excel_row, "DRY-RUN", "Chưa gửi")
                    continue
                res = submit_fn(token, site, tthc_id, form_data)
                result = (res.get("result") or {})
                # DRViewer returns list of records on success
                ok = result.get("isSucceeded") if isinstance(result, dict) else bool(res.get("result"))
                msg = result.get("message", "") if isinstance(result, dict) else "OK"
                if ok or isinstance(res.get("result"), list):
                    status, note = "THANH_CONG", str(msg or "Lưu thành công")
                    print(f"  [{i}] OK {form_key}")
                else:
                    raise RuntimeError(msg or str(res))
            except Exception as e:
                status, note = "LOI", str(e)[:200]
                print(f"  [{i}] FAIL {form_key}: {e}", file=sys.stderr)
            write_status(excel_path, sheet_name, excel_row, status, note)
            results.append({"form": form_key, "row": i, "status": status, "note": note})
            if args.delay > 0 and not args.dry_run:
                time.sleep(args.delay)

    run_form("tsbt", "NhapLieu_TSBT", TSBT_FIELDS, submit_tsbt)
    run_form("kls", "NhapLieu_KLS", KLS_FIELDS, submit_kls)
    run_form("cls", "NhapLieu_CLS", CLS_FIELDS_DINH_KI + CLS_FIELDS_LAO_DONG, submit_cls)

    with open(args.out, "w", encoding="utf-8") as f:
        for r in results:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    ok = sum(1 for r in results if r["status"] == "THANH_CONG")
    fail = sum(1 for r in results if r["status"] == "LOI")
    print(f"\nXong: ok={ok}, fail={fail}, log={args.out}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
