#!/usr/bin/env python3
"""Phase B step 2: import READY_IMPORT cases into Medinet Khám cận lâm sàng.

Reads the latest (or given) preview Excel, imports only READY_IMPORT rows,
skips IMPORTED / SKIP_ALREADY_CLS, writes result Excel under Drive build folder.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import time
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from win_console import safe_print, setup_utf8_stdio  # noqa: E402

setup_utf8_stdio()

from medinet_api import (  # noqa: E402
    LAB_TO_FORM,
    authenticate,
    cls_has_lab_values,
    cls_missing_lab_fields,
    web_cls_looks_incomplete,
    get_cls,
    insert_cls,
    labs_to_form_payload,
    verify_cls_saved,
)
from pdf_extract import extract_pdf  # noqa: E402
from phase_b_preview import (  # noqa: E402
    LAB_COLS,
    build_root,
    fetch_unit_index,
    inbox_dir,
    load_config,
    match_patient,
)


def _preview_stamp(path: Path) -> str:
    """Extract YYYYMMDD-HHMMSS from CLS_preview_*.xlsx name; else empty."""
    m = re.search(r"CLS_preview_(\d{8}-\d{6})", path.name, re.I)
    return m.group(1) if m else ""


def count_preview_statuses(path: Path) -> Counter:
    """Count status_medinet values in preview Excel (first sheet)."""
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(h or "").strip() for h in next(rows_iter)]
        # tolerate slight header variants
        status_idx = None
        for i, h in enumerate(header):
            hl = h.lower().replace(" ", "")
            if hl in {"status_medinet", "statusmedinet", "status"}:
                status_idx = i
                break
        c: Counter = Counter()
        if status_idx is None:
            c["(missing status_medinet column)"] = 1
            return c
        for row in rows_iter:
            if not row or status_idx >= len(row):
                continue
            st = str(row[status_idx] or "").strip().upper()
            if not st:
                st = "(empty)"
            c[st] += 1
        return c
    finally:
        wb.close()


def choose_preview(build: Path, explicit: str = "") -> Path:
    """Pick preview Excel: explicit path, else file with most READY_IMPORT (tie -> newest stamp)."""
    if explicit:
        p = Path(explicit)
        if not p.exists():
            raise FileNotFoundError(p)
        return p
    folder = build / "excel_preview"
    files = [p for p in folder.glob("CLS_preview_*.xlsx") if p.is_file()]
    if not files:
        raise FileNotFoundError(f"No CLS_preview_*.xlsx in {folder}")

    ranked = []
    for p in files:
        counts = count_preview_statuses(p)
        ready = int(counts.get("READY_IMPORT", 0))
        stamp = _preview_stamp(p) or "00000000-000000"
        ranked.append((ready, stamp, p, counts))
        safe_print(f"  preview candidate: {p.name} READY_IMPORT={ready} statuses={dict(counts)}")

    # Prefer most READY_IMPORT, then newest filename stamp
    ranked.sort(key=lambda t: (t[0], t[1]), reverse=True)
    best = ranked[0]
    if best[0] == 0:
        safe_print(
            "WARN: no preview file contains READY_IMPORT. "
            "Re-run phase_b_preview, or pass -Preview path to the approved Excel."
        )
    return best[2]


def read_preview_ready(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows_iter)]
    idx = {h: i for i, h in enumerate(header)}
    # case-insensitive header map
    idx_ci = {h.lower(): i for h, i in idx.items()}

    def get(row, name, default=""):
        i = idx.get(name)
        if i is None:
            i = idx_ci.get(name.lower())
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    out = []
    for row in rows_iter:
        if not row:
            continue
        status = str(get(row, "status_medinet") or "").strip().upper()
        if status != "READY_IMPORT":
            continue
        labs = {}
        for lab in LAB_COLS:
            labs[lab] = {
                "value_raw": get(row, f"{lab}_raw"),
                "unit_raw": get(row, f"{lab}_unit_raw"),
                "value_web": get(row, f"{lab}_web"),
                "unit_web": get(row, f"{lab}_unit_web"),
                "convert_note": get(row, f"{lab}_note"),
            }
        out.append(
            {
                "file_name": str(get(row, "file_name") or ""),
                "sid": str(get(row, "sid") or ""),
                "ho_ten": str(get(row, "ho_ten") or ""),
                "nam_sinh": str(get(row, "nam_sinh") or ""),
                "gioi_tinh": str(get(row, "gioi_tinh") or ""),
                "sdt": str(get(row, "sdt") or ""),
                "mau_kham": str(get(row, "mau_kham") or ""),
                "ngay_co_kq": str(get(row, "ngay_co_kq") or ""),
                "status_medinet": status,
                "medinet_MaPhieu": str(get(row, "medinet_MaPhieu") or ""),
                "medinet_NgayKham": str(get(row, "medinet_NgayKham") or ""),
                "medinet_phieukhamId": get(row, "medinet_phieukhamId") or get(row, "phieukhamId") or "",
                "medinet_cdId": get(row, "medinet_cdId") or get(row, "cdId") or "",
                "labs": labs,
                "source_file": str(get(row, "file_name") or ""),
            }
        )
    wb.close()
    return out


def resolve_phieukham(row: dict, index: dict) -> tuple[str | None, dict | None]:
    """Return (phieukhamId, medinet_rec)."""
    pid = row.get("medinet_phieukhamId")
    if pid not in (None, ""):
        try:
            return str(int(float(str(pid)))), None
        except Exception:
            pass
    mp = str(row.get("medinet_MaPhieu") or "").strip()
    if mp and mp in index.get("by_maphieu", {}):
        rec = index["by_maphieu"][mp]
        return str(rec.get("phieukhamId") or rec.get("Id")), rec
    st, rec = match_patient(row, index)
    if rec:
        return str(rec.get("phieukhamId") or rec.get("Id")), rec
    return None, None


def write_result_excel(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Import_CLS"
    header = [
        "STT",
        "file_name",
        "ho_ten",
        "nam_sinh",
        "mau_kham",
        "medinet_MaPhieu",
        "phieukhamId",
        "import_status",
        "message",
        "verified",
        "fields_sent",
    ]
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(header, 1):
        cell = ws.cell(1, c, h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    for i, row in enumerate(rows, 1):
        vals = [
            i,
            row.get("file_name"),
            row.get("ho_ten"),
            row.get("nam_sinh"),
            row.get("mau_kham"),
            row.get("medinet_MaPhieu"),
            row.get("phieukhamId"),
            row.get("import_status"),
            row.get("message"),
            row.get("verified"),
            row.get("fields_sent"),
        ]
        for c, v in enumerate(vals, 1):
            ws.cell(1 + i, c, v)
    wb.save(path)


def update_cases_csv(cases_path: Path, rows: list[dict]) -> None:
    if not cases_path.exists():
        return
    with cases_path.open(encoding="utf-8-sig", newline="") as f:
        existing = list(csv.DictReader(f))
    by_file = {}
    for r in rows:
        by_file[Path(r.get("file_name") or "").name] = r
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for e in existing:
        src = Path(e.get("source_file") or "").name
        hit = by_file.get(src)
        if not hit:
            continue
        st = hit.get("import_status")
        if st == "IMPORTED":
            e["status"] = "IMPORTED"
            e["imported_at"] = now
            e["notes"] = "phase_b_import:ok"
        elif st:
            e["status"] = st if st.startswith("ERROR") or st in {"SKIP_ALREADY_CLS", "WAITING_ADMIN"} else e.get("status")
            e["notes"] = f"phase_b_import:{st}:{hit.get('message','')}"[:240]
            try:
                e["import_attempts"] = str(int(e.get("import_attempts") or 0) + 1)
            except Exception:
                e["import_attempts"] = "1"
        e["last_checked_at"] = now
        if hit.get("medinet_MaPhieu"):
            e["ma_phieu"] = hit["medinet_MaPhieu"]
    if existing:
        with cases_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(existing[0].keys()))
            w.writeheader()
            w.writerows(existing)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--preview", default="", help="Path to CLS_preview_*.xlsx (default: latest)")
    ap.add_argument("--limit", type=int, default=0, help="Limit READY cases (0=all)")
    ap.add_argument("--dry-run", action="store_true", help="Map payload only, no write")
    ap.add_argument("--force", action="store_true", help="Overwrite if CLS already has values")
    ap.add_argument("--sleep", type=float, default=0.35, help="Delay between imports")
    ap.add_argument("--reparse-pdf", action="store_true", help="Re-extract PDF instead of Excel *_web")
    args = ap.parse_args()

    cfg = load_config()
    build = build_root(cfg)
    safe_print(f"Build: {build}")
    safe_print("Scanning excel_preview for READY_IMPORT ...")
    preview = choose_preview(build, args.preview)
    safe_print(f"Using Preview: {preview}")
    status_counts = count_preview_statuses(preview)
    safe_print(f"Status in file: {dict(status_counts)}")

    ready = read_preview_ready(preview)
    safe_print(f"READY_IMPORT rows: {len(ready)}")
    if args.limit:
        ready = ready[: args.limit]
        safe_print(f"Limited to: {len(ready)}")
    if not ready:
        safe_print("Nothing to import.")
        safe_print(
            "Hint: pass the approved Excel explicitly, e.g.\n"
            '  -Preview "G:\\Drive của tôi\\build for Supper Data\\excel_preview\\CLS_preview_20260802-001204.xlsx"'
        )
        return 0

    user = os.environ.get("MEDINET_USER", "pkdkthuankieu")
    password = os.environ.get("MEDINET_PASS", "P@ssw0rd")
    token_box = {"t": authenticate(user, password)}

    def reauth():
        token_box["t"] = authenticate(user, password)
        return token_box["t"]

    date_from = cfg.get("medinet", {}).get("date_from", "01/07/2026")
    date_to = cfg.get("medinet", {}).get("date_to", "31/07/2026")
    safe_print("Indexing Medinet lists for phieukhamId resolve...")
    index = fetch_unit_index(token_box["t"], date_from, date_to)
    # refresh token after long index
    token_box["t"] = authenticate(user, password)

    inbox = inbox_dir(cfg)
    results = []
    for i, row in enumerate(ready, 1):
        name = row.get("ho_ten")
        safe_print(f"[{i}/{len(ready)}] {name} ...")
        pid, rec = resolve_phieukham(row, index)
        if rec:
            row["medinet_MaPhieu"] = rec.get("MaPhieu") or row.get("medinet_MaPhieu")
        if not pid:
            row.update(
                {
                    "phieukhamId": "",
                    "import_status": "WAITING_ADMIN",
                    "message": "Không resolve được phieukhamId",
                    "verified": "NO",
                    "fields_sent": 0,
                }
            )
            results.append(row)
            continue

        # Optional reparse
        labs = row.get("labs") or {}
        if args.reparse_pdf:
            pdf = inbox / row["file_name"]
            if not pdf.exists():
                # search recursively
                hits = list(inbox.rglob(row["file_name"])) if row.get("file_name") else []
                pdf = hits[0] if hits else pdf
            if pdf.exists():
                labs = extract_pdf(pdf).get("labs") or labs

        # Guard: already has CLS — overwrite if incomplete vs PDF (Urobilinogen, etc.)
        # Urea: often absent on PDF — ignore as required field
        existing, token_box["t"] = get_cls(token_box["t"], pid, reauth=reauth)
        payload = labs_to_form_payload(labs, phieukham_id=pid, gioi_tinh=row.get("gioi_tinh") or "")
        payload["LoaiKham"] = 5152
        fields_sent = len([k for k in payload if k in LAB_TO_FORM.values()])
        missing = [
            k
            for k in (cls_missing_lab_fields(existing, payload) if existing else [])
            if k != "SinhHoaMau_Ure"
        ]
        incomplete = web_cls_looks_incomplete(existing, payload) or bool(missing)
        if cls_has_lab_values(existing) and not args.force and not incomplete:
            row.update(
                {
                    "phieukhamId": pid,
                    "import_status": "SKIP_ALREADY_CLS",
                    "message": "Web đã có CLS đủ — bỏ qua (dùng --force để ghi đè)",
                    "verified": "YES",
                    "fields_sent": 0,
                }
            )
            results.append(row)
            continue
        if incomplete and cls_has_lab_values(existing):
            safe_print(
                f"  OVERWRITE incomplete {row.get('ho_ten')} missing={missing[:8] or 'heuristic'}"
            )

        if args.dry_run:
            row.update(
                {
                    "phieukhamId": pid,
                    "import_status": "DRY_RUN",
                    "message": json.dumps(payload, ensure_ascii=False)[:500],
                    "verified": "NO",
                    "fields_sent": fields_sent,
                }
            )
            results.append(row)
            continue

        ok, msg, _raw, token_box["t"] = insert_cls(token_box["t"], payload, reauth=reauth)
        time.sleep(0.15)
        verified, vdetail, token_box["t"] = verify_cls_saved(
            token_box["t"], pid, payload=payload, reauth=reauth
        )
        msg = f"{msg}; {vdetail}"
        partial_bad = ("SET-no-urine-text" in (msg or "")) or ("SET-urine-all-dropped" in (msg or ""))
        if ok and verified and fields_sent > 0 and not partial_bad:
            import_status = "IMPORTED"
        else:
            import_status = "ERROR_IMPORT"
            msg = msg or "verify failed"

        row.update(
            {
                "phieukhamId": pid,
                "import_status": import_status,
                "message": msg,
                "verified": "YES" if verified else "NO",
                "fields_sent": fields_sent,
            }
        )
        results.append(row)
        safe_print(f"  -> {import_status} fields={fields_sent} verified={verified}")
        if args.sleep:
            time.sleep(args.sleep)

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = build / "excel_preview" / f"CLS_import_result_{stamp}.xlsx"
    write_result_excel(results, out)
    cases_path = ROOT / cfg.get("tracking", {}).get("cases_csv", "tracking/cases.csv")
    update_cases_csv(cases_path, results)

    c = Counter(r.get("import_status") for r in results)
    safe_print("---")
    safe_print(f"Result Excel: {out}")
    safe_print(f"Status: {dict(c)}")
    return 0 if c.get("ERROR_IMPORT", 0) == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
