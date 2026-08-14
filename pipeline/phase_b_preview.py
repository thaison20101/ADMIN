#!/usr/bin/env python3
"""Phase B step 1: parse PDFs → Excel preview + missing/updated lists.

Does NOT import to Medinet yet. Run this first so you can verify values/units.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import time
import urllib.error
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(Path(__file__).resolve().parent))

from win_console import safe_print, setup_utf8_stdio  # noqa: E402

setup_utf8_stdio()

from pdf_extract import extract_pdf  # noqa: E402

DEFAULT_CONFIG = ROOT / "pipeline" / "config.example.json"
LOCAL_CONFIG = ROOT / "pipeline" / "config.local.json"

LAB_COLS = [
    "WBC",
    "Neutrophils_pct",
    "Neutrophils_count",
    "Eosinophils_pct",
    "Eosinophils_count",
    "Monocytes_pct",
    "Monocytes_count",
    "Basophils_pct",
    "Basophils_count",
    "Lymphocytes_pct",
    "Lymphocytes_count",
    "RBC",
    "HGB",
    "HCT",
    "MCV",
    "MCH",
    "MCHC",
    "RDW",
    "PLT",
    "MPV",
    "Glucose",
    "Urea",
    "Creatinine",
    "AST",
    "ALT",
    "Urobilinogen",
    "Glucose_NT",
    "Ketone",
    "Bilirubin_NT",
    "Protein_NT",
    "Nitrite",
    "pH_NT",
    "Mau_NT",
    "Ti_trong",
    "Bach_cau_NT",
]


def load_config() -> dict:
    path = LOCAL_CONFIG if LOCAL_CONFIG.exists() else DEFAULT_CONFIG
    with path.open(encoding="utf-8-sig") as f:
        return json.load(f)


def _resolve_existing_build(raw: str) -> Path:
    """Prefer an existing Drive build folder; try common Google Drive name variants."""
    candidates = []
    if raw:
        candidates.append(Path(raw))
    # Common Google Drive Desktop folder names on Windows
    for drive in ("G:", "H:", "D:"):
        for mid in (
            "Drive của tôi",
            "Drive của Tôi",
            "My Drive",
            "Drive cua toi",
        ):
            candidates.append(Path(f"{drive}/{mid}/build for Supper Data"))
    candidates.append(ROOT / "pipeline" / "work" / "build")

    seen = set()
    for p in candidates:
        key = str(p).lower()
        if key in seen:
            continue
        seen.add(key)
        try:
            if p.exists():
                return p
        except Exception:
            continue
    # default: configured/raw even if missing (will be created)
    return Path(raw) if raw else candidates[0]


def build_root(cfg: dict) -> Path:
    raw = cfg.get("drive", {}).get("build_root") or r"G:\Drive của tôi\build for Supper Data"
    p = _resolve_existing_build(str(raw))
    for sub in ("excel_preview", "missing_or_updated", "logs", "cases_snapshot"):
        (p / sub).mkdir(parents=True, exist_ok=True)
    return p


def inbox_dir(cfg: dict) -> Path:
    sync = Path(cfg.get("drive", {}).get("local_sync_root") or "")
    if sync.exists():
        return sync / cfg["drive"]["inbox_folder"]
    return ROOT / "INBOX_CLS"


def list_pdfs(inbox: Path, limit: int | None) -> list[Path]:
    files = sorted([p for p in inbox.rglob("*.pdf") if p.is_file()])
    if limit:
        files = files[:limit]
    return files


def authenticate(user: str, password: str) -> str:
    BE = "https://be-qlskcd.medinet.org.vn"
    req = urllib.request.Request(
        f"{BE}/api/TokenAuth/Authenticate",
        data=json.dumps(
            {"userNameOrEmailAddress": user, "password": password, "rememberClient": True}
        ).encode(),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=60) as r:
        return json.loads(r.read())["result"]["accessToken"]


def to_fparams(obj: dict) -> list:
    return [{"Varible": k, "Value": "" if v is None else str(v)} for k, v in obj.items()]


def api(token: str, path: str, method: str = "GET", body=None):
    BE = "https://be-qlskcd.medinet.org.vn"
    url = f"{BE}{path}" if path.startswith("/") else path
    data = None if body is None else json.dumps(body, ensure_ascii=False).encode()
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "SessionSiteId": "130",
    }
    for attempt in range(4):
        try:
            req = urllib.request.Request(url, data=data, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=180) as r:
                return r.status, json.loads(r.read())
        except urllib.error.HTTPError as e:
            raw = e.read().decode("utf-8", errors="replace")
            if e.code == 401 and attempt < 3:
                time.sleep(1)
                continue
            try:
                return e.code, json.loads(raw)
            except Exception:
                return e.code, {"error": raw[:500]}
        except Exception as e:
            time.sleep(1 + attempt)
            last = e
    return 0, {"error": str(last)}



def _fold_name(s: str) -> str:
    """Uppercase + strip Vietnamese accents for soft name match."""
    import unicodedata

    s = unicodedata.normalize("NFD", (s or "").upper())
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", s).strip()


def _year_from_ngaysinh(ns) -> str:
    """Extract birth year from Medinet NgaySinh.

    API may return ISO `1943-01-01` OR DMY `01/01/1943`. Taking [:4] on DMY
    wrongly yields `01/0` and breaks all name|year matches → mass NO_TTHC.
    """
    s = str(ns or "").strip()
    if not s:
        return ""
    if re.match(r"^(19|20)\d{2}\b", s):
        return s[:4]
    m = re.search(r"(19\d{2}|20\d{2})\s*$", s)
    if m:
        return m.group(1)
    m = re.search(r"(19\d{2}|20\d{2})", s)
    return m.group(1) if m else ""


def _parse_any_date(value) -> date | None:
    """Parse PDF ngay_co_kq / Medinet NgayKham / filename DDMMYY → date."""
    s = str(value or "").strip()
    if not s:
        return None
    # ISO / Medinet: 2026-07-23 or 2026-07-23T00:00:00
    m = re.match(r"^(20\d{2}|19\d{2})-(\d{2})-(\d{2})", s)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    # DMY: 23/07/2026 or 23/07/2026 14:30
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(20\d{2}|19\d{2})", s)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            return None
    # Filename prefix DDMMYY
    m = re.match(r"^(\d{2})(\d{2})(\d{2})\b", s)
    if m:
        dd, mm, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = 2000 + yy if yy < 100 else yy
        try:
            return date(year, mm, dd)
        except ValueError:
            return None
    return None


def _date_proximity_score(pdf_d: date | None, rec_d: date | None) -> int:
    """Score how well PDF result-print date fits Medinet NgayKham.

    Lab may be printed BEFORE khám (ngay_co_kq earlier than NgayKham).
    Allow exam up to 45 days after print, or print up to 7 days after exam.
    """
    if not pdf_d or not rec_d:
        return 0
    delta = (rec_d - pdf_d).days  # + = khám after print
    if -7 <= delta <= 3:
        return 3  # same day / nearly same
    if 0 <= delta <= 14:
        return 2  # print before exam within 2 weeks
    if -7 <= delta <= 45:
        return 1  # still plausible
    return 0  # too far — demote heavily when disambiguating


def _names_soft_match(a: str, b: str) -> bool:
    """True when folded names are the same person (strict on họ + tên).

    Rejects subset traps like "NGUYEN THI KIEU" ⊂ "NGUYEN THI KIEU DIEM"
    which previously caused false TTHC matches → wrong PROCESSED.
    """
    fa, fb = _fold_name(a), _fold_name(b)
    if not fa or not fb:
        return False
    if fa == fb:
        return True
    ta, tb = fa.split(), fb.split()
    if not ta or not tb:
        return False
    if sorted(ta) == sorted(tb):
        return True
    # Vietnamese identity: họ (first) + tên (last) must both match.
    # Extra middle tokens OK (OCR drop "THI"); different LAST token = other person.
    if len(ta) < 2 or len(tb) < 2:
        return False
    if ta[0] != tb[0] or ta[-1] != tb[-1]:
        return False
    return True


def fetch_unit_index(token: str, date_from: str, date_to: str) -> dict:
    """Build lookup by normalized name+phone and SID-ish MaPhieu for M3/M4."""
    reports = [
        ("M3", "KSKDK_DanhSach_KSK_M13", "NgayTao"),
        ("M4", "KSKDK_DanhSach_KSK_NguoiCaoTuoi_Report", "KSKDK_NgayKham"),
    ]
    index = {
        "by_phone": {},
        "by_name_year": {},
        "by_fold_year": {},
        "by_cccd": {},
        "by_maphieu": {},
        "by_pid": {},
        "no_cls_ids": set(),
        "all_ids": set(),
    }

    def get_report(code):
        s, d = api(token, f"/api/services/app/DRReport/GetIdByCode?Code={code}&SessionSiteId=130")
        items = ((d or {}).get("result") or {}).get("data") or []
        return items[0] if items else None

    from datetime import date, timedelta

    def parse_d(s):
        dd, mm, yy = s.split("/")
        return date(int(yy), int(mm), int(dd))

    d0, d1 = parse_d(date_from), parse_d(date_to)
    # Widen start: M3 list filters NgayTao — phiếu tạo trước khoảng khám vẫn cần index
    d0 = d0 - timedelta(days=60)
    days = []
    cur = d0
    while cur <= d1:
        days.append(cur)
        cur += timedelta(days=1)
    safe_print(f"  Index day span {d0.strftime('%d/%m/%Y')} -> {d1.strftime('%d/%m/%Y')} ({len(days)} days)")

    def _cccd_of(r: dict) -> str:
        for k in (
            "SoDinhDanh",
            "CCCD",
            "CMND",
            "SoCMND",
            "MaDinhDanh",
            "DinhDanhCaNhan",
            "SoDinhDanhCaNhan",
        ):
            v = re.sub(r"\D", "", str(r.get(k) or ""))
            if len(v) >= 9:
                return v
        return ""

    for mau, code, date_field in reports:
        rep = get_report(code)
        if not rep:
            safe_print(f"  report missing {code}")
            continue
        store, ds = rep["sqlContent"], rep["dataSourceId"]
        safe_print(f"Indexing {mau} ({code}) ...", flush=True)
        for day in days:
            dr = f"{day.strftime('%d/%m/%Y')} - {day.strftime('%d/%m/%Y')}"
            s, d = api(
                token,
                f"/api/services/app/DRViewer/ExecuteStoreWithParam_ByDatasource?dataSourceId={ds}&store={store}",
                "POST",
                to_fparams(
                    {
                        date_field: dr,
                        "NgayTao": dr,
                        "KSKDK_NgayKham": dr,
                        "page": 1,
                        "pageSize": 5000,
                    }
                ),
            )
            rows = ((d or {}).get("result") or {}).get("data") or []
            for r in rows:
                pid = r.get("phieukhamId") or r.get("Id")
                index["all_ids"].add(pid)
                phone = re.sub(r"\D", "", str(r.get("SDT") or ""))
                name = (r.get("HoTen") or "").strip().upper()
                year = _year_from_ngaysinh(r.get("NgaySinh"))
                mp = str(r.get("MaPhieu") or "")
                rec = {**r, "_mau": mau}
                if phone:
                    index["by_phone"].setdefault(phone, []).append(rec)
                if name and year:
                    index["by_name_year"].setdefault(f"{name}|{year}", []).append(rec)
                    index["by_fold_year"].setdefault(f"{_fold_name(name)}|{year}", []).append(rec)
                cccd = _cccd_of(r)
                if cccd:
                    index["by_cccd"][cccd] = rec
                if mp:
                    index["by_maphieu"][mp] = rec
                if pid not in (None, ""):
                    index["by_pid"][str(pid)] = rec

            s, d = api(
                token,
                f"/api/services/app/DRViewer/ExecuteStoreWithParam_ByDatasource?dataSourceId={ds}&store={store}",
                "POST",
                to_fparams(
                    {
                        date_field: dr,
                        "NgayTao": dr,
                        "KSKDK_NgayKham": dr,
                        "ChatLuongDuLieu": 4,
                        "page": 1,
                        "pageSize": 5000,
                    }
                ),
            )
            for r in ((d or {}).get("result") or {}).get("data") or []:
                index["no_cls_ids"].add(r.get("phieukhamId") or r.get("Id"))
        safe_print(
            f"  {mau} indexed phones={len(index['by_phone'])} names={len(index['by_name_year'])} "
            f"fold={len(index['by_fold_year'])} cccd={len(index['by_cccd'])}",
            flush=True,
        )
    return index


def match_patient(row: dict, index: dict) -> tuple[str, dict | None]:
    """Return (status, medinet_row).

    Hard keys:
      - nam_sinh (năm sinh) MUST match
      - họ + tên (strict soft: same first+last token; no subset names)
    Soft key:
      - ngày có kết quả (PDF) vs NgayKham (Medinet): may differ because lab
        can be printed BEFORE khám — allow exam up to ~45 days after print.
    """
    phone = re.sub(r"\D", "", str(row.get("sdt") or ""))
    name = (row.get("ho_ten") or "").strip().upper()
    year = str(row.get("nam_sinh") or "").strip()
    sid = str(row.get("sid") or row.get("ma_phieu") or "")
    cccd = re.sub(r"\D", "", str(row.get("cccd") or ""))
    fname = str(row.get("file_name") or row.get("source_file") or "")
    stem = Path(fname).stem if fname else ""

    # PDF result-print date: header "Ngày có kết quả" or filename DDMMYY
    pdf_result_d = _parse_any_date(row.get("ngay_co_kq"))
    if not pdf_result_d:
        pdf_result_d = _parse_any_date(stem[:6] if re.match(r"^\d{6}-", stem) else "")

    fn_name, fn_year = "", ""
    m_fn = re.search(
        r"^\d{6}-\d+\s*-\s*(.+?)\s*-\s*(19\d{2}|20\d{2})\s*-\s*[MF](?:_|\.|$)",
        stem,
        re.I,
    )
    if not m_fn:
        m_fn = re.search(
            r"^\d{6}-\d+\s*-\s*(.+?)\s*-\s*(19\d{2}|20\d{2})\s*-\s*[MF]\b",
            stem,
            re.I,
        )
    if m_fn:
        fn_name = m_fn.group(1).strip().upper()
        fn_year = m_fn.group(2)
        if not name:
            name = fn_name
        # Prefer filename year when present (stable on INBOX names)
        if fn_year:
            year = fn_year
        elif not year:
            year = fn_year

    yr_target = year or fn_year

    def _rec_year(rec: dict) -> str:
        return _year_from_ngaysinh(rec.get("NgaySinh"))

    def _year_ok(rec: dict) -> bool:
        if not yr_target:
            return False
        return _rec_year(rec) == yr_target

    def _rec_ngaykham(rec: dict) -> date | None:
        return _parse_any_date(rec.get("NgayKham") or rec.get("KSKDK_NgayKham") or rec.get("NgayTao"))

    strong = []  # CCCD / MaPhieu / explicit pid — still verify year when we have one
    if cccd and cccd in index.get("by_cccd", {}):
        strong.append(index["by_cccd"][cccd])
    for mp in filter(None, [sid, row.get("ma_phieu")]):
        mp = str(mp).strip()
        if mp and mp in index["by_maphieu"]:
            strong.append(index["by_maphieu"][mp])
        if mp and mp in index.get("by_pid", {}):
            strong.append(index["by_pid"][mp])
    for m in re.finditer(r"KSKDKP\d+", stem, re.I):
        mp = m.group(0).upper()
        if mp in index["by_maphieu"]:
            strong.append(index["by_maphieu"][mp])
    for m in re.finditer(r"_(\d{5,7})(?:_|\.|$)", stem):
        token = m.group(1)
        if token in index.get("by_pid", {}):
            strong.append(index["by_pid"][token])

    candidates = []
    # Phone: only keep same birth year when year known (avoid wrong twin-name)
    if phone and phone in index["by_phone"]:
        for rec in index["by_phone"][phone]:
            if not yr_target or _year_ok(rec):
                candidates.append(rec)

    # Name+year only (required year — never name-only across 1950 vs 1953)
    if yr_target:
        for nm in filter(None, [name, fn_name]):
            key = f"{nm}|{yr_target}"
            if key in index["by_name_year"]:
                candidates.extend(index["by_name_year"][key])
            fk = f"{_fold_name(nm)}|{yr_target}"
            if fk in index.get("by_fold_year", {}):
                candidates.extend(index["by_fold_year"][fk])
            for k, recs in index.get("by_fold_year", {}).items():
                kn, ky = (k.split("|", 1) + [""])[:2]
                if ky == yr_target and _names_soft_match(nm, kn):
                    candidates.extend(recs)

    # Lab SID in "DDMMYY-SID - NAME - YEAR - M/F" is NOT phieukhamId
    lab_sid = ""
    m_lab = re.match(r"^(\d{6})-(\d+)\b", stem)
    if m_lab:
        lab_sid = m_lab.group(2)

    digit_hits = []
    for m in re.finditer(r"(?<!\d)(\d{6,7})(?!\d)", stem):
        token = m.group(1)
        if re.match(r"^\d{6}-\d+", stem) and token == stem.split("-", 1)[0][:6]:
            continue
        if lab_sid and token == lab_sid:
            continue
        if token in index.get("by_pid", {}):
            digit_hits.append(index["by_pid"][token])
        if token in index["by_maphieu"]:
            digit_hits.append(index["by_maphieu"][token])

    fold_target = name or fn_name
    for rec in digit_hits:
        if yr_target and not _year_ok(rec):
            continue
        if fold_target and not _names_soft_match(fold_target, str(rec.get("HoTen") or "")):
            continue
        candidates.append(rec)

    # Strong keys: prefer year agreement; if year known and conflicts → drop
    for rec in strong:
        if yr_target and not _year_ok(rec):
            # Unique id with conflicting year: still trust id (CCCD/MaPhieu/pid)
            # only when name also soft-matches or no name available
            rn = str(rec.get("HoTen") or "")
            if fold_target and rn and not _names_soft_match(fold_target, rn):
                continue
        candidates.append(rec)

    seen = set()
    uniq = []
    for c in candidates:
        pid = c.get("phieukhamId") or c.get("Id")
        if pid in seen:
            continue
        seen.add(pid)
        uniq.append(c)

    if not uniq:
        return "WAITING_ADMIN", None

    # HARD FILTER: when PDF/filename has nam_sinh, only same-year candidates
    if yr_target:
        same_year = [c for c in uniq if _year_ok(c)]
        if same_year:
            uniq = same_year
        else:
            # No year agreement → do not guess among same-name different years
            return "WAITING_ADMIN", None

    def _score(c: dict) -> tuple:
        rn = str(c.get("HoTen") or "")
        ns = _rec_year(c)
        name_ok = 2 if fold_target and _fold_name(fold_target) == _fold_name(rn) else (
            1 if fold_target and _names_soft_match(fold_target, rn) else 0
        )
        year_ok = 2 if yr_target and ns == yr_target else 0
        phone_ok = 1 if phone and re.sub(r"\D", "", str(c.get("SDT") or "")) == phone else 0
        date_ok = _date_proximity_score(pdf_result_d, _rec_ngaykham(c))
        mau = row.get("mau_kham")
        mau_ok = 1 if mau and c.get("_mau") == mau else 0
        # Primary: year+name+phone+date; secondary: mau
        return (year_ok + name_ok + phone_ok + date_ok, mau_ok, date_ok)

    uniq.sort(key=_score, reverse=True)
    # Require name soft-match when we have a name (after year filter)
    if fold_target:
        named = [c for c in uniq if _names_soft_match(fold_target, str(c.get("HoTen") or ""))]
        if named:
            uniq = named
        elif not cccd and not any(
            str(row.get("ma_phieu") or "") and str(row.get("ma_phieu")) in index.get("by_maphieu", {})
            for _ in [0]
        ):
            # keep strong pid-only hits already year-filtered
            pass

    # Prefer exact folded name over soft (ho+ten) matches
    if fold_target:
        exact = [
            c
            for c in uniq
            if _fold_name(fold_target) == _fold_name(str(c.get("HoTen") or ""))
        ]
        if exact:
            uniq = exact

    # When several same name+year: use ngày in KQ vs NgayKham to pick / reject
    if len(uniq) > 1 and pdf_result_d:
        dated = [c for c in uniq if _date_proximity_score(pdf_result_d, _rec_ngaykham(c)) > 0]
        if len(dated) == 1:
            uniq = dated
        elif len(dated) > 1:
            dated.sort(key=_score, reverse=True)
            top = _score(dated[0])
            tied = [c for c in dated if _score(c) == top]
            if len(tied) == 1:
                uniq = tied
            elif phone or cccd:
                uniq = tied
            else:
                return "WAITING_ADMIN", None
        elif not phone and not cccd:
            # Multiple name+year but NONE near PDF print date → do not guess
            return "WAITING_ADMIN", None

    if len(uniq) > 1:
        # Still ambiguous after year+name(+date) — need phone/cccd; else wait
        top = _score(uniq[0])
        tied = [c for c in uniq if _score(c) == top]
        if len(tied) > 1 and not phone and not cccd:
            return "WAITING_ADMIN", None
        uniq = tied

    mau = row.get("mau_kham")
    preferred = [c for c in uniq if c.get("_mau") == mau] or uniq
    preferred.sort(key=_score, reverse=True)
    rec = preferred[0]

    # Single candidate with known PDF date but NgayKham far away: still accept
    # (in trước / khám sau can exceed window for unique name+year). Only block
    # when we had to choose among multiples (handled above).

    pid = rec.get("phieukhamId") or rec.get("Id")
    if pid not in index["no_cls_ids"]:
        return "SKIP_ALREADY_CLS", rec
    return "READY_IMPORT", rec


def search_patient_live(
    token: str,
    *,
    name: str,
    year: str,
    date_from: str,
    date_to: str,
    ngay_co_kq: str = "",
) -> tuple[str, dict | None, str]:
    """Fallback when day-index miss: query M3/M4 by HoTen over full date span.

    Disambiguate same name+year via ngày có kết quả vs NgayKham (print may
    precede exam). Returns (status, rec, token).
    """
    if not name or not year:
        return "WAITING_ADMIN", None, token

    reports = [
        ("M3", "KSKDK_DanhSach_KSK_M13", "NgayTao"),
        ("M4", "KSKDK_DanhSach_KSK_NguoiCaoTuoi_Report", "KSKDK_NgayKham"),
    ]
    fold = _fold_name(name)
    pdf_d = _parse_any_date(ngay_co_kq)
    hits = []

    def get_report(code):
        s, d = api(token, f"/api/services/app/DRReport/GetIdByCode?Code={code}&SessionSiteId=130")
        items = ((d or {}).get("result") or {}).get("data") or []
        return items[0] if items else None

    dr = f"{date_from} - {date_to}"
    for mau, code, date_field in reports:
        rep = get_report(code)
        if not rep:
            continue
        store, ds = rep["sqlContent"], rep["dataSourceId"]
        for ho in filter(None, [name, fold]):
            s, d = api(
                token,
                f"/api/services/app/DRViewer/ExecuteStoreWithParam_ByDatasource?dataSourceId={ds}&store={store}",
                "POST",
                to_fparams(
                    {
                        date_field: dr,
                        "NgayTao": dr,
                        "KSKDK_NgayKham": dr,
                        "HoTen": ho,
                        "page": 1,
                        "pageSize": 200,
                    }
                ),
            )
            for r in ((d or {}).get("result") or {}).get("data") or []:
                if _year_from_ngaysinh(r.get("NgaySinh")) != str(year):
                    continue
                if not _names_soft_match(name, str(r.get("HoTen") or "")):
                    continue
                hits.append({**r, "_mau": mau})

    seen = set()
    uniq = []
    for h in hits:
        pid = h.get("phieukhamId") or h.get("Id")
        if pid in seen:
            continue
        seen.add(pid)
        uniq.append(h)
    if not uniq:
        return "WAITING_ADMIN", None, token

    # Prefer exact folded name
    exact = [h for h in uniq if _fold_name(name) == _fold_name(str(h.get("HoTen") or ""))]
    if exact:
        uniq = exact

    if len(uniq) > 1 and pdf_d:
        dated = [
            h
            for h in uniq
            if _date_proximity_score(
                pdf_d, _parse_any_date(h.get("NgayKham") or h.get("KSKDK_NgayKham"))
            )
            > 0
        ]
        if len(dated) == 1:
            uniq = dated
        elif len(dated) > 1:
            dated.sort(
                key=lambda h: _date_proximity_score(
                    pdf_d, _parse_any_date(h.get("NgayKham") or h.get("KSKDK_NgayKham"))
                ),
                reverse=True,
            )
            best = _date_proximity_score(
                pdf_d, _parse_any_date(dated[0].get("NgayKham") or dated[0].get("KSKDK_NgayKham"))
            )
            tied = [
                h
                for h in dated
                if _date_proximity_score(
                    pdf_d, _parse_any_date(h.get("NgayKham") or h.get("KSKDK_NgayKham"))
                )
                == best
            ]
            if len(tied) == 1:
                uniq = tied
            else:
                return "WAITING_ADMIN", None, token
        else:
            return "WAITING_ADMIN", None, token

    if len(uniq) > 1:
        return "WAITING_ADMIN", None, token
    rec = uniq[0]
    return "READY_IMPORT", rec, token


def write_preview_excel(rows: list[dict], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Preview_CLS"
    header = [
        "STT",
        "file_name",
        "sid",
        "ho_ten",
        "nam_sinh",
        "gioi_tinh",
        "sdt",
        "mau_kham",
        "ngay_co_kq",
        "status_medinet",
        "medinet_MaPhieu",
        "medinet_NgayKham",
        "medinet_phieukhamId",
        "medinet_cdId",
        "parse_ok",
    ]
    for lab in LAB_COLS:
        header += [f"{lab}_raw", f"{lab}_unit_raw", f"{lab}_web", f"{lab}_unit_web", f"{lab}_note"]

    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(color="FFFFFF", bold=True)
    for c, h in enumerate(header, 1):
        cell = ws.cell(1, c, h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(wrap_text=True, horizontal="center")

    for i, row in enumerate(rows, 1):
        labs = row.get("labs") or {}
        vals = [
            i,
            row.get("file_name"),
            row.get("sid"),
            row.get("ho_ten"),
            row.get("nam_sinh"),
            row.get("gioi_tinh"),
            row.get("sdt"),
            row.get("mau_kham"),
            row.get("ngay_co_kq"),
            row.get("status_medinet"),
            row.get("medinet_MaPhieu"),
            row.get("medinet_NgayKham"),
            row.get("medinet_phieukhamId"),
            row.get("medinet_cdId"),
            "YES" if row.get("parse_ok") else "NO",
        ]
        for lab in LAB_COLS:
            item = labs.get(lab) or {}
            vals += [
                item.get("value_raw", ""),
                item.get("unit_raw", ""),
                item.get("value_web", ""),
                item.get("unit_web", ""),
                item.get("convert_note", ""),
            ]
        for c, v in enumerate(vals, 1):
            ws.cell(1 + i, c, v)

    ws2 = wb.create_sheet("Missing_or_Updated")
    h2 = ["STT", "file_name", "sid", "ho_ten", "nam_sinh", "sdt", "mau_kham", "status_medinet", "reason"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(1, c, h)
        cell.fill = head_fill
        cell.font = head_font
    r = 2
    stt = 1
    for row in rows:
        st = row.get("status_medinet")
        if st in ("WAITING_ADMIN", "SKIP_ALREADY_CLS", "PARSE_ERROR"):
            reason = {
                "WAITING_ADMIN": "Chưa thấy TTHC trên Medinet",
                "SKIP_ALREADY_CLS": "Đã có CLS trên web — cần kiểm tra trước khi ghi đè",
                "PARSE_ERROR": "Không đọc được PDF đủ thông tin",
            }.get(st, st)
            for c, v in enumerate(
                [
                    stt,
                    row.get("file_name"),
                    row.get("sid"),
                    row.get("ho_ten"),
                    row.get("nam_sinh"),
                    row.get("sdt"),
                    row.get("mau_kham"),
                    st,
                    reason,
                ],
                1,
            ):
                ws2.cell(r, c, v)
            r += 1
            stt += 1

    wb.save(path)


def update_cases_csv(cases_path: Path, rows: list[dict]) -> None:
    if not cases_path.exists():
        return
    with cases_path.open(encoding="utf-8-sig", newline="") as f:
        existing = list(csv.DictReader(f))
    by_file = {}
    for r in rows:
        by_file[Path(r.get("source_file") or r.get("file_name") or "").name] = r
        by_file[r.get("file_name")] = r

    for e in existing:
        src = Path(e.get("source_file") or "").name
        hit = by_file.get(src)
        if not hit:
            continue
        e["status"] = hit.get("status_medinet") or e.get("status")
        e["ho_ten"] = e.get("ho_ten") or hit.get("ho_ten")
        e["notes"] = f"phase_b:{hit.get('status_medinet')}"
        e["last_checked_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if hit.get("medinet_MaPhieu"):
            e["ma_phieu"] = hit["medinet_MaPhieu"]

    if existing:
        with cases_path.open("w", encoding="utf-8", newline="") as f:
            w = csv.DictWriter(f, fieldnames=list(existing[0].keys()))
            w.writeheader()
            w.writerows(existing)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0, help="Limit PDFs for test (0=all)")
    ap.add_argument("--skip-medinet", action="store_true", help="Only parse PDFs, no Medinet match")
    ap.add_argument("--inbox", default="", help="Override inbox folder")
    args = ap.parse_args()

    cfg = load_config()
    build = build_root(cfg)
    inbox = Path(args.inbox) if args.inbox else inbox_dir(cfg)
    safe_print(f"Inbox: {inbox}")
    safe_print(f"Build: {build}")

    pdfs = list_pdfs(inbox, args.limit or None)
    safe_print(f"PDF count: {len(pdfs)}", flush=True)
    if not pdfs:
        safe_print("No PDFs found.")
        return 1

    rows = []
    for i, p in enumerate(pdfs, 1):
        try:
            data = extract_pdf(p)
        except Exception as e:
            data = {
                "source_file": str(p),
                "file_name": p.name,
                "parse_ok": False,
                "labs": {},
                "status_medinet": "PARSE_ERROR",
                "notes": str(e),
            }
        if not data.get("parse_ok"):
            data["status_medinet"] = data.get("status_medinet") or "PARSE_ERROR"
        rows.append(data)
        if i % 50 == 0 or i == len(pdfs):
            safe_print(f"  parsed {i}/{len(pdfs)}", flush=True)

    index = None
    if not args.skip_medinet:
        import os

        from medinet_creds import get_medinet_creds

        user, password = get_medinet_creds(cfg)
        safe_print("Auth Medinet + index July lists...", flush=True)
        token = authenticate(user, password)
        date_from = cfg.get("medinet", {}).get("date_from") or "01/07/2026"
        date_to = (cfg.get("medinet", {}).get("date_to") or "").strip()
        if not date_to:
            from datetime import date as _date

            date_to = _date.today().strftime("%d/%m/%Y")
        index = fetch_unit_index(token, date_from, date_to)
        for row in rows:
            if row.get("status_medinet") == "PARSE_ERROR":
                continue
            st, rec = match_patient(row, index)
            row["status_medinet"] = st
            if rec:
                row["medinet_MaPhieu"] = rec.get("MaPhieu")
                nk = rec.get("NgayKham") or ""
                row["medinet_NgayKham"] = str(nk).split("T")[0]
                row["medinet_phieukhamId"] = rec.get("phieukhamId") or rec.get("Id")
                row["medinet_cdId"] = rec.get("cdId") or ""
            else:
                row["medinet_MaPhieu"] = ""
                row["medinet_NgayKham"] = ""
                row["medinet_phieukhamId"] = ""
                row["medinet_cdId"] = ""
    else:
        for row in rows:
            row.setdefault("status_medinet", "NOT_CHECKED")

    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    out = build / "excel_preview" / f"CLS_preview_{stamp}.xlsx"
    write_preview_excel(rows, out)
    # also copy missing sheet-only workbook
    missing_rows = [r for r in rows if r.get("status_medinet") in ("WAITING_ADMIN", "SKIP_ALREADY_CLS", "PARSE_ERROR")]
    miss_path = build / "missing_or_updated" / f"missing_or_updated_{stamp}.xlsx"
    write_preview_excel(missing_rows, miss_path)

    cases_path = ROOT / cfg.get("tracking", {}).get("cases_csv", "tracking/cases.csv")
    update_cases_csv(cases_path, rows)

    # summary
    from collections import Counter

    c = Counter(r.get("status_medinet") for r in rows)
    safe_print("---")
    safe_print(f"Preview Excel: {out}")
    safe_print(f"Missing/Updated Excel: {miss_path}")
    safe_print(f"Status: {dict(c)}")
    safe_print("NEXT: open Preview Excel, check value_web/unit_web. Then run import step.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
