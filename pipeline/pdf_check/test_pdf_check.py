"""Unit tests for pipeline/pdf_check (no Medinet / no G: required)."""

from __future__ import annotations

import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
PIPE = ROOT / "pipeline"
sys.path.insert(0, str(PIPE))
sys.path.insert(0, str(PIPE / "pdf_check"))


def test_scan_and_dedup():
    from pdf_check.dedup import mark_duplicates
    from pdf_check.scan_pdfs import scan_pipeline_pdfs

    with tempfile.TemporaryDirectory() as td:
        root = Path(td)
        for name in ("MISSING", "PROCESSED", "INBOX_CLS"):
            (root / name).mkdir()
        pdf_a = root / "MISSING" / "140826-1 - NGUYEN KHAC HOANG - 1990 - M.pdf"
        pdf_b = root / "PROCESSED" / "140826-1 - NGUYEN KHAC HOANG - 1990 - M.pdf"
        pdf_c = root / "INBOX_CLS" / "other.pdf"
        pdf_a.write_bytes(b"%PDF-1.4 same")
        pdf_b.write_bytes(b"%PDF-1.4 same")
        pdf_c.write_bytes(b"%PDF-1.4 other")

        scanned = scan_pipeline_pdfs(root, ("MISSING", "PROCESSED", "INBOX_CLS"))
        assert len(scanned) == 3
        marked = mark_duplicates(scanned, compute_hash=True)
        dups = [r for r in marked if r["is_dup_name"] == "YES"]
        assert len(dups) == 2
        assert "MISSING" in dups[0]["dup_folders"] and "PROCESSED" in dups[0]["dup_folders"]
        hash_dups = [r for r in marked if r.get("same_hash_dup") == "YES"]
        assert len(hash_dups) == 2


def test_suggest_folder_and_cls_summary():
    from pdf_check.check_tthc_cls import _cls_summary, suggest_folder

    assert (
        suggest_folder(
            match_status="READY",
            tthc_scope="TK2",
            coverage="FULL",
            sample_kind="BLOOD_URINE",
            nam_sinh="1986",
            file_name="x.pdf",
            primary_account="pkdk_Thuankieu",
        )
        == "TK2"
    )
    assert (
        suggest_folder(
            match_status="READY",
            tthc_scope="BOTH",
            coverage="FULL",
            sample_kind="BLOOD_URINE",
            nam_sinh="1986",
            file_name="x.pdf",
            primary_account="pkdkthuankieu",
        )
        == "PROCESSED"
    )
    assert (
        suggest_folder(
            match_status="NO_TTHC",
            tthc_scope="NONE",
            coverage="FULL",
            sample_kind="BLOOD_URINE",
            nam_sinh="1986",
            file_name="x.pdf",
            primary_account="",
        )
        == "MISSING"
    )

    assert (
        _cls_summary(
            match_status="READY",
            tthc_scope="BOTH",
            cls_tk1="YES",
            cls_tk2="NO",
        )
        == "PARTIAL_CLS"
    )
    assert (
        _cls_summary(
            match_status="READY",
            tthc_scope="TK2",
            cls_tk1="N/A",
            cls_tk2="YES",
        )
        == "HAS_CLS_ONE"
    )
    assert (
        _cls_summary(
            match_status="READY",
            tthc_scope="BOTH",
            cls_tk1="NO",
            cls_tk2="NO",
        )
        == "NEED_CLS"
    )


def test_write_excel_sheets():
    from pdf_check.write_excel import write_pdf_check_excel

    rows = [
        {
            "folder": "PROCESSED",
            "file_name": "a.pdf",
            "path": r"G:\x\PROCESSED\a.pdf",
            "ho_ten": "TRAN VINH HUNG",
            "nam_sinh": "1986",
            "match_status": "READY",
            "tthc_scope": "TK2",
            "tthc_tk1": "NO",
            "tthc_tk2": "YES",
            "cls_tk1": "N/A",
            "cls_tk2": "YES",
            "cls_summary": "HAS_CLS_ONE",
            "folder_nen": "TK2",
            "is_dup_name": "NO",
        },
        {
            "folder": "MISSING",
            "file_name": "b.pdf",
            "path": r"G:\x\MISSING\b.pdf",
            "ho_ten": "NGUYEN A",
            "match_status": "READY",
            "tthc_scope": "BOTH",
            "tthc_tk1": "YES",
            "tthc_tk2": "YES",
            "cls_tk1": "NO",
            "cls_tk2": "NO",
            "cls_summary": "NEED_CLS",
            "folder_nen": "PROCESSED",
            "is_dup_name": "YES",
            "dup_folders": "MISSING|PROCESSED",
        },
    ]
    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "PDF_CHECK_test.xlsx"
        write_pdf_check_excel(rows, out)
        assert out.exists() and out.stat().st_size > 1000
        from openpyxl import load_workbook

        wb = load_workbook(out)
        assert set(wb.sheetnames) >= {
            "All",
            "Dup",
            "NeedCLS",
            "HasCLS",
            "NoTTHC",
            "Ambiguous",
            "FolderMismatch",
            "Summary",
        }


def test_no_insert_cls_import_in_pdf_check():
    """pdf_check runtime modules must not call insert_cls."""
    root = Path(__file__).resolve().parents[1] / "pdf_check"
    for p in root.glob("*.py"):
        if p.name.startswith("test_"):
            continue
        text = p.read_text(encoding="utf-8")
        assert "insert_cls(" not in text, p.name


def test_ps1_ascii():
    p = Path(__file__).resolve().parents[1] / "CHAY_PDF_CHECK.ps1"
    p.read_bytes().decode("ascii")


if __name__ == "__main__":
    test_scan_and_dedup()
    test_suggest_folder_and_cls_summary()
    test_write_excel_sheets()
    test_no_insert_cls_import_in_pdf_check()
    test_ps1_ascii()
    print("OK: pdf_check tests passed")
