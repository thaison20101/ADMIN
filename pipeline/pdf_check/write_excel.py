"""Write multi-sheet PDF_CHECK Excel report — tiếng Việt có dấu."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

# Internal keys → cột tiếng Việt
COL_VI = {
    "folder": "Folder",
    "file_name": "Tên file",
    "path": "Đường dẫn",
    "ho_ten": "Họ tên",
    "nam_sinh": "Năm sinh",
    "ngay_sinh": "Ngày sinh",
    "sdt": "SĐT",
    "cccd": "CCCD",
    "pdf_coverage": "Độ phủ PDF",
    "sample_kind": "Loại mẫu",
    "parse_ok": "Parse OK",
    "match_status": "Trạng thái match",
    "match_mode": "Chế độ match",
    "tthc_tk1": "TTHC TK1",
    "tthc_tk2": "TTHC TK2",
    "tthc_scope": "Phạm vi TTHC",
    "pid_tk1": "PID TK1",
    "pid_tk2": "PID TK2",
    "maphieu_tk1": "Mã phiếu TK1",
    "maphieu_tk2": "Mã phiếu TK2",
    "cls_tk1": "CLS TK1",
    "cls_tk2": "CLS TK2",
    "cls_summary": "Tóm tắt CLS",
    "folder_nen": "Folder nên",
    "is_dup_name": "Trùng tên",
    "dup_folders": "Folder trùng",
    "dup_count": "Số bản trùng",
    "same_hash_dup": "Trùng hash",
    "hash_dup_folders": "Folder hash trùng",
    "file_hash": "Hash file",
}

ALL_COLS = list(COL_VI.keys())
HEADER_VI = [COL_VI[k] for k in ALL_COLS]

SCOPE_VI = {"BOTH": "Cả 2 TK", "TK1": "TK1", "TK2": "TK2", "NONE": "Không có"}


def _row_vi(r: dict[str, Any]) -> list[Any]:
    out: list[Any] = []
    for k in ALL_COLS:
        v = r.get(k, "")
        if k == "tthc_scope":
            v = SCOPE_VI.get(str(v), v)
        out.append(v)
    return out


def _write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    ws.append(HEADER_VI)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append(_row_vi(r))


def write_pdf_check_excel(rows: list[dict[str, Any]], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "Tất cả"
    _write_sheet(ws_all, rows)

    dup = [r for r in rows if str(r.get("is_dup_name") or "") == "YES"]
    ws_dup = wb.create_sheet("Trùng tên")
    _write_sheet(ws_dup, dup)

    need = [
        r
        for r in rows
        if str(r.get("cls_summary") or "") in {"NEED_CLS", "PARTIAL_CLS"}
    ]
    ws_need = wb.create_sheet("Cần điền CLS")
    _write_sheet(ws_need, need)

    has = [
        r
        for r in rows
        if str(r.get("cls_summary") or "") in {"HAS_CLS_ONE", "HAS_CLS_BOTH"}
    ]
    ws_has = wb.create_sheet("Đã có CLS")
    _write_sheet(ws_has, has)

    no_tthc = [
        r
        for r in rows
        if str(r.get("cls_summary") or "") in {"NO_TTHC", "PARSE_FAIL"}
        or str(r.get("match_status") or "") in {"NO_TTHC", "PARSE_FAIL"}
    ]
    ws_miss = wb.create_sheet("Không TTHC")
    _write_sheet(ws_miss, no_tthc)

    amb = [r for r in rows if str(r.get("match_status") or "") == "AMBIGUOUS"]
    ws_amb = wb.create_sheet("Mơ hồ")
    _write_sheet(ws_amb, amb)

    mismatch = [
        r
        for r in rows
        if r.get("folder_nen")
        and r.get("folder")
        and str(r.get("folder")) != str(r.get("folder_nen"))
        and str(r.get("match_status") or "") == "READY"
    ]
    ws_mm = wb.create_sheet("Sai folder")
    _write_sheet(ws_mm, mismatch)

    ws_sum = wb.create_sheet("Tóm tắt")
    ws_sum.append(["Chỉ số", "Giá trị"])
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["B1"].font = Font(bold=True)
    ws_sum.append(["Tổng PDF", len(rows)])
    ws_sum.append(["Trùng tên", len(dup)])
    ws_sum.append(["Cần điền CLS", len(need)])
    ws_sum.append(["Đã có CLS", len(has)])
    ws_sum.append(["Không TTHC", len(no_tthc)])
    ws_sum.append(["Mơ hồ", len(amb)])
    ws_sum.append(["Sai folder", len(mismatch)])
    ws_sum.append([])
    ws_sum.append(["Tóm tắt CLS", "Số lượng"])
    for k, v in Counter(str(r.get("cls_summary") or "") for r in rows).most_common():
        ws_sum.append([k, v])
    ws_sum.append([])
    ws_sum.append(["Phạm vi TTHC", "Số lượng"])
    for k, v in Counter(str(r.get("tthc_scope") or "") for r in rows).most_common():
        ws_sum.append([SCOPE_VI.get(k, k), v])
    ws_sum.append([])
    ws_sum.append(["Folder", "Số lượng"])
    for k, v in Counter(str(r.get("folder") or "") for r in rows).most_common():
        ws_sum.append([k, v])

    wb.save(out_path)
    return out_path
