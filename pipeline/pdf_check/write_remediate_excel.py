"""Excel báo cáo remediation — tiếng Việt có dấu."""

from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

ACTION_COLS = [
    "Tên file",
    "Họ tên",
    "Năm sinh",
    "Folder hiện tại",
    "Folder đích",
    "Phạm vi TTHC",
    "Độ phủ PDF",
    "CLS TK1",
    "CLS TK2",
    "filled_ok",
    "n_accts",
    "Hành động",
    "Kết quả",
    "Ghi chú",
    "Đường dẫn",
    "Đã xóa trùng",
]

SCOPE_VI = {
    "BOTH": "Cả 2 TK",
    "TK1": "TK1",
    "TK2": "TK2",
    "NONE": "Không có",
    "": "",
}

CLS_VI = {
    "YES": "Có",
    "NO": "Chưa",
    "N/A": "Không áp dụng",
    "SKIP": "Bỏ qua",
    "ERR": "Lỗi",
}


def _vi_scope(val: str) -> str:
    return SCOPE_VI.get(str(val or ""), str(val or ""))


def _vi_cls(val: str) -> str:
    s = str(val or "")
    if s.startswith("ERR"):
        return "Lỗi"
    return CLS_VI.get(s, s)


def _write_sheet(ws, rows: list[dict[str, Any]], cols: list[str]) -> None:
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])


def row_to_action_dict(r: dict[str, Any]) -> dict[str, Any]:
    deleted = r.get("dedup_deleted") or r.get("Đã xóa trùng") or ""
    if isinstance(deleted, list):
        deleted = "; ".join(str(x) for x in deleted)
    return {
        "Tên file": r.get("file_name") or r.get("Tên file") or "",
        "Họ tên": r.get("ho_ten") or r.get("Họ tên") or "",
        "Năm sinh": r.get("nam_sinh") or r.get("Năm sinh") or "",
        "Folder hiện tại": r.get("folder_from") or r.get("folder") or r.get("Folder hiện tại") or "",
        "Folder đích": r.get("folder_to") or r.get("Folder đích") or "",
        "Phạm vi TTHC": _vi_scope(str(r.get("tthc_scope") or "")),
        "Độ phủ PDF": r.get("pdf_coverage") or r.get("Độ phủ PDF") or "",
        "CLS TK1": _vi_cls(str(r.get("cls_tk1") or "")),
        "CLS TK2": _vi_cls(str(r.get("cls_tk2") or "")),
        "filled_ok": r.get("filled_ok", ""),
        "n_accts": r.get("n_accts", ""),
        "Hành động": r.get("action") or r.get("Hành động") or "",
        "Kết quả": r.get("result") or r.get("Kết quả") or "",
        "Ghi chú": r.get("notes") or r.get("Ghi chú") or "",
        "Đường dẫn": r.get("path_final") or r.get("path") or r.get("Đường dẫn") or "",
        "Đã xóa trùng": deleted,
    }


def write_remediate_excel(rows: list[dict[str, Any]], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    action_rows = [row_to_action_dict(r) for r in rows]
    wb = Workbook()

    ws_all = wb.active
    ws_all.title = "Tất cả"
    _write_sheet(ws_all, action_rows, ACTION_COLS)

    ws_detail = wb.create_sheet("Chi tiết hành động")
    _write_sheet(ws_detail, action_rows, ACTION_COLS)

    deduped = [r for r in action_rows if str(r.get("Đã xóa trùng") or "").strip()]
    ws_dup = wb.create_sheet("Trùng đã xóa")
    _write_sheet(ws_dup, deduped, ACTION_COLS)

    errors = [
        r
        for r in action_rows
        if str(r.get("Kết quả") or "") in {"Lỗi", "ERROR"}
        or "Lỗi" in str(r.get("Ghi chú") or "")
    ]
    ws_err = wb.create_sheet("Lỗi")
    _write_sheet(ws_err, errors, ACTION_COLS)

    ws_sum = wb.create_sheet("Tóm tắt")
    ws_sum.append(["Chỉ số", "Giá trị"])
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["B1"].font = Font(bold=True)
    ws_sum.append(["Tổng case", len(action_rows)])
    ws_sum.append(["Có xóa trùng", len(deduped)])
    ws_sum.append(["Lỗi", len(errors)])
    ws_sum.append([])
    ws_sum.append(["Folder đích", "Số lượng"])
    for k, v in Counter(str(r.get("Folder đích") or "") for r in action_rows).most_common():
        if k:
            ws_sum.append([k, v])
    ws_sum.append([])
    ws_sum.append(["Kết quả", "Số lượng"])
    for k, v in Counter(str(r.get("Kết quả") or "") for r in action_rows).most_common():
        if k:
            ws_sum.append([k, v])
    ws_sum.append([])
    ws_sum.append(["Hành động", "Số lượng"])
    for k, v in Counter(str(r.get("Hành động") or "") for r in action_rows).most_common():
        if k:
            ws_sum.append([k, v])

    wb.save(out_path)
    return out_path
