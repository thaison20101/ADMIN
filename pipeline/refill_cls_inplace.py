#!/usr/bin/env python3
"""Điền lại CLS từ PDF tại chỗ — không move file.

Ưu tiên folder:
  P. BÌNH TÂY - TRƯỜNG THCS NGUYỄN ĐỨC CẢNH - NGÀY 13-08-2026 - 165 CASE

  python pipeline/refill_cls_inplace.py
  python pipeline/refill_cls_inplace.py --apply
  python pipeline/refill_cls_inplace.py --toan-bo --apply
"""

from __future__ import annotations

import argparse
import re
import sys
import time
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
PIPE = Path(__file__).resolve().parent
sys.path.insert(0, str(PIPE))

from win_console import safe_print, setup_utf8_stdio  # noqa: E402

setup_utf8_stdio()

from drive_paths import g_pipeline_live, local_work_build, require_g_on_windows, resolve_g_sync  # noqa: E402
from medinet_api import (  # noqa: E402
    LAB_TO_FORM,
    authenticate,
    cls_has_lab_values,
    cls_missing_lab_fields,
    insert_cls,
    labs_to_form_payload,
    load_cls_view,
    verify_cls_saved,
)
from medinet_creds import MEDINET_ACCOUNTS  # noqa: E402
from pdf_extract import extract_pdf  # noqa: E402
from phase_b_preview import load_config, load_or_fetch_merged_unit_index, resolve_name_year  # noqa: E402
from single_instance import acquire_lock, release_lock  # noqa: E402
from tthc_match import ACCOUNT_TK1, ACCOUNT_TK2, resolve_tthc_matches  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

# Names to never fill (manual exclude). Fold-matched against PDF ho_ten / filename.
DEFAULT_SKIP_NAMES = (
    "TRAN DUY NHAT",
)
DEFAULT_FOLDER_HINTS = (
    "BINH TAY",
    "NGUYEN DUC CANH",
    "13-08-2026",
    "165 CASE",
)

# Exact folder name under sync root — refill this first (no move).
FIRST_FOLDER_NAMES = ("first", "First", "FIRST")

TOAN_BO_FOLDERS = (
    "PROCESSED",
    "TK1",
    "TK2",
    "UNDER 18",
    "ERROR",
    "INBOX_CLS",
)

ACTION_COLS = [
    "Tên file",
    "Họ tên",
    "Năm sinh",
    "Folder nguồn",
    "Phạm vi TTHC",
    "Trường trên PDF",
    "Thiếu trước",
    "Đã điền",
    "Kết quả",
    "Ghi chú",
]


def _today_dmy() -> str:
    return date.today().strftime("%d/%m/%Y")


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).upper().strip()


def find_first_folder(sync: Path) -> Path | None:
    """Find sync/<first> (case-insensitive exact name). Prefer this for refill."""
    if not sync.exists():
        return None
    try:
        children = list(sync.iterdir())
    except OSError:
        return None
    wanted = {n.lower() for n in FIRST_FOLDER_NAMES}
    for p in children:
        try:
            if p.is_dir() and p.name.lower() in wanted:
                return p
        except OSError:
            continue
    return None


def find_priority_folder(sync: Path, hints: tuple[str, ...] = DEFAULT_FOLDER_HINTS) -> Path | None:
    """Prefer folder `first`, else Bình Tây 165-case under sync root."""
    first = find_first_folder(sync)
    if first is not None:
        return first
    if not sync.exists():
        return None
    best: Path | None = None
    best_score = 0
    try:
        children = list(sync.iterdir())
    except OSError:
        return None
    for p in children:
        try:
            if not p.is_dir():
                continue
        except OSError:
            continue
        name = _fold(p.name)
        score = sum(1 for h in hints if h in name)
        if score > best_score:
            best_score = score
            best = p
    if best is not None and best_score >= 2:
        return best
    return None


def _name_is_skipped(ho_ten: str, file_name: str, skip_folds: set[str]) -> bool:
    if not skip_folds:
        return False
    for cand in (ho_ten, file_name):
        f = _fold(cand)
        if not f:
            continue
        if f in skip_folds:
            return True
        # filename often: "... - TRAN DUY NHAT - 1990 - M.pdf"
        for sk in skip_folds:
            if sk and sk in f:
                return True
    return False


def list_pdfs_rglob(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    try:
        return sorted(folder.rglob("*.pdf"), key=lambda p: p.name.lower())
    except OSError:
        return []


def _norm_aid(aid: str) -> str:
    a = (aid or "").strip()
    if a in {ACCOUNT_TK1, "pkdkthuankieu"}:
        return ACCOUNT_TK1
    if a in {ACCOUNT_TK2, "pkdk_Thuankieu"}:
        return ACCOUNT_TK2
    return a


def _payload_lab_keys(payload: dict) -> list[str]:
    return sorted(k for k in payload if k in LAB_TO_FORM.values())


def refill_one(
    pdf: Path,
    *,
    folder_label: str,
    index: dict,
    accounts: list[dict],
    tokens: dict[str, str],
    apply: bool,
    skip_filled: bool = False,
    skip_folds: set[str] | None = None,
) -> dict[str, Any]:
    row: dict[str, Any] = {
        "Tên file": pdf.name,
        "Họ tên": "",
        "Năm sinh": "",
        "Folder nguồn": folder_label,
        "Phạm vi TTHC": "Không có",
        "Trường trên PDF": "",
        "Thiếu trước": "",
        "Đã điền": "",
        "Kết quả": "Dry-run" if not apply else "Thành công",
        "Ghi chú": "",
    }
    try:
        data = extract_pdf(pdf)
    except Exception as e:
        row["Kết quả"] = "Lỗi"
        row["Ghi chú"] = f"parse:{e}"[:120]
        return row

    data["file_name"] = pdf.name
    data["source_file"] = str(pdf)
    name, year = resolve_name_year(
        {
            "ho_ten": data.get("ho_ten") or "",
            "nam_sinh": data.get("nam_sinh") or "",
            "file_name": pdf.name,
            "source_file": str(pdf),
        }
    )
    if name:
        data["ho_ten"] = name
    if year:
        data["nam_sinh"] = year
    row["Họ tên"] = str(data.get("ho_ten") or "")
    row["Năm sinh"] = str(data.get("nam_sinh") or "")

    if skip_folds and _name_is_skipped(row["Họ tên"], pdf.name, skip_folds):
        row["Kết quả"] = "Bỏ qua"
        row["Ghi chú"] = "skip_name"
        return row

    if not data.get("parse_ok"):
        row["Kết quả"] = "Lỗi"
        row["Ghi chú"] = "parse_ok=false"
        return row

    tthc = resolve_tthc_matches(data, index, accounts=accounts)
    if tthc.status == "AMBIGUOUS_NAME":
        row["Kết quả"] = "Bỏ qua"
        row["Ghi chú"] = f"ambiguous:{tthc.mode}"
        return row
    if tthc.status != "READY_IMPORT" or not tthc.matches:
        row["Kết quả"] = "Bỏ qua"
        row["Ghi chú"] = f"no_tthc:{tthc.mode}"
        return row

    by_aid: dict[str, dict] = {}
    for rec in tthc.matches:
        aid = _norm_aid(str(rec.get("_medinet_account") or ""))
        if aid:
            by_aid[aid] = rec
    if ACCOUNT_TK1 in by_aid and ACCOUNT_TK2 in by_aid:
        row["Phạm vi TTHC"] = "Cả 2 TK"
    elif ACCOUNT_TK2 in by_aid:
        row["Phạm vi TTHC"] = "TK2"
    elif ACCOUNT_TK1 in by_aid:
        row["Phạm vi TTHC"] = "TK1"

    def make_reauth(aid: str):
        def _r():
            for acct in accounts:
                if acct["id"] == aid:
                    tokens[aid] = authenticate(acct["user"], acct["password"])
                    return tokens[aid]
            return tokens.get(aid) or ""

        return _r

    all_pdf_fields: list[str] = []
    all_missing: list[str] = []
    all_filled: list[str] = []
    notes: list[str] = []

    for aid, mrec in by_aid.items():
        pid = str(mrec.get("phieukhamId") or mrec.get("Id") or "")
        cdid = mrec.get("cdId")
        if not pid:
            notes.append(f"{aid}:no_pid")
            continue
        payload = labs_to_form_payload(
            data.get("labs") or {},
            phieukham_id=pid,
            gioi_tinh=data.get("gioi_tinh") or "",
        )
        payload["LoaiKham"] = 5152
        if cdid not in (None, ""):
            payload["cdId"] = int(cdid)
        pdf_fields = _payload_lab_keys(payload)
        all_pdf_fields.extend(f"{aid}:{k}" for k in pdf_fields)

        # Always compare web vs FULL PDF payload (every field PDF has).
        pdf_has_urea = "Urea" in (data.get("labs") or {})

        def _miss_wo_optional_urea(existing_row: dict | None) -> list[str]:
            if not cls_has_lab_values(existing_row):
                miss0 = list(pdf_fields)
            else:
                miss0 = cls_missing_lab_fields(existing_row, payload)
            if not pdf_has_urea:
                miss0 = [k for k in miss0 if k != "SinhHoaMau_Ure"]
            return miss0

        if not apply:
            try:
                existing, tokens[aid] = load_cls_view(
                    tokens[aid], pid, reauth=make_reauth(aid)
                )
                miss = _miss_wo_optional_urea(existing)
                all_missing.extend(f"{aid}:{k}" for k in miss)
                all_filled.extend(f"{aid}:{k}" for k in miss)  # would fill
            except Exception as e:
                notes.append(f"{aid}:dry_load:{e}"[:40])
                all_missing.extend(f"{aid}:{k}" for k in pdf_fields)
            continue

        # APPLY
        existing, tokens[aid] = load_cls_view(tokens[aid], pid, reauth=make_reauth(aid))
        miss = _miss_wo_optional_urea(existing)
        all_missing.extend(f"{aid}:{k}" for k in miss)

        # skip_filled: web already has every PDF field → do not Set again
        if skip_filled and pdf_fields and not miss:
            all_filled.extend(f"{aid}:{k}" for k in pdf_fields)
            notes.append(f"{aid}:da_du_skip")
            continue

        still: list[str] = list(pdf_fields)
        ok, msg, verified, vdetail = False, "", False, ""
        max_set = 3
        for attempt in range(max_set):
            ok, msg, _raw, tokens[aid] = insert_cls(
                tokens[aid], payload, reauth=make_reauth(aid)
            )
            time.sleep(0.12 * (attempt + 1))
            verified, vdetail, tokens[aid] = verify_cls_saved(
                tokens[aid], pid, payload=payload, reauth=make_reauth(aid)
            )
            existing2, tokens[aid] = load_cls_view(
                tokens[aid], pid, reauth=make_reauth(aid)
            )
            still = _miss_wo_optional_urea(existing2)
            if not still:
                break
            # Partial persist — retry full Set (urine format branches inside insert_cls)
            notes.append(
                f"{aid}:retry_set={attempt + 1}/{max_set};con_thieu={still[:8]}"
            )

        filled = [k for k in pdf_fields if k not in still]
        if not pdf_has_urea:
            filled = [k for k in filled if k != "SinhHoaMau_Ure"]
        all_filled.extend(f"{aid}:{k}" for k in filled)
        notes.append(
            f"{aid}:ok={ok};ver={verified};miss_truoc={len(miss)};{vdetail}"[:120]
        )
        if still:
            row["Kết quả"] = "Một phần"
            notes.append(f"{aid}:con_thieu={still[:12]}")
        elif not verified:
            # Get+FormViewer empty / mismatch — never false Thành công
            row["Kết quả"] = "Một phần"
            notes.append(f"{aid}:verify_fail:{vdetail}"[:80])
        elif not ok:
            row["Kết quả"] = "Một phần"
            notes.append(f"{aid}:set_fail:{msg}"[:80])

    row["Trường trên PDF"] = ", ".join(sorted(set(all_pdf_fields)))[:500]
    row["Thiếu trước"] = ", ".join(sorted(set(all_missing)))[:500]
    row["Đã điền"] = ", ".join(sorted(set(all_filled)))[:500]
    row["Ghi chú"] = "; ".join(notes)[:300]
    if row["Kết quả"] not in {"Lỗi", "Bỏ qua", "Một phần"}:
        if not all_pdf_fields:
            row["Kết quả"] = "Bỏ qua"
            row["Ghi chú"] = (row["Ghi chú"] + ";khong_co_truong_pdf")[:300]
        elif apply and skip_filled and all_pdf_fields and not all_missing:
            # Had TTHC + web already complete — skipped Set
            row["Kết quả"] = "Đã đủ"
        elif apply:
            row["Kết quả"] = "Thành công"
        else:
            row["Kết quả"] = "Dry-run"
    return row


def write_refill_excel(rows: list[dict], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Chi tiết hành động"
    ws.append(ACTION_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws.append([r.get(c, "") for c in ACTION_COLS])

    ws_all = wb.create_sheet("Tất cả")
    ws_all.append(ACTION_COLS)
    for c in ws_all[1]:
        c.font = Font(bold=True)
    for r in rows:
        ws_all.append([r.get(c, "") for c in ACTION_COLS])

    need = [r for r in rows if r.get("Thiếu trước")]
    ws_need = wb.create_sheet("Cần điền")
    ws_need.append(ACTION_COLS)
    for c in ws_need[1]:
        c.font = Font(bold=True)
    for r in need:
        ws_need.append([r.get(c, "") for c in ACTION_COLS])

    skip = [r for r in rows if r.get("Kết quả") == "Bỏ qua"]
    ws_skip = wb.create_sheet("Khong TTHC - Bo qua")
    ws_skip.append(ACTION_COLS)
    for c in ws_skip[1]:
        c.font = Font(bold=True)
    for r in skip:
        ws_skip.append([r.get(c, "") for c in ACTION_COLS])

    ws_sum = wb.create_sheet("Tóm tắt")
    ws_sum.append(["Chỉ số", "Giá trị"])
    ws_sum["A1"].font = Font(bold=True)
    ws_sum.append(["Tổng PDF", len(rows)])
    ws_sum.append(["Cần điền", len(need)])
    ws_sum.append(["Bỏ qua", len(skip)])
    ws_sum.append([])
    ws_sum.append(["Kết quả", "Số lượng"])
    for k, v in Counter(str(r.get("Kết quả") or "") for r in rows).most_common():
        ws_sum.append([k, v])
    wb.save(out_path)
    return out_path


def run_refill(
    *,
    apply: bool = False,
    toan_bo: bool = False,
    folder: str = "",
    limit: int = 0,
    resume: bool = False,
    lock_name: str = "refill_cls_inplace",
    skip_filled: bool = False,
    refresh_index: bool = False,
    skip_names: list[str] | None = None,
) -> dict:
    lock = acquire_lock(lock_name or "refill_cls_inplace")
    if lock is None:
        safe_print(f"ABORT: refill dang chay (lock={lock_name or 'refill_cls_inplace'}).")
        return {"abort": "locked"}
    try:
        if sys.platform.startswith("win") and g_pipeline_live() is None:
            safe_print("ABORT: G: chua mount.")
            return {"abort": "g_missing"}
        cfg = load_config()
        sync = resolve_g_sync(cfg)
        if sys.platform.startswith("win") and not require_g_on_windows(sync):
            safe_print(f"ABORT: chi G: sync={sync}")
            return {"abort": "not_g"}

        targets: list[tuple[str, Path]] = []
        if folder.strip():
            p = Path(folder.strip())
            if not p.is_absolute():
                p = sync / folder.strip()
            targets.append((p.name, p))
        elif toan_bo:
            for name in TOAN_BO_FOLDERS:
                targets.append((name, sync / name))
        else:
            found = find_priority_folder(sync)
            if found is None:
                safe_print(
                    "ABORT: khong tim thay folder 'first' hoac Binh Tay / 165 CASE"
                )
                safe_print(f"SYNC={sync}")
                try:
                    for c in sorted(sync.iterdir()):
                        if c.is_dir():
                            safe_print(f"  - {c.name}")
                except OSError:
                    pass
                return {"abort": "folder_not_found"}
            targets.append((found.name, found))
            if found.name.lower() == "first":
                safe_print("Uu tien: folder first (khong move)")

        mode = "APPLY" if apply else "DRY-RUN"
        safe_print(f"========== DIEN LAI CLS ({mode}) ==========")
        safe_print(f"SYNC: {sync}")
        safe_print("Rule: FULL PDF fields -> web (trong/ngoai khoang), KHONG MOVE, KHONG Excel")
        if skip_filled:
            safe_print("skip_filled=ON: web da du -> bo qua Set; Bo qua/no_TTHC se thu lai")

        skip_folds = {_fold(n) for n in DEFAULT_SKIP_NAMES}
        for n in skip_names or []:
            f = _fold(n)
            if f:
                skip_folds.add(f)
        if skip_folds:
            safe_print(f"skip_names: {sorted(skip_folds)}")

        pdfs: list[tuple[str, Path]] = []
        for label, d in targets:
            found_pdfs = list_pdfs_rglob(d)
            safe_print(f"  {label}: {len(found_pdfs)} pdf")
            for p in found_pdfs:
                pdfs.append((label, p))
        if limit > 0:
            pdfs = pdfs[:limit]
            safe_print(f"Limited: {len(pdfs)}")

        accounts = [dict(a) for a in MEDINET_ACCOUNTS[:2]]
        tokens: dict[str, str] = {}
        for acct in accounts:
            tokens[acct["id"]] = authenticate(acct["user"], acct["password"])

        date_from = (cfg.get("medinet") or {}).get("date_from") or "01/07/2026"
        date_to = ((cfg.get("medinet") or {}).get("date_to") or "").strip() or _today_dmy()
        cache_dir = ROOT / "pipeline" / "work" / "index_cache"
        # refresh_index=0h: bat buoc fetch lai TTHC (case moi nhap sau lan chay truoc)
        idx_age = 0.0 if refresh_index else 3.0
        if refresh_index:
            safe_print("refresh_index=ON: fetch TTHC moi (khong dung cache 3h)")
        index = load_or_fetch_merged_unit_index(
            accounts, date_from, date_to, cache_dir=cache_dir, max_age_hours=idx_age
        )

        results: list[dict] = []
        t0 = time.time()
        build = local_work_build()
        build.mkdir(parents=True, exist_ok=True)
        (build / "logs").mkdir(parents=True, exist_ok=True)
        tag = "TOANBO" if toan_bo else (
            "FIRST"
            if folder.strip().lower() == "first"
            else (
                "PDF"
                if folder.strip().lower() == "pdf"
                else "FOLDER"
            )
        )
        ck_path = build / f"REFILL_CHECKPOINT_{tag}.txt"
        # Only skip these on --resume (NOT "Bỏ qua" — those must retry when TTHC appears)
        _OK_STATUSES = {"Thành công", "Đã đủ"}
        done_ok: set[str] = set()
        if resume and ck_path.exists():
            try:
                for ln in ck_path.read_text(encoding="utf-8").splitlines():
                    ln = ln.strip()
                    if not ln or ln.startswith("#"):
                        continue
                    if "\t" in ln:
                        path_s, st = ln.split("\t", 1)
                        if st.strip() in _OK_STATUSES:
                            done_ok.add(path_s.strip())
                    else:
                        # Legacy bare path = treated as OK (crash-continue)
                        done_ok.add(ln)
                safe_print(
                    f"RESUME: skip {len(done_ok)} pdf da Thanh cong/Da du ({ck_path.name})"
                )
            except OSError as e:
                safe_print(f"RESUME: khong doc duoc checkpoint: {e}")

        # Fresh run without --resume: start new checkpoint
        if not resume:
            try:
                ck_path.write_text(
                    f"# refill checkpoint {tag} {datetime.now().isoformat()}\n"
                    f"# format: path<TAB>status  (chi Thanh cong/Da du de --resume)\n",
                    encoding="utf-8",
                )
            except OSError:
                pass

        skipped = 0
        try:
            for i, (label, pdf) in enumerate(pdfs, 1):
                key = str(pdf.resolve()) if pdf.exists() else str(pdf)
                if resume and key in done_ok:
                    skipped += 1
                    if i == 1 or i % 100 == 0 or i == len(pdfs):
                        safe_print(f"  [{i}/{len(pdfs)}] SKIP (resume OK) {pdf.name}")
                    continue
                # Quick name skip from filename before parse (Trần Duy Nhất, …)
                if _name_is_skipped("", pdf.name, skip_folds):
                    r = {
                        "Tên file": pdf.name,
                        "Họ tên": "",
                        "Năm sinh": "",
                        "Folder nguồn": label,
                        "Phạm vi TTHC": "Không có",
                        "Trường trên PDF": "",
                        "Thiếu trước": "",
                        "Đã điền": "",
                        "Kết quả": "Bỏ qua",
                        "Ghi chú": "skip_name",
                    }
                    results.append(r)
                    if i == 1 or i % 25 == 0 or i == len(pdfs):
                        safe_print(f"  [{i}/{len(pdfs)}] SKIP name {pdf.name}")
                    continue
                try:
                    r = refill_one(
                        pdf,
                        folder_label=label,
                        index=index,
                        accounts=accounts,
                        tokens=tokens,
                        apply=apply,
                        skip_filled=skip_filled,
                        skip_folds=skip_folds,
                    )
                except Exception as e:
                    r = {
                        "Tên file": pdf.name,
                        "Họ tên": "",
                        "Năm sinh": "",
                        "Folder nguồn": label,
                        "Phạm vi TTHC": "",
                        "Trường trên PDF": "",
                        "Thiếu trước": "",
                        "Đã điền": "",
                        "Kết quả": "Lỗi",
                        "Ghi chú": f"crash:{e}"[:200],
                    }
                    safe_print(f"  [{i}/{len(pdfs)}] CRASH {pdf.name}: {e}")
                results.append(r)
                ket = str(r.get("Kết quả") or "")
                # Checkpoint only OK — "Bỏ qua" (no TTHC) must be retryable next run
                if ket in _OK_STATUSES:
                    try:
                        with ck_path.open("a", encoding="utf-8") as fh:
                            fh.write(f"{key}\t{ket}\n")
                        done_ok.add(key)
                    except OSError:
                        pass
                if i == 1 or i % 25 == 0 or i == len(pdfs):
                    safe_print(
                        f"  [{i}/{len(pdfs)}] {r.get('Họ tên') or pdf.name} "
                        f"scope={r.get('Phạm vi TTHC')} ketqua={ket}"
                    )
                    try:
                        sys.stdout.flush()
                    except Exception:
                        pass
        except KeyboardInterrupt:
            safe_print(
                f"STOPPED KeyboardInterrupt tai {len(results)} xu ly "
                f"(skipped_resume={skipped}). Chay lai voi --resume."
            )
            _write_refill_progress_log(
                build, tag, mode, toan_bo, results, t0, skipped, interrupted=True
            )
            return {
                "abort": "interrupted",
                "total": len(results),
                "skipped": skipped,
            }
        except Exception as e:
            safe_print(f"STOPPED exception: {e} (xu ly={len(results)}; --resume de tiep)")
            _write_refill_progress_log(
                build, tag, mode, toan_bo, results, t0, skipped, interrupted=True
            )
            return {"abort": f"crash:{e}", "total": len(results), "skipped": skipped}

        # No Excel this run (hourly later). Short console + txt log only.
        out_log = _write_refill_progress_log(
            build, tag, mode, toan_bo, results, t0, skipped, interrupted=False
        )
        counts = Counter(str(r.get("Kết quả") or "") for r in results)
        for k, v in counts.most_common():
            safe_print(f"  {k}={v}")
        if skipped:
            safe_print(f"  resume_skipped={skipped}")
        safe_print(f"DONE ({mode}) - khong move PDF, khong Excel")
        return {
            "ok": True,
            "log": str(out_log),
            "total": len(results),
            "skipped": skipped,
            "counts": dict(counts),
        }
    finally:
        release_lock(lock)


def _write_refill_progress_log(
    build: Path,
    tag: str,
    mode: str,
    toan_bo: bool,
    results: list[dict],
    t0: float,
    skipped: int,
    *,
    interrupted: bool,
) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    suffix = "PARTIAL" if interrupted else tag
    out_log = build / "logs" / f"REFILL_{suffix}_{stamp}.txt"
    counts = Counter(str(r.get("Kết quả") or "") for r in results)
    lines = [
        f"mode={mode}",
        f"toan_bo={toan_bo}",
        f"interrupted={interrupted}",
        f"total_processed={len(results)}",
        f"resume_skipped={skipped}",
        f"elapsed_s={time.time() - t0:.0f}",
        "excel=SKIP",
        "",
        "ket_qua:",
    ]
    for k, v in counts.most_common():
        lines.append(f"  {k}={v}")
    skips = [r for r in results if r.get("Kết quả") == "Bỏ qua"]
    if skips:
        lines.append("")
        lines.append(f"bo_qua ({len(skips)}):")
        for r in skips:
            lines.append(
                f"  {r.get('Họ tên') or '?'}|{r.get('Tên file')}|ly_do={r.get('Ghi chú')}"
            )
    partials = [r for r in results if r.get("Kết quả") == "Một phần"][:20]
    if partials:
        lines.append("")
        lines.append("mau_mot_phan:")
        for r in partials:
            lines.append(
                f"  {r.get('Họ tên')}|{r.get('Tên file')}|thieu={r.get('Thiếu trước')}"
            )
    if interrupted:
        lines.append("")
        lines.append("NOTE: chay lai voi --resume de bo qua PDF da checkpoint.")
    try:
        out_log.write_text("\n".join(lines) + "\n", encoding="utf-8")
        safe_print(f"Log: {out_log}")
    except OSError as e:
        safe_print(f"Log write failed: {e}")
    # Always echo skipped names to console
    if skips:
        safe_print(f"BO QUA chi tiet ({len(skips)}):")
        for r in skips:
            safe_print(
                f"  - {r.get('Họ tên') or '?'} | {r.get('Tên file')} | {r.get('Ghi chú')}"
            )
    return out_log


def main() -> int:
    ap = argparse.ArgumentParser(description="Dien lai CLS tu PDF — khong move")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--toan-bo", action="store_true", help="Quet PROCESSED/TK1/TK2/...")
    ap.add_argument("--folder", default="", help="Path hoac ten folder duoi sync")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument(
        "--resume",
        action="store_true",
        help="Bo qua PDF da co trong checkpoint (tiep tuc sau crash)",
    )
    ap.add_argument(
        "--lock-name",
        default="refill_cls_inplace",
        help="Ten lock rieng (vd refill_cls_pdf) de khong ngat refill khac",
    )
    ap.add_argument(
        "--skip-filled",
        action="store_true",
        help="Web da du field PDF -> bo qua Set; case Bo qua/no_TTHC van thu lai",
    )
    ap.add_argument(
        "--refresh-index",
        action="store_true",
        help="Fetch TTHC moi (khong dung cache 3h) — can khi vua nhap them benh nhan",
    )
    ap.add_argument(
        "--skip-name",
        action="append",
        default=[],
        help="Bo qua ten (co the lap lai). Mac dinh luon skip TRAN DUY NHAT",
    )
    args = ap.parse_args()
    extra_skips = [str(x) for x in (args.skip_name or []) if str(x).strip()]
    res = run_refill(
        apply=bool(args.apply),
        toan_bo=bool(args.toan_bo),
        folder=str(args.folder or ""),
        limit=int(args.limit or 0),
        resume=bool(args.resume),
        lock_name=str(args.lock_name or "refill_cls_inplace"),
        skip_filled=bool(args.skip_filled),
        refresh_index=bool(args.refresh_index),
        skip_names=extra_skips or None,
    )
    if res.get("abort") == "interrupted":
        return 3
    if res.get("abort"):
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
